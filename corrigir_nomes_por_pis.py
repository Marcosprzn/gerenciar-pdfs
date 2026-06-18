#!/usr/bin/env python3
import os, re, math, datetime
import tkinter as tk
from tkinter import filedialog

import xlrd
import openpyxl
from odf import opendocument
from odf.table import Table, TableRow, TableCell
from odf.text import P
import pandas as pd

SHEET_NAME = 'FGTS EM ATRASO - PROCESSOS'

IGNORAR_NOMES = {
    'NOMES', 'NOME', 'TITULAR', 'FUNCIONARIO',
    'PIS', 'PROC.', 'DATA', 'OBS',
    'TOTAL', 'TOTAIS', 'SUBTOTAL',
}

def normalizar_pis(valor):
    if valor is None:
        return ''
    if isinstance(valor, float):
        if math.isnan(valor) or math.isinf(valor) or valor == 0:
            return ''
        return str(int(valor)).zfill(11)
    texto = str(valor).strip()
    if not texto or texto in ('0', '0.0'):
        return ''
    digitos = re.sub(r'\D', '', texto)
    return digitos.zfill(11)

def eh_nome_valido(nome):
    if not isinstance(nome, str):
        return False
    nome = nome.strip().upper()
    if not nome or len(nome) <= 5 or ' ' not in nome:
        return False
    if nome in IGNORAR_NOMES:
        return False
    return bool(re.match(r"^[A-ZÁÀÂÃÉÊÍÓÔÕÚÜÇ0-9 \.\-]+$", nome))

def ler_mapa_referencia_xls(caminho):
    wb = xlrd.open_workbook(caminho)
    if SHEET_NAME not in wb.sheet_names():
        return {}
    sheet = wb.sheet_by_name(SHEET_NAME)
    mapa = {}
    for row in range(sheet.nrows):
        nome_raw = sheet.cell_value(row, 3)
        if not eh_nome_valido(nome_raw):
            continue
        nome = nome_raw.strip().upper()
        pis = normalizar_pis(sheet.cell_value(row, 2))
        if not pis:
            pis = normalizar_pis(sheet.cell_value(row, 1))
        if pis and pis not in mapa:
            mapa[pis] = nome
    return mapa

def ler_mapa_referencia_xlsx(caminho):
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        return {}
    sheet = wb[SHEET_NAME]
    mapa = {}
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
        nome_raw = row[3] if len(row) > 3 else None
        if not eh_nome_valido(nome_raw):
            continue
        nome = str(nome_raw).strip().upper()
        pis = normalizar_pis(row[2] if len(row) > 2 else None)
        if not pis:
            pis = normalizar_pis(row[1] if len(row) > 1 else None)
        if pis and pis not in mapa:
            mapa[pis] = nome
    wb.close()
    return mapa

def construir_mapa_referencia(caminho):
    ext = os.path.splitext(caminho)[1].lower()
    if ext == '.xlsx':
        return ler_mapa_referencia_xlsx(caminho)
    if ext == '.ods':
        return ler_mapa_referencia_ods(caminho)
    return ler_mapa_referencia_xls(caminho)

def localizar_correcoes(caminho, mapa_ref):
    ext = os.path.splitext(caminho)[1].lower()
    if ext == '.xlsx':
        return localizar_correcoes_xlsx(caminho, mapa_ref)
    if ext == '.ods':
        return localizar_correcoes_ods(caminho, mapa_ref)
    return localizar_correcoes_xls(caminho, mapa_ref)

def localizar_correcoes_xls(caminho, mapa_ref):
    wb = xlrd.open_workbook(caminho)
    if SHEET_NAME not in wb.sheet_names():
        print(f'    Aba "{SHEET_NAME}" nao encontrada.')
        return []
    sheet = wb.sheet_by_name(SHEET_NAME)
    correcoes = []
    for row in range(sheet.nrows):
        nome_raw = sheet.cell_value(row, 3)
        if not eh_nome_valido(nome_raw):
            continue
        nome_atual = nome_raw.strip().upper()
        pis = normalizar_pis(sheet.cell_value(row, 2))
        if not pis:
            pis = normalizar_pis(sheet.cell_value(row, 1))
        if not pis:
            continue
        nome_correto = mapa_ref.get(pis)
        if nome_correto and nome_correto != nome_atual:
            correcoes.append((row, pis, nome_atual, nome_correto))
    return correcoes

def localizar_correcoes_xlsx(caminho, mapa_ref):
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f'    Aba "{SHEET_NAME}" nao encontrada.')
        wb.close()
        return []
    sheet = wb[SHEET_NAME]
    correcoes = []
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
        nome_raw = row[3] if len(row) > 3 else None
        if not eh_nome_valido(nome_raw):
            continue
        nome_atual = str(nome_raw).strip().upper()
        pis = normalizar_pis(row[2] if len(row) > 2 else None)
        if not pis:
            pis = normalizar_pis(row[1] if len(row) > 1 else None)
        if not pis:
            continue
        nome_correto = mapa_ref.get(pis)
        if nome_correto and nome_correto != nome_atual:
            correcoes.append((i, pis, nome_atual, nome_correto))
    wb.close()
    return correcoes

def aplicar_correcoes_xlsx(caminho, correcoes):
    wb = openpyxl.load_workbook(caminho)
    ws = wb[SHEET_NAME]
    for idx, pis, antigo, novo in correcoes:
        ws.cell(row=idx, column=4).value = novo
    wb.save(caminho)
    wb.close()

def ler_df_ods(caminho):
    return pd.read_excel(caminho, sheet_name=SHEET_NAME, header=None, engine='odf')

def ler_mapa_referencia_ods(caminho):
    df = ler_df_ods(caminho)
    mapa = {}
    for i in range(len(df)):
        nome_raw = df.iloc[i, 3]
        if not eh_nome_valido(nome_raw):
            continue
        nome = str(nome_raw).strip().upper()
        pis = normalizar_pis(df.iloc[i, 2])
        if not pis:
            pis = normalizar_pis(df.iloc[i, 1])
        if pis and pis not in mapa:
            mapa[pis] = nome
    return mapa

def localizar_correcoes_ods(caminho, mapa_ref):
    df = ler_df_ods(caminho)
    correcoes = []
    for i in range(len(df)):
        nome_raw = df.iloc[i, 3]
        if not eh_nome_valido(nome_raw):
            continue
        nome_atual = str(nome_raw).strip().upper()
        pis = normalizar_pis(df.iloc[i, 2])
        if not pis:
            pis = normalizar_pis(df.iloc[i, 1])
        if not pis:
            continue
        nome_correto = mapa_ref.get(pis)
        if nome_correto and nome_correto != nome_atual:
            correcoes.append((i + 1, pis, nome_atual, nome_correto))
    return correcoes

def _ods_set_cell_value(rows, row_idx, col_idx, value):
    if row_idx >= len(rows):
        return
    cells = rows[row_idx].getElementsByType(TableCell)
    while len(cells) <= col_idx:
        new_cell = TableCell()
        rows[row_idx].addElement(new_cell)
        cells = rows[row_idx].getElementsByType(TableCell)
    cell = cells[col_idx]
    for old_p in cell.getElementsByType(P):
        cell.removeChild(old_p)
    cell.setAttribute('valuetype', 'string')
    p = P(text=value)
    cell.addElement(p)

def aplicar_correcoes_ods(caminho, correcoes):
    doc = opendocument.load(caminho)
    tables = doc.spreadsheet.getElementsByType(Table)
    table = None
    for t in tables:
        if t.getAttribute('name') == SHEET_NAME:
            table = t
            break
    if table is None:
        print(f'    Aba "{SHEET_NAME}" nao encontrada.')
        return
    rows = table.getElementsByType(TableRow)
    for idx, pis, antigo, novo in correcoes:
        _ods_set_cell_value(rows, idx - 1, 3, novo)
    doc.save(caminho)

def aplicar_correcoes_xls_com_excel(caminho, correcoes):
    import win32com.client
    import pythoncom
    pythoncom.CoInitialize()
    xl = win32com.client.Dispatch('Excel.Application')
    try:
        xl.Application.DisplayAlerts = False
        xl.Visible = False
        abs_path = os.path.abspath(caminho)
        wb = xl.Workbooks.Open(abs_path)
        ws = wb.Sheets(SHEET_NAME)
        for idx, pis, antigo, novo in correcoes:
            ws.Cells(idx + 1, 4).Value = novo
        wb.Close(SaveChanges=True)
    except Exception:
        import traceback
        traceback.print_exc()
        raise
    finally:
        try:
            xl.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()

def dialogo_arquivo(titulo, tipos):
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    root.update()
    arquivo = filedialog.askopenfilename(title=titulo, filetypes=tipos)
    root.destroy()
    return arquivo

def dialogo_arquivos(titulo, tipos):
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    root.update()
    arquivos = filedialog.askopenfilenames(title=titulo, filetypes=tipos)
    root.destroy()
    return list(arquivos)

def main():
    print('=' * 60)
    print('  CORRECAO DE NOMES POR PIS')
    print('  (.xls via Excel | .xlsx via openpyxl | .ods via odf)')
    print('=' * 60)
    print()
    print('Selecione a planilha REFERENCIA (ja corrigida manualmente)')
    print()
    ref = dialogo_arquivo(
        'Selecione a planilha REFERENCIA (corrigida)',
        [('Planilhas', '*.xls *.xlsx *.ods')]
    )
    if not ref:
        print('Nenhuma referencia selecionada.')
        return

    print(f'Referencia: {os.path.basename(ref)}')
    print(f'\nLendo referencia...')
    mapa = construir_mapa_referencia(ref)
    if not mapa:
        print('Nenhum funcionario encontrado na referencia.')
        print(f'Verifique se a aba "{SHEET_NAME}" existe e tem dados.')
        input('\nPressione Enter para sair...')
        return
    print(f'  {len(mapa)} PIS unicos lidos da referencia')
    print()

    alvos = dialogo_arquivos(
        'Selecione as planilhas para CORRIGIR',
        [('Planilhas', '*.xls *.xlsx *.ods')]
    )
    if not alvos:
        print('Nenhuma planilha selecionada.')
        return
    alvos = [a for a in alvos if os.path.normpath(a) != os.path.normpath(ref)]
    if not alvos:
        print('Nenhuma planilha para corrigir (excluindo a referencia).')
        return

    print(f'\n{len(alvos)} planilha(s) selecionada(s):')
    for a in alvos:
        print(f'  - {os.path.basename(a)}  ({os.path.splitext(a)[1]})')

    print(f'\nATENCAO: Os arquivos serao modificados IN-PLACE!')
    resp = input('Continuar? (S/n): ').strip().lower()
    if resp == 'n':
        print('Operacao cancelada.')
        return

    log = []
    log.append(f'Correcao por PIS - {datetime.datetime.now().strftime("%d/%m/%Y %H:%M")}')
    log.append(f'Referencia: {ref}')
    log.append('')
    log.append('=' * 70)
    log.append('')

    total_geral = 0
    for alvo in alvos:
        nome = os.path.basename(alvo)
        ext = os.path.splitext(alvo)[1].lower()
        print(f'\n--- {nome} ---')
        try:
            correcoes = localizar_correcoes(alvo, mapa)
            if not correcoes:
                print(f'  Nenhuma correcao necessaria.')
                log.append(f'{nome}: 0 correcoes')
                continue

            for idx, pis, antigo, novo in correcoes:
                print(f'    {antigo:45s} -> {novo}  (PIS: {pis})')
                log.append(f'  L{idx+1}: {antigo} -> {novo}')

            if ext == '.xlsx':
                print(f'  Aplicando correcoes no .xlsx...')
                aplicar_correcoes_xlsx(alvo, correcoes)
            elif ext == '.ods':
                print(f'  Aplicando correcoes no .ods...')
                aplicar_correcoes_ods(alvo, correcoes)
            else:
                print(f'  Aplicando correcoes via Excel...')
                aplicar_correcoes_xls_com_excel(alvo, correcoes)

            print(f'  {len(correcoes)} nome(s) corrigido(s)')
            log.append(f'{nome}: {len(correcoes)} correcao(oes)')
            total_geral += len(correcoes)
        except Exception as e:
            msg = f'  ERRO: {e}'
            print(msg)
            log.append(f'{nome}: ERRO - {e}')
            import traceback
            traceback.print_exc()

    log.append('')
    log.append('=' * 70)
    log.append(f'Total corrigidos: {total_geral}')

    log_dir = os.path.dirname(alvos[0]) if alvos else '.'
    log_path = os.path.join(log_dir, f'relatorio_correcoes_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.txt')
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(log))

    print(f'\n{"=" * 60}')
    print(f'  CONCLUIDO!')
    print(f'  Total de nomes corrigidos: {total_geral}')
    print(f'  Relatorio: {log_path}')
    print(f'{"=" * 60}')
    input('\nPressione Enter para sair...')

if __name__ == '__main__':
    main()
