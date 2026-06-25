#!/usr/bin/env python3
"""
Consolida relatorios .xlsx em uma planilha formatada com duas abas:
- Falta PDF: pessoas na planilha sem PDF correspondente
- Falta no Excel: PDFs na pasta sem registro na planilha

Escaneia automaticamente a estrutura:
pasta_raiz/ -> ANO/ -> ANO CONFERIDOS/ -> relatorio.xlsx
"""
import os, sys, re
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

root = tk.Tk()
root.withdraw()
root.attributes('-topmost', True)
root.update()

pasta_raiz = filedialog.askdirectory(
    title="Selecione a PASTA RAIZ (que contem as pastas 2007, 2008, 2009...)"
)
root.destroy()

if not pasta_raiz:
    print("Nenhuma pasta selecionada.")
    sys.exit(1)

# Escaneia estrutura: ANO/ -> ANO CONFERIDOS/ -> .xlsx
arquivos = []
for item in sorted(os.listdir(pasta_raiz)):
    caminho_ano = os.path.join(pasta_raiz, item)
    if not os.path.isdir(caminho_ano):
        continue
    # Procura subpastas com "CONFERIDOS" no nome
    for sub in sorted(os.listdir(caminho_ano)):
        caminho_sub = os.path.join(caminho_ano, sub)
        if not os.path.isdir(caminho_sub):
            continue
        if 'CONFERIDOS' in sub.upper() or 'CONFERIDO' in sub.upper():
            for f in sorted(os.listdir(caminho_sub)):
                if f.lower().endswith('.xlsx'):
                    arquivos.append(os.path.join(caminho_sub, f))

if not arquivos:
    print("Nenhum arquivo .xlsx encontrado na estrutura esperada.")
    sys.exit(1)

print(f"Encontrados {len(arquivos)} relatorios:\n")
for a in arquivos:
    print(f"  {os.path.basename(a)}  ({os.path.dirname(a)})")
print()

falta_pdf = []
falta_excel = []

for arq in sorted(arquivos):
    nome = os.path.basename(arq)
    m = re.search(r"(?<!\d)(\d{2})[-_ ](\d{2,4})(?!\d)", nome)
    mes_ano = None
    if m:
        mes = m.group(1)
        ano = "20" + m.group(2) if len(m.group(2)) == 2 else m.group(2)
        mes_ano = f"{mes}/{ano}"

    try:
        df = pd.read_excel(arq, sheet_name='Dados')
    except Exception:
        try:
            df = pd.read_excel(arq)
        except Exception as e:
            print(f"  ERRO ao ler {nome}: {e}")
            continue

    col_status = None
    col_nome = None
    col_pdf = None
    for col in df.columns:
        cs = str(col).lower()
        if 'status' in cs:
            col_status = col
        elif 'nome' in cs:
            col_nome = col
        elif 'arquivo' in cs or 'encontrado' in cs:
            col_pdf = col

    if not col_status:
        print(f"  Coluna nao encontrada em {nome}")
        continue

    for _, row in df.iterrows():
        status = str(row[col_status])
        nome_pessoa = str(row.get(col_nome or 'NOMES', ''))
        nome_pdf = str(row.get(col_pdf or 'Nome do Arquivo Encontrado', ''))

        if 'NÃO ENCONTRADO NO .PDF' in status.upper():
            if nome_pessoa and nome_pessoa not in ('', 'nan', 'None'):
                falta_pdf.append({
                    'mes': mes_ano or nome,
                    'nome': nome_pessoa,
                    'proc': str(row.get('PROC.', str(row.get('Processo', str(row.get('PROC', '')))))),
                    'pis': str(row.get('PIS', '')),
                })
        elif 'PDF NA PASTA' in status.upper():
            if nome_pdf and nome_pdf not in ('', 'nan', 'None'):
                falta_excel.append({
                    'mes': mes_ano or nome,
                    'nome_pdf': nome_pdf,
                    'arquivo': nome,
                })

# Ordena
falta_pdf.sort(key=lambda x: (x['mes'], x['nome']))
falta_excel.sort(key=lambda x: (x['mes'], x['nome_pdf']))

# Cria planilha
from collections import Counter
pessoas_count = Counter(item['nome'] for item in falta_pdf)

caminho_saida = os.path.join(pasta_raiz, 'consolidado_faltantes.xlsx')
with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
    df_falta_pdf = pd.DataFrame(falta_pdf) if falta_pdf else pd.DataFrame(columns=['mes', 'nome', 'proc', 'pis'])
    df_falta_excel = pd.DataFrame(falta_excel) if falta_excel else pd.DataFrame(columns=['mes', 'nome_pdf', 'arquivo'])
    df_falta_pdf.to_excel(writer, sheet_name='Falta PDF', index=False)
    df_falta_excel.to_excel(writer, sheet_name='Falta no Excel', index=False)

# Formata
wb = load_workbook(caminho_saida)

bdr = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
fill_cinza = PatternFill('solid', fgColor='D3D3D3')
font_branca = Font(color='FFFFFF', bold=True, size=11)
fill_azul = PatternFill('solid', fgColor='2F5496')

def formatar_aba(ws, titulo, col_larguras):
    ws.insert_rows(1, 2)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_larguras))
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    for col_idx, largura in enumerate(col_larguras, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = font_branca
        cell.fill = fill_azul
        cell.border = bdr
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = largura

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.border = bdr

ws_falta = wb['Falta PDF']
formatar_aba(ws_falta, 'FALTA PDF - Pessoas na planilha sem PDF correspondente', [12, 50, 20, 20])

ws_excel = wb['Falta no Excel']
formatar_aba(ws_excel, 'FALTA NO EXCEL - PDFs na pasta sem registro na planilha', [12, 50, 30])

wb.save(caminho_saida)

# Resumo no terminal
print(f"\nTotal faltando PDF: {len(falta_pdf)}")
print(f"Total faltando Excel: {len(falta_excel)}")
print(f"\nPessoas que mais faltam PDF:")
print(f"{'Nome':<50} {'Vezes'}")
print("=" * 60)
for nome, count in pessoas_count.most_common(20):
    print(f"{nome:<50} {count}")

print(f"\nPlanilha salva: {caminho_saida}")
os.startfile(caminho_saida)
input("\nPressione Enter para sair...")
