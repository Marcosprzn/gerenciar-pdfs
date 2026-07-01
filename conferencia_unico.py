#!/usr/bin/env python3
import os, re, sys, math, datetime, unicodedata, shutil
import tkinter as tk
from tkinter import filedialog
from difflib import SequenceMatcher

import pandas as pd
import xlrd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    import llm_matcher
except ImportError:
    llm_matcher = None

EXT = ('.pdf', '.tif', '.tiff')
SHEET_NAME = 'FGTS EM ATRASO - PROCESSOS'
EXCLUIR_PADROES = [
    'ARQUIVO SEFIP', 'SEFIP', 'GRRF', 'DEPOSITADO',
    'COMPROVANTE', 'RECIBO', 'EXTRATO', 'FOLHA',
    'RELATORIO', 'GUIA', 'GPS', 'GFIP',
    'PROTOCOLO', 'COMPENSACAO', 'DECLARACAO',
    'ARQUIVO FGTS', 'ARQUIVO FGST', 'IMG_', 'Thumbs',
]
CONECTORES = {'DE', 'DA', 'DO', 'DOS', 'DAS'}
IGNORAR_NOMES = {
    'NOMES', 'NOME', 'TITULAR', 'FUNCIONARIO',
    'PIS', 'PROC.', 'DATA', 'OBS',
    'TOTAL', 'TOTAIS', 'SUBTOTAL',
}
RE_VARIANTE = re.compile(r'(REC\s*\.?\s*\d+|\b115\b)', re.IGNORECASE)

# ============================================================
# UTILITARIAS
# ============================================================

def normalizar(texto):
    if not isinstance(texto, str): return ""
    nfkd = unicodedata.normalize('NFKD', texto)
    s = "".join(c for c in nfkd if not unicodedata.combining(c)).upper()
    s = s.replace('.', '').replace('-', ' ').replace('\u2013', ' ').replace('\u2014', ' ').replace('_', ' ').replace('(', '').replace(')', '')
    return " ".join(s.split())

def levenshtein(a, b):
    m, n = len(a), len(b)
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev = dp[0]; dp[0] = i
        for j in range(1, n + 1):
            temp = dp[j]
            dp[j] = min(dp[j] + 1, dp[j - 1] + 1, prev + (0 if a[i-1] == b[j-1] else 1))
            prev = temp
    return dp[n]

def similaridade(a, b):
    return SequenceMatcher(None, a, b).ratio()

def extrair_variante(nome):
    m = RE_VARIANTE.search(nome)
    if m:
        raw = m.group(1)
        tag = re.sub(r'[\s\.]', '', raw).upper()
        if tag == '115':
            tag = 'REC115'
        stem = nome[:m.start()].strip() + ' ' + nome[m.end():].strip()
        return stem.strip(), tag
    return nome, None

def eh_nome_valido(texto):
    if not isinstance(texto, str): return False
    t = texto.strip().upper()
    if t in IGNORAR_NOMES or len(t) <= 5 or ' ' not in t: return False
    if re.match(r'^[A-ZÁÀÂÃÉÊÍÓÔÕÚÜÇ][\.\s]+[A-ZÁÀÂÃÉÊÍÓÔÕÚÜÇ]', t):
        if t.replace(' ', '').replace('.', '') in ('TOTAIS', 'TOTAL', 'SUBTOTAL'):
            return False
    return bool(re.match(r"^[A-ZÁÀÂÃÉÊÍÓÔÕÚÜÇ0-9 \.\-]+$", t))

def eh_arquivo_valido(nome_arquivo):
    if nome_arquivo.lower().endswith(('.xls', '.xlsx', '.ods')):
        return False
    n = os.path.splitext(nome_arquivo)[0].upper()
    return not any(p in n for p in EXCLUIR_PADROES)

# ============================================================
# MATCHING DE NOMES
# ============================================================

def verificar_abreviacao(a, b):
    ta = [t for t in a.split() if t not in CONECTORES]
    tb = [t for t in b.split() if not t.isdigit() and t not in CONECTORES]
    if abs(len(ta) - len(tb)) > 2:
        return False
    i = j = match = 0
    while i < len(ta) and j < len(tb):
        pa, pb = ta[i], tb[j]
        if pa == pb:
            match += 1; i += 1; j += 1; continue
        if (len(pa) <= 4 and pb.startswith(pa)) or (len(pb) <= 4 and pa.startswith(pb)):
            match += 1; i += 1; j += 1; continue
        if len(pa) >= 4 and len(pb) >= 4 and len(ta) == len(tb):
            if levenshtein(pa, pb) <= max(1, min(len(pa), len(pb)) // 3):
                match += 1; i += 1; j += 1; continue
            return False
        if len(pa) >= 4 and len(pb) >= 4:
            i += 1; j += 1
        elif len(ta) > len(tb): i += 1
        elif len(tb) > len(ta): j += 1
        else: i += 1; j += 1
    return match >= max(len(ta), len(tb)) - 1

def buscar_melhor_match(nome_busca, pdfs_disponiveis, usar_llm=False, nivel_rigor=3, tipo_lista="PADRAO"):
    stem, var = extrair_variante(nome_busca)
    nome_norm = normalizar(stem)

    def _buscar(grupo):
        if not grupo:
            return None, None, None

        # 1. Exata
        cand = []
        for p in grupo:
            s, _ = extrair_variante(os.path.splitext(p["real"])[0])
            n = normalizar(s)
            if nome_norm == n:
                cand.append((p, n))
        if cand:
            return cand[0][0], "ENCONTRADO", None

        if nivel_rigor == 1:
            return None, None, None

        # 2. Abreviacao
        melhor = None
        for p in grupo:
            s, _ = extrair_variante(os.path.splitext(p["real"])[0])
            n = normalizar(s)
            if verificar_abreviacao(nome_norm, n):
                if melhor is None or abs(len(n) - len(nome_norm)) < abs(len(melhor[1]) - len(nome_norm)):
                    melhor = (p, n)
        if melhor:
            razao = similaridade(nome_norm, melhor[1])
            if razao < 0.80:
                if usar_llm and llm_matcher:
                    resp = llm_matcher.verificar_com_llm(nome_norm, melhor[1])
                    if resp is True:
                        print(f'    [LLM] {nome_norm} x {melhor[1]} -> SIM')
                        return melhor[0], "ENCONTRADO VIA LLM", None
                    elif resp is False:
                        print(f'    [LLM] {nome_norm} x {melhor[1]} -> NAO')
                    else:
                        print(f'    [LLM] {nome_norm} x {melhor[1]} -> DUVIDA')
                        if nivel_rigor == 3:
                            return melhor[0], "POSSIVEL ERRO NOMINAL", None
                    if nivel_rigor == 1:
                        melhor = None
                else:
                    if nivel_rigor == 3:
                        return melhor[0], "POSSIVEL ERRO NOMINAL", None
            if melhor:
                return melhor[0], "ENCONTRADO COM ABREVIACAO", None

        # 3. Similaridade (primeiro com abreviacao, depois fallback direto)
        melhor = None
        melhor_razao = 0
        for p in grupo:
            s, _ = extrair_variante(os.path.splitext(p["real"])[0])
            n = normalizar(s)
            if not verificar_abreviacao(nome_norm, n):
                continue
            r = similaridade(nome_norm, n)
            if r > melhor_razao:
                melhor_razao = r
                melhor = (p, n)
        if melhor and melhor_razao >= 0.80:
            if usar_llm and llm_matcher:
                resp = llm_matcher.verificar_com_llm(nome_norm, melhor[1])
                if resp is True:
                    print(f'    [LLM] {nome_norm} x {melhor[1]} -> SIM')
                    return melhor[0], "ENCONTRADO VIA LLM", None
                elif resp is False:
                    print(f'    [LLM] {nome_norm} x {melhor[1]} -> NAO')
                else:
                    print(f'    [LLM] {nome_norm} x {melhor[1]} -> DUVIDA')
                    if nivel_rigor == 3:
                        return melhor[0], "POSSIVEL ERRO NOMINAL", None
            else:
                if nivel_rigor == 3:
                    return melhor[0], "POSSIVEL ERRO NOMINAL", None

        # 3b. Similaridade direta (fallback sem abreviacao, com primeiro nome igual)
        if not melhor:
            melhor = None
            melhor_razao = 0
            tokens_nome = [t for t in nome_norm.split() if t not in CONECTORES]
            if tokens_nome:
                for p in grupo:
                    s, _ = extrair_variante(os.path.splitext(p["real"])[0])
                    n = normalizar(s)
                    tokens_file = [t for t in n.split() if t not in CONECTORES]
                    if not tokens_file:
                        continue
                    # Primeiro nome deve ser igual ou abreviacao
                    p1, p2 = tokens_nome[0], tokens_file[0]
                    if p1 != p2 and not ((len(p1) <= 4 and p2.startswith(p1)) or (len(p2) <= 4 and p1.startswith(p2))):
                        continue
                    # Lista menor deve ser subconjunto da maior (todos os tokens devem casar)
                    if len(tokens_nome) <= len(tokens_file):
                        menor, maior = tokens_nome, tokens_file
                    else:
                        menor, maior = tokens_file, tokens_nome
                    todos_casam = True
                    for tm in menor:
                        casa = any(
                            tm == tm2 or
                            (len(tm) <= 4 and tm2.startswith(tm)) or
                            (len(tm2) <= 4 and tm.startswith(tm2)) or
                            (len(tm) >= 4 and len(tm2) >= 4 and levenshtein(tm, tm2) <= max(1, min(len(tm), len(tm2)) // 3))
                            for tm2 in maior
                        )
                        if not casa:
                            todos_casam = False
                            break
                    if not todos_casam:
                        continue
                    r = similaridade(nome_norm, n)
                    if r >= 0.85 and r > melhor_razao:
                        melhor_razao = r
                        melhor = (p, n)
            if melhor:
                if usar_llm and llm_matcher:
                    resp = llm_matcher.verificar_com_llm(nome_norm, melhor[1])
                    if resp is True:
                        print(f'    [LLM] {nome_norm} x {melhor[1]} -> SIM')
                        return melhor[0], "ENCONTRADO VIA LLM", None
                    elif resp is False:
                        print(f'    [LLM] {nome_norm} x {melhor[1]} -> NAO')
                    else:
                        if nivel_rigor == 3:
                            return melhor[0], "POSSIVEL ERRO NOMINAL", None
                else:
                    if nivel_rigor == 3:
                        return melhor[0], "POSSIVEL ERRO NOMINAL", None

        if nivel_rigor == 2:
            return None, None, None

        # 4. LLM fallback (so o melhor candidato por similaridade pura, sem abreviacao)
        if usar_llm and llm_matcher:
            melhor = None
            melhor_r = 0
            for p in grupo:
                s, _ = extrair_variante(os.path.splitext(p["real"])[0])
                n = normalizar(s)
                r = similaridade(nome_norm, n)
                if r > melhor_r:
                    melhor_r = r; melhor = (p, n)
            if melhor and melhor_r >= 0.75:
                resp = llm_matcher.verificar_com_llm(nome_norm, melhor[1])
                if resp is True:
                    print(f'    [LLM] {nome_norm} x {melhor[1]} -> SIM (fallback)')
                    return melhor[0], "ENCONTRADO VIA LLM", None
                elif resp == "DUVIDA":
                    print(f'    [LLM] {nome_norm} x {melhor[1]} -> DUVIDA (fallback, erro nominal)')
                    if nivel_rigor == 3:
                        return melhor[0], "POSSIVEL ERRO NOMINAL", None

        return None, None, None

    # Separa por variante (prioridade: tipo_lista 115 > variante do nome)
    if tipo_lista == "115":
        grupo_pri = [p for p in pdfs_disponiveis if p["variante"] == "REC115"]
        grupo_sec = [p for p in pdfs_disponiveis if p not in grupo_pri]
    elif var:
        grupo_pri = [p for p in pdfs_disponiveis if p["variante"] == var]
        grupo_sec = [p for p in pdfs_disponiveis if p not in grupo_pri]
    else:
        grupo_pri = [p for p in pdfs_disponiveis if not p["variante"]]
        grupo_sec = [p for p in pdfs_disponiveis if p not in grupo_pri]

    match, status, _ = _buscar(grupo_pri)
    var_diff = False
    if not match:
        match, status, _ = _buscar(grupo_sec)
        if match:
            var_diff = True

    return match, status, var_diff

# ============================================================
# LEITURA DA PLANILHA
# ============================================================

def ler_planilha(caminho):
    ext = os.path.splitext(caminho)[1].lower()
    engine = "odf" if ext == '.ods' else None
    try:
        df = pd.read_excel(caminho, sheet_name=SHEET_NAME, header=None, engine=engine)
    except Exception as e:
        print(f"  ERRO ao ler planilha: {e}")
        return None

    dados = []
    bloco = "PADRAO"
    achou = False
    gaps = 0
    for i in range(len(df)):
        nome = df.iloc[i, 3]
        col0 = str(df.iloc[i, 0]).strip().upper() if pd.notna(df.iloc[i, 0]) else ''
        if re.search(r'T\s*O\s*T\s*(A\s*I\s*S|A\s*L)', col0):
            continue
        if eh_nome_valido(nome):
            achou = True; gaps = 0
            dados.append({"PROC.": df.iloc[i, 1], "PIS": df.iloc[i, 2],
                          "NOMES": nome, "TIPO_LISTA": bloco})
        elif achou:
            gaps += 1
            if gaps > 5: bloco = "115"
            if gaps > 40: break
    return pd.DataFrame(dados) if dados else None

# ============================================================
# CONFERENCIA
# ============================================================

def conferir(planilha_path, pasta_pdfs, usar_llm=False):
    print(f"\n--- Lendo planilha: {os.path.basename(planilha_path)} ---")
    df = ler_planilha(planilha_path)
    if df is None or df.empty:
        print("  Nenhum dado valido na planilha.")
        return None

    # Lista PDFs
    pdfs_brutos = sorted(f for f in os.listdir(pasta_pdfs) if eh_arquivo_valido(f))
    excluidos = sorted(f for f in os.listdir(pasta_pdfs)
                       if not eh_arquivo_valido(f) and not f.lower().endswith(('.xls', '.xlsx', '.ods')))

    # Debug: mostra distribuicao por extensao
    ext_count = {}
    for f in pdfs_brutos:
        ext = os.path.splitext(f)[1].lower() or '(sem ext)'
        ext_count[ext] = ext_count.get(ext, 0) + 1
    print(f'  Arquivos por tipo: {", ".join(f"{k}: {v}" for k, v in sorted(ext_count.items()))}')

    arquivos = []
    for f in pdfs_brutos:
        stem, var = extrair_variante(os.path.splitext(f)[0])
        arquivos.append({"real": f, "norm": normalizar(stem), "variante": var})

    print(f"  Pessoas na planilha: {len(df)}")
    print(f"  Arquivos na pasta: {len(pdfs_brutos)}")
    if excluidos:
        print(f"  PDFs ignorados (sistema): {len(excluidos)}")
    print()

    status_lista = [""] * len(df)
    arquivo_lista = [""] * len(df)
    erros_nominais = []

    # 3 Passagens: 1 (Exato), 2 (Bom), 3 (Duvidoso/Fuzzy)
    for rigor in [1, 2, 3]:
        for idx, row in df.iterrows():
            if status_lista[idx]: 
                continue
                
            nome = row['NOMES']
            tipo = row.get('TIPO_LISTA', 'PADRAO')

            match, status, var_diff = buscar_melhor_match(nome, arquivos, usar_llm, nivel_rigor=rigor, tipo_lista=tipo)

            if match:
                arquivos.remove(match)
                if tipo == "115" and "ENCONTRADO" in status:
                    status = "ENCONTRADO COMO 115"
                if var_diff and "ENCONTRADO" in status:
                    status += " (VAR DIFF)"
                if "POSSIVEL ERRO NOMINAL" in status:
                    erros_nominais.append((nome, os.path.splitext(match["real"])[0]))
                status_lista[idx] = status
                arquivo_lista[idx] = os.path.splitext(match["real"])[0]

    for idx in range(len(df)):
        if not status_lista[idx]:
            # Debug: mostra o melhor candidato possivel
            nome_plan = df.iloc[idx]['NOMES']
            n_norm = normalizar(nome_plan)
            melhor_arq = None
            melhor_r = 0
            for a in arquivos:
                r = similaridade(n_norm, a["norm"])
                if r > melhor_r:
                    melhor_r = r
                    melhor_arq = a
            if melhor_arq and melhor_r > 0:
                print(f'    [DEBUG] NAO ENCONTRADO: {nome_plan}')
                print(f'            Melhor candidato: {melhor_arq["real"]} (similaridade: {melhor_r:.0%})')
                abrev = verificar_abreviacao(n_norm, melhor_arq["norm"])
                print(f'            Abreviacao: {abrev}')
            status_lista[idx] = "NAO ENCONTRADO NO PDF"
            arquivo_lista[idx] = ""

    df_result = df[['PROC.', 'PIS', 'NOMES']].copy()
    df_result['Status PDF'] = status_lista
    df_result['Nome do Arquivo Encontrado'] = arquivo_lista

    # Sobras
    sobras = []
    for a in arquivos:
        var_tag = f" [{a['variante']}]" if a['variante'] else ""
        sobras.append({"PROC.": "", "PIS": "", "NOMES": "",
                       "Status PDF": f"PDF NA PASTA, MAS NAO NA PLANILHA{var_tag}",
                       "Nome do Arquivo Encontrado": os.path.splitext(a["real"])[0]})
    if sobras:
        df_result = pd.concat([df_result, pd.DataFrame(sobras)], ignore_index=True)

    # Resumo
    totais = df_result['Status PDF'].value_counts()
    resumo = {
        'encontrado': totais.get('ENCONTRADO', 0) + totais.get('ENCONTRADO COMO 115', 0),
        'abreviacao': totais.get('ENCONTRADO COM ABREVIACAO', 0),
        'erro_nominal': totais.get('POSSIVEL ERRO NOMINAL', 0),
        'nao_encontrado': totais.get('NAO ENCONTRADO NO PDF', 0),
        'via_llm': totais.get('ENCONTRADO VIA LLM', 0),
        'pdf_extra': sum(1 for s in totais.index if str(s).startswith('PDF NA PASTA')),
    }

    return df_result, resumo, erros_nominais, excluidos, len(pdfs_brutos)

# ============================================================
# SALVAR RELATORIO
# ============================================================

def salvar_relatorio(df, caminho, total_pdfs, erros_nominais, excluidos):
    prioridade = {
        "NAO ENCONTRADO NO PDF": 1, "POSSIVEL ERRO NOMINAL": 2,
        "ENCONTRADO VIA LLM": 3, "PDF NA PASTA, MAS NAO NA PLANILHA": 4,
        "ENCONTRADO COMO 115": 5, "ENCONTRADO COM ABREVIACAO": 6, "ENCONTRADO": 7,
    }
    df = df.copy()
    df['_ordem'] = df['Status PDF'].map(prioridade).fillna(8)
    df = df.sort_values(['_ordem', 'NOMES']).drop(columns=['_ordem'])

    with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Dados', index=False)

    wb = openpyxl.load_workbook(caminho)

    # Aba Resumo
    ws = wb.create_sheet('Resumo', 0)
    ws.merge_cells('A1:B1')
    ws['A1'] = 'RESUMO DA CONFERENCIA'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    totais = df['Status PDF'].value_counts()
    qtd_encontrado = totais.get('ENCONTRADO', 0) + totais.get('ENCONTRADO COMO 115', 0)
    qtd_abrev = totais.get('ENCONTRADO COM ABREVIACAO', 0)
    qtd_erro = totais.get('POSSIVEL ERRO NOMINAL', 0)
    qtd_nao = totais.get('NAO ENCONTRADO NO PDF', 0)
    qtd_llm = totais.get('ENCONTRADO VIA LLM', 0)
    qtd_extra = sum(1 for s in totais.index if str(s).startswith('PDF NA PASTA'))
    total_pessoas = qtd_encontrado + qtd_abrev + qtd_erro + qtd_nao + qtd_llm

    linhas = [
        ('', ''),
        ('Planilha', os.path.basename(caminho)),
        ('', ''),
        ('PESSOAS NA PLANILHA', total_pessoas),
        ('ARQUIVOS PDF NA PASTA', total_pdfs),
    ]
    if qtd_extra > 0: linhas.append(('', ''))
    if qtd_encontrado > 0: linhas.append(('[OK] Encontrados', qtd_encontrado))
    if qtd_llm > 0: linhas.append(('[LLM] Encontrados via LLM', qtd_llm))
    if qtd_abrev > 0: linhas.append(('[~] Encontrados com abreviacao', qtd_abrev))
    if qtd_erro > 0: linhas.append(('[?] Possivel erro nominal', qtd_erro))
    if qtd_nao > 0: linhas.append(('[X] Nao encontrados no PDF', qtd_nao))
    if qtd_extra > 0: linhas.append(('[PDF] PDF na pasta mas nao na planilha', qtd_extra))
    linhas += [('', ''), ('TOTAL ENCONTRADOS', qtd_encontrado + qtd_abrev + qtd_llm),
               ('TOTAL NAO ENCONTRADOS', qtd_nao)]

    bdr = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    fill_ok = PatternFill('solid', fgColor='C6EFCE')
    fill_err = PatternFill('solid', fgColor='FFC7CE')
    fill_115 = PatternFill('solid', fgColor='FFC000')
    font_ok = Font(color='006100', bold=True)
    font_err = Font(color='9C0006', bold=True)
    font_115 = Font(color='9C5700', bold=True)

    for i, (desc, val) in enumerate(linhas, 3):
        ws.cell(row=i, column=1, value=desc).border = bdr
        c = ws.cell(row=i, column=2, value=val).border = bdr
        c.alignment = Alignment(horizontal='center')
        if 'Encontrados' in desc and 'LLM' not in desc:
            c.fill = fill_ok; c.font = font_ok
        elif 'erro' in desc.lower():
            c.fill = fill_err; c.font = font_err

    # Lista de nao encontrados
    linha_atual = 3 + len(linhas) + 2
    nao = df[df['Status PDF'] == 'NAO ENCONTRADO NO PDF']
    if not nao.empty:
        ws.cell(row=linha_atual, column=1, value='NAO ENCONTRADOS NO PDF:').font = Font(bold=True, color='FF0000')
        linha_atual += 1
        for _, r in nao.iterrows():
            ws.cell(row=linha_atual, column=1, value=r['NOMES'])
            ws.cell(row=linha_atual, column=2, value=r.get('PROC.', ''))
            linha_atual += 1

    # Erros nominais
    if erros_nominais:
        linha_atual += 1
        ws.cell(row=linha_atual, column=1, value='POSSIVEIS ERROS NOMINAIS:').font = Font(bold=True, color='FF8C00')
        linha_atual += 1
        for nome_p, nome_f in erros_nominais:
            ws.cell(row=linha_atual, column=1, value=nome_p)
            ws.cell(row=linha_atual, column=2, value=f'-> {nome_f}')
            linha_atual += 1

    # Excluidos
    if excluidos:
        linha_atual += 1
        ws.cell(row=linha_atual, column=1, value='ARQUIVOS IGNORADOS (sistema):').font = Font(bold=True, color='666666')
        linha_atual += 1
        for e in excluidos:
            ws.cell(row=linha_atual, column=1, value=e).font = Font(italic=True, color='888888')
            linha_atual += 1
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 25
    ws2 = wb['Dados']
    bdr2 = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    fill_cinza = PatternFill('solid', fgColor='D3D3D3')
    fonte_vermelha = Font(color='FF0000', bold=True)
    fonte_verde = Font(color='006400', bold=True)
    fonte_azul = Font(color='0000FF', bold=True)
    fonte_roxo = Font(color='800080', bold=True)
    fonte_laranja = Font(color='FF8C00', bold=True)
    fill_amarelo = PatternFill('solid', fgColor='FFFF00')

    col_status = None
    for cell in ws2[1]:
        if cell.value and 'Status' in str(cell.value):
            col_status = cell.column; break

    for row in ws2.iter_rows():
        for cell in row:
            cell.border = bdr2
            if cell.row == 1:
                cell.fill = fill_cinza; cell.font = Font(bold=True)
            elif col_status and cell.column == col_status:
                v = str(cell.value)
                if 'NAO ENCONTRADO' in v: cell.font = fonte_vermelha
                elif v.startswith('ENCONTRADO COMO 115'): cell.font = fonte_roxo
                elif 'ENCONTRADO VIA LLM' in v: cell.font = fonte_laranja
                elif 'ENCONTRADO' in v: cell.font = fonte_verde
                elif v.startswith('PDF NA PASTA'): cell.font = fonte_azul
                elif 'POSSIVEL ERRO NOMINAL' in v:
                    cell.fill = fill_amarelo; cell.font = Font(color='000000', bold=True)

    for col in ws2.columns:
        col_letter = col[0].column_letter
        length = max((len(str(c.value or '')) for c in col), default=10)
        ws2.column_dimensions[col_letter].width = min(length + 2, 50)

    wb.save(caminho)

# ============================================================
# ORGANIZAR PDFs (115 / 660)
# ============================================================

def organizar_pdfs(df, pasta_pdfs, pasta_destino, nome_ref=""):
    if not pasta_pdfs or not os.path.exists(pasta_pdfs):
        return
    if not pasta_destino or not os.path.exists(pasta_destino):
        return

    nome_pasta = nome_ref or os.path.basename(pasta_pdfs)
    padrao_subst = r'( - COMPETENCIA)'
    pasta_115 = re.sub(padrao_subst, r' 115\1', nome_pasta, flags=re.IGNORECASE)
    pasta_660 = re.sub(padrao_subst, r' 660\1', nome_pasta, flags=re.IGNORECASE)
    if pasta_115 == nome_pasta:
        pasta_115 = nome_pasta + ' 115'
        pasta_660 = nome_pasta + ' 660'

    caminho_115 = os.path.join(pasta_destino, pasta_115)
    caminho_660 = os.path.join(pasta_destino, pasta_660)
    count_115 = count_660 = 0

    for _, row in df.iterrows():
        status = str(row.get('Status PDF', ''))
        nome_arq = str(row.get('Nome do Arquivo Encontrado', ''))
        if not nome_arq or status.startswith('PDF NA PASTA'):
            continue

        caminho_pdf = os.path.join(pasta_pdfs, nome_arq + '.pdf')
        if not os.path.exists(caminho_pdf):
            caminho_pdf = os.path.join(pasta_pdfs, nome_arq + '.tif')
        if not os.path.exists(caminho_pdf):
            caminho_pdf = os.path.join(pasta_pdfs, nome_arq + '.tiff')
        if not os.path.exists(caminho_pdf):
            for f in os.listdir(pasta_pdfs):
                if os.path.splitext(f)[0] == nome_arq:
                    caminho_pdf = os.path.join(pasta_pdfs, f)
                    break
            else:
                continue

        if 'ENCONTRADO COMO 115' in status:
            os.makedirs(caminho_115, exist_ok=True)
            shutil.copy2(caminho_pdf, os.path.join(caminho_115, os.path.basename(caminho_pdf)))
            count_115 += 1
        else:
            os.makedirs(caminho_660, exist_ok=True)
            shutil.copy2(caminho_pdf, os.path.join(caminho_660, os.path.basename(caminho_pdf)))
            count_660 += 1

    if count_115 > 0 or count_660 > 0:
        print(f'  PDFs organizados:')
        if count_115 > 0: print(f'    {count_115} na pasta 115: {caminho_115}')
        if count_660 > 0: print(f'    {count_660} na pasta 660: {caminho_660}')


# ============================================================
# CORRECAO EM MASSA
# ============================================================

def corrigir_nomes_em_massa(pasta_raiz, pasta_referencia):
    print(f"\n--- Iniciando Correcao em Massa ---")
    gabarito = {}
    for f in os.listdir(pasta_referencia):
        if not f.lower().endswith(('.xls', '.xlsx', '.ods')) and not eh_arquivo_valido(f): continue
        stem, var = extrair_variante(os.path.splitext(f)[0])
        n = normalizar(stem)
        if n and n not in gabarito:
            gabarito[n] = stem
            
    print(f"  Gabarito lido: {len(gabarito)} nomes base na pasta referencia.")
    
    cache_correcoes = {}
    
    pastas_alvo = []
    for d in os.listdir(pasta_raiz):
        caminho_dir = os.path.join(pasta_raiz, d)
        if os.path.isdir(caminho_dir) and os.path.normpath(caminho_dir) != os.path.normpath(pasta_referencia):
            pastas_alvo.append(caminho_dir)
                   
    print(f"  Verificando {len(pastas_alvo)} pastas para correcao em massa...\n")
    
    for pasta in pastas_alvo:
        print(f"    Inspecionando: {os.path.basename(pasta)}")
        for f in os.listdir(pasta):
            if not eh_arquivo_valido(f): continue
        if any(p in f.upper() for p in EXCLUIR_PADROES): continue

        nome_original = os.path.splitext(f)[0]
        ext_original = os.path.splitext(f)[1]
        
        # Ignora arquivos que terminam com numero (ex: homonimos como "MANOEL TEIXEIRA 2")
        if re.search(r'\d$', nome_original.strip()):
            continue
            
        stem, var = extrair_variante(nome_original)
        n_atual = normalizar(stem)
        
        if not n_atual: continue
        
        if n_atual in gabarito:
            continue
            
        if n_atual in cache_correcoes:
            novo_nome_base = cache_correcoes[n_atual]
            if novo_nome_base is None:
                continue
            novo_nome = novo_nome_base + (f" {var}" if var else "") + ext_original
            if novo_nome != f:
                try:
                    target = os.path.join(pasta, novo_nome)
                    if not os.path.exists(target):
                        os.rename(os.path.join(pasta, f), target)
                except Exception as e:
                    print(f"      [Erro ao renomear {f}]: {e}")
            continue
            
        melhor = None
        melhor_r = 0
        for g_norm, g_stem in gabarito.items():
            if verificar_abreviacao(g_norm, n_atual):
                r = similaridade(g_norm, n_atual)
                if r > melhor_r:
                    melhor_r = r; melhor = g_norm
        
        if not melhor:
            for g_norm, g_stem in gabarito.items():
                r = similaridade(g_norm, n_atual)
                if r > melhor_r:
                    melhor_r = r; melhor = g_norm
        
        if melhor and melhor_r > 0.65:
            nome_bonito_gabarito = gabarito[melhor]
            nome_pasta_atual = os.path.basename(pasta)
            nome_pasta_ref = os.path.basename(pasta_referencia)
            print(f"\n      [DUVIDA] O arquivo '{nome_original}' (pasta: {nome_pasta_atual}) eh a mesma pessoa que '{nome_bonito_gabarito}' (pasta: {nome_pasta_ref})?")
            while True:
                resp = input(f"      (1 para SIM / 2 para NAO): ").strip()
                if resp in ('1', '2'): break
            
            if resp == '1':
                cache_correcoes[n_atual] = nome_bonito_gabarito
                novo_nome = nome_bonito_gabarito + (f" {var}" if var else "") + ext_original
                print(f"      -> Renomeando para '{novo_nome}' e salvo no cache.")
                try:
                    target = os.path.join(pasta, novo_nome)
                    if os.path.exists(target):
                        print(f"      [Aviso] '{novo_nome}' ja existe. Ignorando.")
                    else:
                        os.rename(os.path.join(pasta, f), target)
                except Exception as e:
                    print(f"      [Erro] {e}")
            else:
                cache_correcoes[n_atual] = None
                print(f"      -> Ignorado e salvo no cache.")


# ============================================================
# DIALOGOS
# ============================================================

def dialogo(titulo, tipo='file', multiplo=False):
    root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True); root.update()
    if tipo == 'file':
        if multiplo:
            r = filedialog.askopenfilenames(title=titulo, filetypes=[("Planilhas", "*.xls *.xlsx *.ods")])
        else:
            r = filedialog.askopenfilename(title=titulo, filetypes=[("Planilhas", "*.xls *.xlsx *.ods")])
    else:
        r = filedialog.askdirectory(title=titulo)
    root.destroy()
    return r

def dialogo_pastas(t1, t2):
    pastas = []
    root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True)
    primeiro = True
    while True:
        root.update()
        t = t1 if primeiro else t2
        p = filedialog.askdirectory(title=t)
        if not p: break
        if p not in pastas: pastas.append(p)
        primeiro = False
    root.destroy()
    return pastas

# ============================================================
# MAIN
# ============================================================

def main():
    usar_llm = False

    print('=' * 60)
    print('  CONFERENCIA FGTS x PDFs')
    print('=' * 60)
    print()
    print('Modos:')
    print('  1 — Executar')
    print('  2 — Executar com LLM (Gemini)')
    while True:
        modo = input('Opcao (1 ou 2): ').strip()
        if modo in ('1', '2'): break

    if modo == '2':
        usar_llm = True
        print()
        if llm_matcher:
            r = llm_matcher.verificar_com_llm('TESTE', 'TESTE')
            if r is None:
                print('  LLM indisponivel. Configure config_llm.py ou ignore.')
                input('Enter para continuar...')
            else:
                print('  LLM OK')
        else:
            print('  llm_matcher.py nao encontrado. Modo 2 indisponivel.')
            input('Enter para continuar sem LLM...')

    print()
    arquivos = dialogo('1. Selecione as Planilhas FGTS', 'file', multiplo=True)
    if not arquivos: return
    print(f'{len(arquivos)} planilha(s) selecionada(s).')

    pasta_raiz = dialogo('2. Selecione a PASTA RAIZ dos PDFs', 'dir')
    if not pasta_raiz: return

    print()
    while True:
        resp_massa = input("Voce ja corrigiu os nomes em alguma pasta de PDF e deseja aplicar essa correcao em massa nas outras pastas? (S/N): ").strip().upper()
        if resp_massa in ('S', 'N'): break
        
    if resp_massa == 'S':
        pasta_referencia = dialogo('Selecione a PASTA DE REFERENCIA (a que ja esta corrigida)', 'dir')
        if pasta_referencia:
            corrigir_nomes_em_massa(pasta_raiz, pasta_referencia)
        else:
            print("  Operacao cancelada ou pasta invalida.")

    pasta_salvamento = dialogo('3. Selecione a PASTA DE DESTINO', 'dir')
    if not pasta_salvamento: return

    print(f'\nMapeando pastas dentro de: {pasta_raiz} ...')
    mapa_pastas = {}
    for item in os.listdir(pasta_raiz):
        caminho = os.path.join(pasta_raiz, item)
        if os.path.isdir(caminho):
            nome_norm = normalizar(item.upper())
            for mes_num in ('01','02','03','04','05','06','07','08','09','10','11','12'):
                for sep in ('', ' '):
                    for ano_s in ('2007','2008','2009','2010','2011','2012','2013','2014','2015','2016','2017','2018','2019','2020','2021','2022','2023','2024','2025','2026'):
                        chave = f'{mes_num}_{ano_s}'
                        if f'{mes_num}{sep}{ano_s}' in nome_norm or chave in nome_norm.replace(' ', '_'):
                            mapa_pastas[chave] = caminho
                            break
                    if chave in mapa_pastas: break
                if chave in mapa_pastas: break

    if mapa_pastas:
        print(f'  Pastas encontradas: {", ".join(sorted(mapa_pastas.keys()))}')
    else:
        print('  Nenhuma pasta com data reconhecida.')

    for planilha in arquivos:
        nome = os.path.basename(planilha)
        print(f'\n{"="*60}')
        print(f'  Processando: {nome}')
        print(f'{"="*60}')

        mes, ano = None, None
        m = re.search(r"(?<!\d)(\d{2})[-_ ](\d{2,4})(?!\d)", nome)
        if m:
            mes = m.group(1)
            ano = "20" + m.group(2) if len(m.group(2)) == 2 else m.group(2)

        pasta_pdfs = None
        if mes and ano:
            chave = f'{mes}_{ano}'
            pasta_pdfs = mapa_pastas.get(chave)

        if not pasta_pdfs:
            print('  Pasta de PDFs nao encontrada automaticamente.')
            pasta_pdfs = dialogo(f'  Selecione a pasta para {nome}', 'dir')
            if not pasta_pdfs:
                print('  Pulando...')
                continue

        try:
            resultado = conferir(planilha, pasta_pdfs, usar_llm)
        except Exception as e:
            import traceback
            print(f'  ERRO NA CONFERENCIA: {e}')
            traceback.print_exc()
            resultado = None
        if resultado is None: continue

        df_result, resumo, erros_nominais, excluidos, total_pdfs = resultado

        nome_saida = f"{mes}_{ano}.xlsx" if mes else f"Relatorio_{os.path.splitext(nome)[0]}.xlsx"
        caminho_saida = os.path.join(pasta_salvamento, nome_saida)

        try:
            salvar_relatorio(df_result, caminho_saida, total_pdfs, erros_nominais, excluidos)
            print(f'  Salvo: {nome_saida}')
        except Exception as e:
            print(f'  Erro ao salvar: {e}')
            continue

        nao = df_result[df_result['Status PDF'] == 'NAO ENCONTRADO NO PDF']
        if not nao.empty:
            print(f'  NAO ENCONTRADOS ({len(nao)}):')
            for _, r in nao.iterrows():
                print(f'    {r["NOMES"]}')

        if erros_nominais:
            print(f'  POSSIVEIS ERROS NOMINAIS ({len(erros_nominais)}):')
            for nome_p, nome_f in erros_nominais:
                print(f'    Planilha: {nome_p}  ->  PDF: {nome_f}')

        extras = df_result[df_result['Status PDF'].apply(lambda s: str(s).startswith('PDF NA PASTA'))]
        if not extras.empty:
            print(f'  PDFS EXTRAS ({len(extras)}):')
            for _, r in extras.iterrows():
                print(f'    {r["Nome do Arquivo Encontrado"]}')

        if excluidos:
            print(f'  IGNORADOS (sistema):')
            for e in excluidos:
                print(f'    {e}')

        organizar_pdfs(df_result, pasta_pdfs, pasta_salvamento, os.path.basename(pasta_pdfs))

    print(f'\n{"="*60}')
    print('  PROCESSO CONCLUIDO!')
    print(f'  Relatorios em: {pasta_salvamento}')
    print(f'{"="*60}')
    os.startfile(pasta_salvamento)
    input('\nPressione Enter para sair...')

if __name__ == '__main__':
    main()
