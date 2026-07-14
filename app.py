#!/usr/bin/env python3
"""
Gerenciador de PDFs — Renomeação por Planilha
Flask web app para corrigir nomes de PDFs usando planilhas de referência.
"""
from flask import Flask, render_template, request, jsonify
import os, re, unicodedata, datetime, json, shutil
from difflib import SequenceMatcher
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import tempfile
import sys
import xlrd
from odf import opendocument
from odf.table import Table, TableRow, TableCell
from odf.text import P
import llm_matcher
import confinicial
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import math

app = Flask(__name__)
app.secret_key = 'gerenciador_pdfs_2024'

# ─── Logging ───────────────────────────────────────────────────────────
# Erros vao para o console E para o arquivo app.log (rotativo, 3 x 1 MB),
# em vez de sumirem silenciosamente. Consulte app.log quando algo falhar.
import logging
from logging.handlers import RotatingFileHandler
_LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app.log')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        RotatingFileHandler(_LOG_FILE, maxBytes=1_000_000, backupCount=3, encoding='utf-8'),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger('gerenciador_pdfs')

# ─── Estado global (app local, single-user) ────────────────────────────
_state = {
    'pasta_raiz': '',
    'pastas_ref': [],
    'nomes_ref': [],      # nomes normalizados de todas as pastas de referência
    'analise': None,
    'needs_update': False,
    'planilhas': {}       # { '01-2008': {'caminho': '...', 'df': DataFrame, 'nome_original': '...'} }
}

# Progresso da analise de conciliacao (lido pelo frontend via polling)
_progresso_conc = {'atual': 0, 'total': 0, 'rodando': False, 'label': ''}

HISTORICO_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'historico.json')

def load_historico():
    if os.path.exists(HISTORICO_FILE):
        try:
            with open(HISTORICO_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            log.exception("Falha ao ler o historico (%s)", HISTORICO_FILE)
    return []

def save_historico(h):
    try:
        with open(HISTORICO_FILE, 'w', encoding='utf-8') as f:
            json.dump(h, f, indent=2, ensure_ascii=False)
    except Exception:
        log.exception("Falha ao salvar o historico (%s)", HISTORICO_FILE)

def add_historico(acao, arquivo_original, novo_nome, pasta, status, msg=""):
    h = load_historico()
    h.append({
        'data': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'acao': acao,
        'arquivo_original': arquivo_original,
        'novo_nome': novo_nome,
        'pasta': pasta,
        'status': status,
        'msg': msg
    })
    # Keep last 1000
    if len(h) > 1000: h = h[-1000:]
    save_historico(h)

class PDFWatcher(FileSystemEventHandler):
    def on_any_event(self, event):
        if event.is_directory: return
        p1 = getattr(event, 'src_path', '').lower()
        p2 = getattr(event, 'dest_path', '').lower()
        if p1.endswith(('.pdf', '.tif', '.tiff')) or p2.endswith(('.pdf', '.tif', '.tiff')):
            _state['needs_update'] = True

observer = None
def setup_watchdog(path):
    global observer
    if observer:
        observer.stop()
        observer.join()
    observer = Observer()
    observer.schedule(PDFWatcher(), path, recursive=True)
    observer.start()

# ─── Constantes ────────────────────────────────────────────────────────
EXCLUIR_PADROES = [
    'ARQUIVO SEFIP', 'SEFIP', 'GRRF', 'DEPOSITADO', 'COMPROVANTE',
    'RECIBO', 'EXTRATO', 'FOLHA', 'RELATORIO', 'GUIA', 'GPS', 'GFIP',
    'PROTOCOLO', 'COMPENSACAO', 'DECLARACAO',
    'ARQUIVO FGTS', 'ARQUIVO FGST', 'IMG_', 'Thumbs',
]
CONECTORES = {'DE', 'DA', 'DO', 'DOS', 'DAS'}
RE_VARIANTE = re.compile(r'(REC\s*\.?\s*\d+|\b115\b)', re.IGNORECASE)

# ─── Utilitárias ───────────────────────────────────────────────────────
# normalizar / similaridade / levenshtein agora vivem em utils.py (compartilhado)
from utils import normalizar, similaridade, levenshtein

def extrair_variante(nome):
    m = RE_VARIANTE.search(nome)
    if m:
        raw = m.group(1)
        tag = re.sub(r'[\s\.]', '', raw).upper()
        if tag == '115':
            tag = 'REC115'
        stem = nome[:m.start()].strip() + ' ' + nome[m.end():].strip()
        return stem.strip(), tag
    return nome, None

def eh_pdf_valido(f):
    if not f.lower().endswith(('.pdf', '.tif', '.tiff')): return False
    n = os.path.splitext(f)[0].upper()
    return not any(p in n for p in EXCLUIR_PADROES)

def eh_homonimo(nome_sem_ext):
    return bool(re.search(r'\s\d+$', nome_sem_ext.strip()))

def verificar_abreviacao(a, b):
    ta = [t for t in a.split() if t not in CONECTORES]
    tb = [t for t in b.split() if not t.isdigit() and t not in CONECTORES]
    if abs(len(ta) - len(tb)) > 2: return False
    i = j = mc = 0
    while i < len(ta) and j < len(tb):
        pa, pb = ta[i], tb[j]
        if pa == pb: mc += 1; i += 1; j += 1; continue
        if (len(pa) <= 4 and pb.startswith(pa)) or (len(pb) <= 4 and pa.startswith(pb)):
            mc += 1; i += 1; j += 1; continue
        if len(pa) >= 4 and len(pb) >= 4 and len(ta) == len(tb):
            if levenshtein(pa, pb) <= max(1, min(len(pa), len(pb)) // 3):
                mc += 1; i += 1; j += 1; continue
            return False
        if len(pa) >= 4 and len(pb) >= 4: i += 1; j += 1
        elif len(ta) > len(tb): i += 1
        elif len(tb) > len(ta): j += 1
        else: i += 1; j += 1
    return mc >= max(len(ta), len(tb)) - 1

def ler_nomes_pasta_ref(pasta):
    nomes = []
    if not pasta or not os.path.isdir(pasta): return nomes
    for f in sorted(os.listdir(pasta)):
        if not eh_pdf_valido(f): continue
        stem, _ = extrair_variante(os.path.splitext(f)[0])
        if eh_homonimo(stem): continue
        n = normalizar(stem)
        if n and n not in nomes:
            nomes.append(n)
    return nomes

def buscar_todos_pdfs(pasta_raiz):
    result = []
    if not pasta_raiz or not os.path.isdir(pasta_raiz): return result
    for d in sorted(os.listdir(pasta_raiz)):
        cp = os.path.join(pasta_raiz, d)
        if not os.path.isdir(cp): continue
        for f in sorted(os.listdir(cp)):
            if not eh_pdf_valido(f): continue
            stem, var = extrair_variante(os.path.splitext(f)[0])
            result.append({
                'pasta': d,
                'pasta_path': cp,
                'arquivo': f,
                'stem': stem,
                'nome_norm': normalizar(stem),
                'variante': var,
            })
    return result

def encontrar_match_ref(nome_norm, nome_stem, nomes_ref, ignorar_acentos=False):
    """Retorna (nome_ref, tipo, razao). Tipos: correto|rename_auto|duvida|sem_match"""
    # 1. Exata
    for ref in nomes_ref:
        if normalizar(ref) == nome_norm:
            if ignorar_acentos or ref == nome_stem:
                return ref, 'correto', 1.0
            return ref, 'rename_auto', 1.0

    # 2. Abreviação
    melhor = None; best_diff = 9999
    for ref in nomes_ref:
        rn = normalizar(ref)
        if verificar_abreviacao(nome_norm, rn):
            diff = abs(len(rn) - len(nome_norm))
            if diff < best_diff:
                best_diff = diff; melhor = ref
    if melhor:
        r = similaridade(nome_norm, normalizar(melhor))
        return melhor, ('rename_auto' if r >= 0.82 else 'duvida'), round(r, 3)

    # 3. Similaridade (exige primeiro nome igual + token overlap)
    melhor = None; best_r = 0
    for ref in nomes_ref:
        rn = normalizar(ref)
        tn = [t for t in nome_norm.split() if t not in CONECTORES]
        tr = [t for t in rn.split() if t not in CONECTORES]
        if not tn or not tr: continue
        # Primeiro nome deve ser igual ou abreviacao
        if tn[0] != tr[0] and not ((len(tn[0]) <= 4 and tr[0].startswith(tn[0])) or (len(tr[0]) <= 4 and tn[0].startswith(tr[0]))):
            continue
        # Lista menor deve ser subconjunto da maior (com levenshtein)
        if len(tn) <= len(tr): menor, maior = tn, tr
        else: menor, maior = tr, tn
        todos_casam = True
        for tm in menor:
            casa = any(
                tm == tm2 or
                (len(tm) <= 4 and tm2.startswith(tm)) or
                (len(tm2) <= 4 and tm.startswith(tm2)) or
                (len(tm) >= 4 and len(tm2) >= 4 and levenshtein(tm, tm2) <= max(1, min(len(tm), len(tm2)) // 3))
                for tm2 in maior
            )
            if not casa: todos_casam = False; break
        if not todos_casam: continue
        r = similaridade(nome_norm, rn)
        if r > best_r:
            best_r = r; melhor = ref
    if melhor and best_r >= 0.60:
        return melhor, 'duvida', round(best_r, 3)

    return None, 'sem_match', 0

def abrir_pasta():
    try:
        dlg = tk.Tk()
        dlg.withdraw()
        dlg.attributes('-topmost', True)
        dlg.update()
        p = filedialog.askdirectory(title='Selecione uma pasta')
        dlg.destroy()
        return p or ''
    except Exception:
        return ''

def abrir_arquivos():
    try:
        dlg = tk.Tk()
        dlg.withdraw()
        dlg.attributes('-topmost', True)
        dlg.update()
        files = filedialog.askopenfilenames(
            title='Selecione planilhas',
            filetypes=[('Planilhas', '*.xls *.xlsx *.ods')]
        )
        dlg.destroy()
        return list(files)
    except Exception:
        return []

# ─── Rotas ────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/browse-folder', methods=['POST'])
def browse_folder():
    return jsonify({'path': abrir_pasta()})

@app.route('/api/set-pasta-raiz', methods=['POST'])
def set_pasta_raiz():
    p = request.json.get('path', '')
    if not p or not os.path.isdir(p):
        return jsonify({'ok': False, 'erro': 'Pasta inválida ou não encontrada'})
    _state['pasta_raiz'] = p
    subs = [d for d in os.listdir(p) if os.path.isdir(os.path.join(p, d))]
    total = sum(len([f for f in os.listdir(os.path.join(p, d)) if eh_pdf_valido(f)]) for d in subs)
    
    setup_watchdog(p)
    _state['needs_update'] = False
    
    return jsonify({'ok': True, 'pasta': p, 'subpastas': len(subs), 'total_pdfs': total})

@app.route('/api/add-pasta-ref', methods=['POST'])
def add_pasta_ref():
    p = request.json.get('path', '')
    if not p or not os.path.isdir(p):
        return jsonify({'ok': False, 'erro': 'Pasta de referência inválida ou não encontrada'})
    if p in _state['pastas_ref']:
        return jsonify({'ok': False, 'erro': 'Esta pasta já foi adicionada'})
        
    nomes = ler_nomes_pasta_ref(p)
    if not nomes:
        return jsonify({'ok': False, 'erro': 'Nenhum nome válido (PDF) encontrado na pasta'})
    
    _state['pastas_ref'].append(p)
    _state['nomes_ref'].extend(n for n in nomes if n not in _state['nomes_ref'])
    
    return jsonify({
        'ok': True,
        'pastas': _state['pastas_ref'],
        'total': len(_state['nomes_ref']),
    })

@app.route('/api/clear-pastas-ref', methods=['POST'])
def clear_pastas_ref():
    _state['pastas_ref'] = []
    _state['nomes_ref'] = []
    return jsonify({'ok': True})

@app.route('/api/nomes-ref', methods=['GET'])
def get_nomes_ref():
    return jsonify({'nomes': _state['nomes_ref']})

@app.route('/api/check-update', methods=['GET'])
def check_update():
    if _state['needs_update']:
        _state['needs_update'] = False
        return jsonify({'update': True})
    return jsonify({'update': False})

@app.route('/api/historico', methods=['GET'])
def get_historico():
    h = load_historico()
    h.reverse()
    return jsonify({'historico': h})

@app.route('/api/shutdown', methods=['POST'])
def shutdown():
    global observer
    if observer:
        try:
            observer.stop()
        except: pass
    
    def kill_server():
        import time
        time.sleep(0.5)
        os._exit(0)
    
    import threading
    threading.Thread(target=kill_server).start()
    return jsonify({'ok': True})

@app.route('/api/abrir-pdf', methods=['POST'])
def abrir_pdf():
    caminho = request.json.get('caminho')
    if not caminho or not os.path.exists(caminho):
        return jsonify({'ok': False, 'erro': 'Arquivo não encontrado no computador.'})
    try:
        os.startfile(caminho)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'erro': f'Erro ao abrir: {str(e)}'})

@app.route('/api/analisar', methods=['POST'])
def analisar():
    if not _state['pasta_raiz']:
        return jsonify({'ok': False, 'erro': 'Selecione a pasta raiz primeiro'})

    # SEMPRE recarrega as pastas de referencia do disco
    nomes_ref = []
    for p in _state['pastas_ref']:
        nomes_ref.extend(n for n in ler_nomes_pasta_ref(p) if n not in nomes_ref)
    _state['nomes_ref'] = nomes_ref

    if not nomes_ref:
        return jsonify({'ok': False, 'erro': 'Carregue a pasta de referência primeiro'})

    ignorar_acentos = request.json.get('ignorar_acentos', False)
    pdfs = buscar_todos_pdfs(_state['pasta_raiz'])
    nomes_ref = _state['nomes_ref']
    corretos = []; rename_auto = []; duvidas = []; sem_match = []

    for pdf in pdfs:
        sem_ext = os.path.splitext(pdf['arquivo'])[0]
        if eh_homonimo(sem_ext):
            corretos.append({**pdf, 'nome_ref': pdf['stem'], 'tipo': 'homonimo', 'razao': 1.0})
            continue
        ref, tipo, razao = encontrar_match_ref(pdf['nome_norm'], pdf['stem'], nomes_ref, ignorar_acentos)
        entry = {**pdf, 'nome_ref': ref, 'tipo': tipo, 'razao': razao}
        if tipo == 'correto':
            corretos.append(entry)
        elif tipo == 'rename_auto':
            rename_auto.append(entry)
        elif tipo == 'duvida':
            duvidas.append(entry)
        else:
            sem_match.append(entry)

    _state['analise'] = {
        'corretos': corretos,
        'rename_auto': rename_auto,
        'duvidas': duvidas,
        'sem_match': sem_match,
    }

    return jsonify({
        'ok': True,
        'corretos': len(corretos),
        'rename_auto': len(rename_auto),
        'duvidas': len(duvidas),
        'sem_match': len(sem_match),
        'rename_auto_data': rename_auto,
        'duvidas_data': duvidas,
        'sem_match_data': sem_match,
    })

@app.route('/api/executar', methods=['POST'])
def executar():
    items = request.json.get('items', [])
    resultados = []
    for item in items:
        pasta_path = item['pasta_path']
        arquivo = item['arquivo']
        nome_ref = item.get('nome_ref', '')
        if not nome_ref:
            resultados.append({'arquivo': arquivo, 'status': 'erro', 'msg': 'Nome de referência não informado'})
            continue
        origem = os.path.join(pasta_path, arquivo)
        if not os.path.exists(origem):
            resultados.append({'arquivo': arquivo, 'status': 'erro', 'msg': 'Arquivo não encontrado'})
            continue
        stem, var = extrair_variante(os.path.splitext(arquivo)[0])
        ext_original = os.path.splitext(arquivo)[1] or '.pdf'
        novo_nome = nome_ref + (' ' + var if var else '') + ext_original
        destino = os.path.join(pasta_path, novo_nome)
        pasta_nome = os.path.basename(pasta_path)
        if arquivo == novo_nome:
            resultados.append({'arquivo': arquivo, 'pasta': pasta_nome, 'status': 'igual', 'msg': 'Já estava correto'})
            continue
        if os.path.exists(destino):
            if os.path.getsize(origem) == os.path.getsize(destino):
                try:
                    os.remove(origem)
                    resultados.append({'arquivo': arquivo, 'pasta': pasta_nome, 'status': 'excluido',
                                       'msg': f'Duplicado removido — "{novo_nome}" já existia'})
                    add_historico('Lote (Excluído)', arquivo, '-', pasta_nome, 'Sucesso', 'Duplicado removido')
                except Exception as e:
                    resultados.append({'arquivo': arquivo, 'pasta': pasta_nome, 'status': 'erro', 'msg': str(e)})
                    add_historico('Lote (Excluído)', arquivo, '-', pasta_nome, 'Erro', str(e))
            else:
                resultados.append({'arquivo': arquivo, 'pasta': pasta_nome, 'status': 'conflito',
                                   'msg': f'Conflito: "{novo_nome}" existe com conteúdo diferente'})
            continue
        try:
            os.rename(origem, destino)
            resultados.append({'arquivo': arquivo, 'pasta': pasta_nome, 'status': 'ok', 'msg': f'→ {novo_nome}'})
            add_historico('Lote', arquivo, novo_nome, pasta_nome, 'Sucesso')
        except Exception as e:
            resultados.append({'arquivo': arquivo, 'pasta': pasta_nome, 'status': 'erro', 'msg': str(e)})
            add_historico('Lote', arquivo, novo_nome, pasta_nome, 'Erro', str(e))

    return jsonify({'ok': True, 'resultados': resultados})

@app.route('/api/buscar', methods=['POST'])
def buscar():
    termo = normalizar(request.json.get('termo', ''))
    if not termo or len(termo) < 3:
        return jsonify({'ok': False, 'erro': 'Digite pelo menos 3 caracteres'})
    if not _state['pasta_raiz']:
        return jsonify({'ok': False, 'erro': 'Selecione a pasta raiz primeiro (aba Configuração)'})
    pdfs = buscar_todos_pdfs(_state['pasta_raiz'])
    resultados = []
    for pdf in pdfs:
        n = pdf['nome_norm']
        r = similaridade(termo, n)
        if r >= 0.5 or termo in n:
            resultados.append({
                'pasta': pdf['pasta'],
                'pasta_path': pdf['pasta_path'],
                'arquivo': pdf['arquivo'],
                'stem': pdf['stem'],
                'razao': round(r, 3),
            })
    resultados.sort(key=lambda x: -x['razao'])
    return jsonify({'ok': True, 'resultados': resultados[:80]})

@app.route('/api/renomear', methods=['POST'])
def renomear():
    data = request.json
    pasta_path = data.get('pasta_path', '')
    arquivo = data.get('arquivo', '')
    novo_nome = data.get('novo_nome', '').strip()
    if not novo_nome:
        return jsonify({'ok': False, 'erro': 'Nome não informado'})
    nome_base, ext = os.path.splitext(arquivo)
    if not novo_nome.lower().endswith(('.pdf', '.tif', '.tiff')):
        novo_nome += ext if ext else '.pdf'
    origem = os.path.join(pasta_path, arquivo)
    destino = os.path.join(pasta_path, novo_nome)
    if not os.path.exists(origem):
        return jsonify({'ok': False, 'erro': 'Arquivo não encontrado'})
    if os.path.normcase(origem) == os.path.normcase(destino):
        return jsonify({'ok': False, 'erro': 'O nome é igual ao atual'})
    if os.path.exists(destino):
        return jsonify({'ok': False, 'erro': f'"{novo_nome}" já existe nessa pasta'})
    try:
        os.rename(origem, destino)
        add_historico('Busca (Individual)', arquivo, novo_nome, os.path.basename(pasta_path), 'Sucesso')
        return jsonify({'ok': True, 'novo_nome': novo_nome})
    except Exception as e:
        add_historico('Busca (Individual)', arquivo, novo_nome, os.path.basename(pasta_path), 'Erro', str(e))
        return jsonify({'ok': False, 'erro': str(e)})

# ─── Funções Auxiliares de Conciliação ──────────────────────────────────
def converter_xls_para_xlsx(caminho_xls):
    caminho_xlsx = os.path.splitext(caminho_xls)[0] + '.xlsx'
    if os.path.exists(caminho_xlsx):
        os.remove(caminho_xlsx)

    try:
        import win32com.client, pythoncom
        pythoncom.CoInitialize()
        xl = win32com.client.DispatchEx('Excel.Application')
        try:
            xl.DisplayAlerts = False
        except:
            pass
        wb = xl.Workbooks.Open(caminho_xls)
        wb.SaveAs(caminho_xlsx, FileFormat=51)  # 51 = xlOpenXMLWorkbook
        wb.Close()
        xl.Quit()
        pythoncom.CoUninitialize()
        return caminho_xlsx
    except Exception as e:
        try:
            pythoncom.CoUninitialize()
        except:
            pass
        try:
            import xlrd
            rb = xlrd.open_workbook(caminho_xls)
            nome_aba = 'FGTS EM ATRASO - PROCESSOS'
            if nome_aba not in rb.sheet_names():
                nome_aba = rb.sheet_names()[0]
            df_raw = pd.read_excel(caminho_xls, sheet_name=nome_aba, header=None)
            df_raw.to_excel(caminho_xlsx, sheet_name='FGTS EM ATRASO - PROCESSOS', index=False, header=False)
            return caminho_xlsx
        except:
            return None

def normalizar_pis(valor):
    if valor is None:
        return ''
    if isinstance(valor, float):
        if math.isnan(valor) or valor == 0:
            return ''
        return str(int(valor)).zfill(11)
    texto = str(valor).strip()
    if not texto or texto == '0':
        return ''
    digitos = re.sub(r'\D', '', texto)
    return digitos.zfill(11)

# ─── Conciliação de Planilhas ──────────────────────────────────────────

@app.route('/api/carregar-planilhas', methods=['POST'])
def carregar_planilhas():
    try:
        dlg = tk.Tk(); dlg.withdraw(); dlg.attributes('-topmost', True); dlg.update()
        caminhos = filedialog.askopenfilenames(
            title="Selecione as Planilhas FGTS",
            filetypes=[("Planilhas", "*.xlsx *.xls *.ods"), ("Todos", "*.*")]
        )
        dlg.destroy()
    except Exception as e:
        return jsonify({'ok': False, 'erro': f'Erro ao abrir seletor de arquivos: {str(e)}'})
    
    if not caminhos:
        return jsonify({'ok': False, 'erro': 'Nenhum arquivo selecionado'})
        
    sucessos = 0
    erros = []
    
    for c in caminhos:
        mes, ano = confinicial.extrair_competencia_excel(os.path.basename(c))
        if not mes or not ano:
            erros.append(f"Não foi possível extrair a competência do arquivo: {os.path.basename(c)}")
            continue
            
        comp_str = f"{mes}-{ano}"
        df = confinicial.ler_planilha_fgts(c)
        if df is None or df.empty:
            erros.append(f"Planilha sem dados válidos ou falha na leitura: {os.path.basename(c)}")
            continue
            
        _state['planilhas'][comp_str] = {
            'caminho': c,
            'nome_original': os.path.basename(c),
            'df': df
        }
        sucessos += 1
        
    return jsonify({'ok': True, 'sucessos': sucessos, 'erros': erros})

@app.route('/api/status-conciliacao', methods=['GET'])
def status_conciliacao():
    if not _state['planilhas']:
        return jsonify({'ok': True, 'cards': [], 'conflitos': []})
        
    # Agrupa por competência para os cards
    cards = []
    pis_dict = {} # PIS -> dict de nomes (nome -> lista de competencias)
    
    for comp, p_data in _state['planilhas'].items():
        df = p_data['df']
        cards.append({
            'competencia': comp,
            'arquivo': p_data['nome_original'],
            'total': len(df)
        })
        
        for idx, row in df.iterrows():
            pis = str(row['PIS']).strip()
            nome = str(row['NOMES']).strip()
            if pis and nome and str(pis).lower() not in ('nan', 'none', ''):
                if pis not in pis_dict:
                    pis_dict[pis] = {}
                # normalizar espacos para comparação
                n_clean = " ".join(nome.split())
                if n_clean not in pis_dict[pis]:
                    pis_dict[pis][n_clean] = []
                if comp not in pis_dict[pis][n_clean]:
                    pis_dict[pis][n_clean].append(comp)
                    
    # Encontra conflitos
    conflitos = []
    # Usar _state['nomes_ref'] (normalizado) para sugerir melhor nome
    nomes_ref_norm = _state['nomes_ref']
    
    for pis, nomes_map in pis_dict.items():
        if len(nomes_map) > 1:
            opcoes = list(nomes_map.keys())
            # Tenta achar um que bate com a pasta de referencia
            sugerido = None
            for op in opcoes:
                if normalizar(op) in nomes_ref_norm:
                    sugerido = op
                    break
            if not sugerido:
                sugerido = sorted(opcoes, key=len, reverse=True)[0]
                
            opcoes_detalhe = []
            for op in opcoes:
                opcoes_detalhe.append({
                    'nome': op,
                    'competencias': nomes_map[op]
                })
                
            conflitos.append({
                'pis': pis,
                'opcoes': opcoes_detalhe,
                'sugerido': sugerido
            })
            
    # Ordena cards
    cards.sort(key=lambda x: x['competencia'])
    return jsonify({'ok': True, 'cards': cards, 'conflitos': conflitos})

@app.route('/api/corrigir-pis', methods=['POST'])
def corrigir_pis():
    data = request.json
    pis = str(data.get('pis')).strip()
    novo_nome = data.get('novo_nome', '').strip()
    
    if not pis or not novo_nome:
        return jsonify({'ok': False, 'erro': 'Dados incompletos'})
        
    afetados = 0
    for comp, p_data in _state['planilhas'].items():
        df = p_data['df']
        # Usamos apply/lambda pra converter column pra str sem alterar tipo original
        mascara = df['PIS'].astype(str).str.strip() == pis
        if mascara.any():
            # Acha os que estao diferentes
            diferentes = mascara & (df['NOMES'] != novo_nome)
            qtd = diferentes.sum()
            if qtd > 0:
                df.loc[diferentes, 'NOMES'] = novo_nome
                afetados += qtd
                
    return jsonify({'ok': True, 'afetados': int(afetados)})

@app.route('/api/limpar-conciliacao', methods=['POST'])
def limpar_conciliacao():
    _state['planilhas'] = {}
    return jsonify({'ok': True})

@app.route('/api/analisar-conciliacao', methods=['POST'])
def analisar_conciliacao():
    if not _state['pasta_raiz']:
        return jsonify({'ok': False, 'erro': 'Selecione a Pasta Raiz de PDFs na aba de Configuração primeiro.'})
    if not _state['planilhas']:
        return jsonify({'ok': False, 'erro': 'Nenhuma planilha carregada.'})

    mapa_pastas = {}
    for d in os.listdir(_state['pasta_raiz']):
        caminho_dir = os.path.join(_state['pasta_raiz'], d)
        if os.path.isdir(caminho_dir):
            mes, ano = confinicial.extrair_competencia_pasta(d)
            if mes and ano:
                mapa_pastas[f"{mes}-{ano}"] = caminho_dir

    resultados_audit = {}

    def _status_base(s):
        for prefixo in [
            "NÃO ENCONTRADO NO .PDF",
            "DUPLICADO NA PLANILHA",
            "POSSÍVEL ERRO NOMINAL",
            "ENCONTRADO VIA LLM",
            "ENCONTRADO COMO 115",
            "ENCONTRADO COM ABREVIAÇÃO",
            "ENCONTRADO",
        ]:
            if str(s).startswith(prefixo):
                return prefixo
        if str(s).startswith("PDF NA PASTA"):
            return "PDF NA PASTA, MAS NÃO NA PLANILHA"
        return str(s)

    _progresso_conc.update(atual=0, total=len(_state['planilhas']), rodando=True, label='')

    for comp, p_data in _state['planilhas'].items():
        _progresso_conc['label'] = comp
        _progresso_conc['atual'] += 1
        df = p_data['df'].copy()
        pasta_pdfs = mapa_pastas.get(comp)

        if not pasta_pdfs:
            resultados_audit[comp] = {
                'erro_pasta': True,
                'estatisticas': {},
                'encontrados': [],
                'encontrados_abreviacao': [],
                'erros_nominais': [],
                'nao_encontrados': [],
                'pdfs_extras': [],
                'caminho_planilha': p_data['caminho']
            }
            continue
            
        df_verificado = confinicial.verificar_pdfs(df, pasta_pdfs)
        df_verificado['Status Base'] = df_verificado['Status PDF'].apply(_status_base)
        
        totais = df_verificado['Status Base'].value_counts().to_dict()
        
        estatisticas = {
            'encontrados': totais.get('ENCONTRADO', 0) + totais.get('ENCONTRADO COMO 115', 0) + totais.get('ENCONTRADO COM ABREVIAÇÃO', 0) + totais.get('ENCONTRADO VIA LLM', 0),
            'erros_nominais': totais.get('POSSÍVEL ERRO NOMINAL', 0),
            'nao_encontrados': totais.get('NÃO ENCONTRADO NO .PDF', 0),
            'pdfs_extras': totais.get('PDF NA PASTA, MAS NÃO NA PLANILHA', 0),
            'llm': totais.get('ENCONTRADO VIA LLM', 0)
        }
        
        encontrados = []
        encontrados_abreviacao = []
        erros_nominais = []
        nao_encontrados = []
        pdfs_extras = []
        
        for idx, row in df_verificado.iterrows():
            status = str(row.get('Status Base', ''))
            pis = normalizar_pis(row.get('PIS', ''))
            nome_planilha = str(row.get('NOMES', ''))
            nome_pdf = str(row.get('Nome do Arquivo Encontrado', ''))
            proc = str(row.get('PROC.', ''))
            
            if status == 'POSSÍVEL ERRO NOMINAL':
                erros_nominais.append({'pis': pis, 'nome_planilha': nome_planilha, 'nome_pdf': nome_pdf})
            elif status == 'NÃO ENCONTRADO NO .PDF':
                nao_encontrados.append({'pis': pis, 'nome_planilha': nome_planilha, 'proc': proc})
            elif status == 'DUPLICADO NA PLANILHA':
                # Mesma pessoa (mesmo PIS) digitada 2x na planilha; ja conferida
                # na outra linha. Nao e PDF faltando.
                encontrados.append({'pis': pis, 'nome_planilha': nome_planilha, 'nome_pdf': '(duplicado na planilha)', 'proc': proc, 'detalhe': 'DUPLICADO NA PLANILHA'})
            elif status == 'PDF NA PASTA, MAS NÃO NA PLANILHA':
                pdfs_extras.append({'nome_pdf': nome_pdf})
            elif status in ['ENCONTRADO', 'ENCONTRADO COMO 115', 'ENCONTRADO VIA LLM']:
                encontrados.append({'pis': pis, 'nome_planilha': nome_planilha, 'nome_pdf': nome_pdf, 'proc': proc, 'detalhe': str(row.get('Status PDF', ''))})
            elif status == 'ENCONTRADO COM ABREVIAÇÃO':
                encontrados_abreviacao.append({'pis': pis, 'nome_planilha': nome_planilha, 'nome_pdf': nome_pdf, 'proc': proc, 'detalhe': str(row.get('Status PDF', ''))})

        resultados_audit[comp] = {
            'erro_pasta': False,
            'estatisticas': estatisticas,
            'encontrados': encontrados,
            'encontrados_abreviacao': encontrados_abreviacao,
            'erros_nominais': erros_nominais,
            'nao_encontrados': nao_encontrados,
            'pdfs_extras': pdfs_extras,
            'caminho_planilha': p_data['caminho']
        }

    _progresso_conc['rodando'] = False
    return jsonify({'ok': True, 'resultados': resultados_audit})

# ─── Modo MULTI-ANO ─────────────────────────────────────────────────────
def _status_base_conc(s):
    for prefixo in [
        "NÃO ENCONTRADO NO .PDF", "DUPLICADO NA PLANILHA", "POSSÍVEL ERRO NOMINAL",
        "ENCONTRADO VIA LLM", "ENCONTRADO COMO 115", "ENCONTRADO COM ABREVIAÇÃO",
        "ENCONTRADO",
    ]:
        if str(s).startswith(prefixo):
            return prefixo
    if str(s).startswith("PDF NA PASTA"):
        return "PDF NA PASTA, MAS NÃO NA PLANILHA"
    return str(s)


def _analisar_planilhas(planilhas, pdf_root):
    """Concilia um conjunto de planilhas (por competencia) contra a pasta-raiz
    de PDFs (com subpastas por competencia). Retorna resultados_audit (mesmo
    formato de /api/analisar-conciliacao)."""
    mapa_pastas = {}
    if pdf_root and os.path.isdir(pdf_root):
        for d in os.listdir(pdf_root):
            caminho_dir = os.path.join(pdf_root, d)
            if os.path.isdir(caminho_dir):
                mes, ano = confinicial.extrair_competencia_pasta(d)
                if mes and ano:
                    mapa_pastas[f"{mes}-{ano}"] = caminho_dir

    resultados_audit = {}
    _progresso_conc.update(atual=0, total=len(planilhas), rodando=True, label='')
    try:
        for comp, p_data in planilhas.items():
            _progresso_conc['label'] = comp
            _progresso_conc['atual'] += 1
            df = p_data['df'].copy()
            pasta_pdfs = mapa_pastas.get(comp)
            if not pasta_pdfs:
                resultados_audit[comp] = {
                    'erro_pasta': True, 'estatisticas': {}, 'encontrados': [],
                    'encontrados_abreviacao': [], 'erros_nominais': [],
                    'nao_encontrados': [], 'pdfs_extras': [],
                    'caminho_planilha': p_data['caminho']
                }
                continue

            df_verificado = confinicial.verificar_pdfs(df, pasta_pdfs)
            df_verificado['Status Base'] = df_verificado['Status PDF'].apply(_status_base_conc)
            totais = df_verificado['Status Base'].value_counts().to_dict()
            estatisticas = {
                'encontrados': totais.get('ENCONTRADO', 0) + totais.get('ENCONTRADO COMO 115', 0) + totais.get('ENCONTRADO COM ABREVIAÇÃO', 0) + totais.get('ENCONTRADO VIA LLM', 0),
                'erros_nominais': totais.get('POSSÍVEL ERRO NOMINAL', 0),
                'nao_encontrados': totais.get('NÃO ENCONTRADO NO .PDF', 0),
                'pdfs_extras': totais.get('PDF NA PASTA, MAS NÃO NA PLANILHA', 0),
                'llm': totais.get('ENCONTRADO VIA LLM', 0)
            }
            encontrados, encontrados_abreviacao = [], []
            erros_nominais, nao_encontrados, pdfs_extras = [], [], []
            for idx, row in df_verificado.iterrows():
                status = str(row.get('Status Base', ''))
                pis = normalizar_pis(row.get('PIS', ''))
                nome_planilha = str(row.get('NOMES', ''))
                nome_pdf = str(row.get('Nome do Arquivo Encontrado', ''))
                proc = str(row.get('PROC.', ''))
                if status == 'POSSÍVEL ERRO NOMINAL':
                    erros_nominais.append({'pis': pis, 'nome_planilha': nome_planilha, 'nome_pdf': nome_pdf})
                elif status == 'NÃO ENCONTRADO NO .PDF':
                    nao_encontrados.append({'pis': pis, 'nome_planilha': nome_planilha, 'proc': proc})
                elif status == 'DUPLICADO NA PLANILHA':
                    encontrados.append({'pis': pis, 'nome_planilha': nome_planilha, 'nome_pdf': '(duplicado na planilha)', 'proc': proc, 'detalhe': 'DUPLICADO NA PLANILHA'})
                elif status == 'PDF NA PASTA, MAS NÃO NA PLANILHA':
                    pdfs_extras.append({'nome_pdf': nome_pdf})
                elif status in ['ENCONTRADO', 'ENCONTRADO COMO 115', 'ENCONTRADO VIA LLM']:
                    encontrados.append({'pis': pis, 'nome_planilha': nome_planilha, 'nome_pdf': nome_pdf, 'proc': proc, 'detalhe': str(row.get('Status PDF', ''))})
                elif status == 'ENCONTRADO COM ABREVIAÇÃO':
                    encontrados_abreviacao.append({'pis': pis, 'nome_planilha': nome_planilha, 'nome_pdf': nome_pdf, 'proc': proc, 'detalhe': str(row.get('Status PDF', ''))})
            resultados_audit[comp] = {
                'erro_pasta': False, 'estatisticas': estatisticas,
                'encontrados': encontrados, 'encontrados_abreviacao': encontrados_abreviacao,
                'erros_nominais': erros_nominais, 'nao_encontrados': nao_encontrados,
                'pdfs_extras': pdfs_extras, 'caminho_planilha': p_data['caminho']
            }
    finally:
        _progresso_conc['rodando'] = False
    return resultados_audit


def _carregar_planilhas_pasta(excel_dir):
    """Carrega todas as planilhas (.xls/.xlsx/.ods) de uma pasta, por competencia."""
    planilhas = {}
    if not excel_dir or not os.path.isdir(excel_dir):
        return planilhas
    for f in sorted(os.listdir(excel_dir)):
        if not f.lower().endswith(('.xls', '.xlsx', '.ods')):
            continue
        mes, ano = confinicial.extrair_competencia_excel(f)
        if not mes:
            continue
        cam = os.path.join(excel_dir, f)
        try:
            df = confinicial.ler_planilha_fgts(cam)
        except Exception:
            df = None
        if df is None or df.empty:
            continue
        planilhas[f"{mes}-{ano}"] = {'caminho': cam, 'nome_original': f, 'df': df}
    return planilhas


def _detectar_anos(parent):
    """Detecta as pastas de ANO em `parent`. Cada ano tem 3 subpastas:
    'ANO PDF', 'ANO EXCEL', 'ANO CONFERENCIA'. Retorna {ano: {pdf,excel,conferencia,nome}}."""
    anos = {}
    if not parent or not os.path.isdir(parent):
        return anos
    for d in sorted(os.listdir(parent)):
        cam = os.path.join(parent, d)
        if not os.path.isdir(cam):
            continue
        sub = {}
        try:
            for s in os.listdir(cam):
                p = os.path.join(cam, s)
                if os.path.isdir(p):
                    sub[s.upper()] = p
        except Exception:
            continue
        pdf_dir = next((v for k, v in sub.items() if k.endswith('PDF')), None)
        excel_dir = next((v for k, v in sub.items() if 'EXCEL' in k), None)
        conf_dir = next((v for k, v in sub.items() if 'CONFEREN' in k), None)
        if pdf_dir and excel_dir:
            m = re.search(r'(19|20)\d{2}', d) or re.search(r'(19|20)\d{2}', os.path.basename(pdf_dir))
            ano = m.group(0) if m else d
            anos[ano] = {'nome': d, 'pdf': pdf_dir, 'excel': excel_dir, 'conferencia': conf_dir}
    return anos


@app.route('/api/selecionar-pasta-anos', methods=['POST'])
def selecionar_pasta_anos():
    try:
        dlg = tk.Tk(); dlg.withdraw(); dlg.attributes('-topmost', True)
        parent = filedialog.askdirectory(title="Selecione a pasta que contem os ANOS (2007, 2008, ...)")
        dlg.destroy()
    except Exception as e:
        return jsonify({'ok': False, 'erro': f'Erro ao abrir seletor: {str(e)}'})
    if not parent:
        return jsonify({'ok': False, 'erro': 'Nenhuma pasta selecionada'})
    anos = _detectar_anos(parent)
    _state['anos'] = anos
    _state['pasta_anos'] = parent
    lista = []
    for ano in sorted(anos):
        info = anos[ano]
        try:
            n_xls = len([f for f in os.listdir(info['excel']) if f.lower().endswith(('.xls', '.xlsx', '.ods'))])
        except Exception:
            n_xls = 0
        try:
            n_pdfsub = len([d for d in os.listdir(info['pdf']) if os.path.isdir(os.path.join(info['pdf'], d))])
        except Exception:
            n_pdfsub = 0
        lista.append({'ano': ano, 'nome': info['nome'], 'excel': n_xls, 'pdf_pastas': n_pdfsub})
    return jsonify({'ok': True, 'pasta': parent, 'anos': lista})


@app.route('/api/analisar-ano', methods=['POST'])
def analisar_ano():
    ano = str(request.json.get('ano', ''))
    info = _state.get('anos', {}).get(ano)
    if not info:
        return jsonify({'ok': False, 'erro': f'Ano {ano} nao carregado. Selecione a pasta dos anos primeiro.'})
    planilhas = _carregar_planilhas_pasta(info['excel'])
    if not planilhas:
        return jsonify({'ok': False, 'erro': f'Nenhuma planilha valida em "{os.path.basename(info["excel"])}".'})
    resultados = _analisar_planilhas(planilhas, info['pdf'])
    return jsonify({'ok': True, 'ano': ano, 'resultados': resultados})


@app.route('/api/gerar-ano', methods=['POST'])
def gerar_ano():
    """Salva e organiza um ANO: APAGA tudo na pasta 'ANO CONFERENCIA' e gera de
    novo (relatorios .xlsx + organizacao dos PDFs em 115/660), por competencia."""
    ano = str(request.json.get('ano', ''))
    info = _state.get('anos', {}).get(ano)
    if not info:
        return jsonify({'ok': False, 'erro': f'Ano {ano} nao carregado. Selecione a pasta dos anos primeiro.'})

    # Pasta de saida = "ANO CONFERENCIA" (dentro da pasta do ano)
    year_dir = os.path.dirname(info['pdf'])
    destino = info.get('conferencia') or os.path.join(year_dir, f"{ano} CONFERENCIA")
    # Trava de seguranca: so mexe se o nome contiver CONFERENCIA (evita apagar
    # a pasta errada por acidente).
    if 'CONFEREN' not in os.path.basename(destino).upper():
        return jsonify({'ok': False, 'erro': 'Pasta de conferencia invalida (nome nao contem CONFERENCIA).'})

    planilhas = _carregar_planilhas_pasta(info['excel'])
    if not planilhas:
        return jsonify({'ok': False, 'erro': f'Nenhuma planilha valida em "{os.path.basename(info["excel"])}".'})

    # APAGA tudo na pasta CONFERENCIA e recria vazia
    try:
        if os.path.isdir(destino):
            shutil.rmtree(destino)
        os.makedirs(destino, exist_ok=True)
    except Exception as e:
        return jsonify({'ok': False, 'erro': f'Falha ao limpar a pasta CONFERENCIA: {e}'})

    # Mapa de pastas de PDF por competencia
    mapa_pastas = {}
    for d in os.listdir(info['pdf']):
        cam = os.path.join(info['pdf'], d)
        if os.path.isdir(cam):
            mes, a = confinicial.extrair_competencia_pasta(d)
            if mes and a:
                mapa_pastas[f"{mes}-{a}"] = cam

    resultados = []
    _progresso_conc.update(atual=0, total=len(planilhas), rodando=True, label='')
    try:
        for comp, p_data in planilhas.items():
            _progresso_conc['label'] = comp
            _progresso_conc['atual'] += 1
            df = p_data['df'].copy()
            pasta_pdfs = mapa_pastas.get(comp)
            if not pasta_pdfs:
                df['Status PDF'] = "PASTA NÃO ENCONTRADA"
                df['Nome do Arquivo Encontrado'] = ""
                df_final = df
                total_pdfs = 0
            else:
                total_pdfs = len([f for f in os.listdir(pasta_pdfs) if f.lower().endswith(('.pdf', '.tif', '.tiff'))])
                df_final = confinicial.verificar_pdfs(df, pasta_pdfs)
                confinicial.organizar_pdfs_por_resultado(df_final, pasta_pdfs, destino)
            nome_saida = f"Conciliacao_{comp}.xlsx"
            try:
                confinicial.salvar_com_resumo(df_final, os.path.join(destino, nome_saida), total_pdfs=total_pdfs)
                resultados.append(f"Gerado: {nome_saida}")
            except Exception as e:
                resultados.append(f"Erro ao salvar {nome_saida}: {str(e)}")
    finally:
        _progresso_conc['rodando'] = False

    return jsonify({'ok': True, 'ano': ano, 'destino': destino, 'resultados': resultados})

@app.route('/api/progresso-conciliacao', methods=['GET'])
def progresso_conciliacao():
    p = dict(_progresso_conc)
    p['pct'] = round(p['atual'] / p['total'] * 100) if p['total'] else 0
    return jsonify(p)

@app.route('/api/buscar-planilhas', methods=['POST'])
def buscar_planilhas():
    termo = normalizar(request.json.get('termo', ''))
    if not termo or len(termo) < 3:
        return jsonify({'ok': False, 'erro': 'Digite pelo menos 3 caracteres'})
        
    if not _state['planilhas']:
        return jsonify({'ok': False, 'erro': 'Nenhuma planilha carregada. Carregue as planilhas na aba de Conciliação primeiro.'})
        
    # Verify if term is digits only (PIS search)
    termo_pis = normalizar_pis(termo) if len(re.sub(r'\D', '', termo)) >= 10 else None
    
    resultados_dict = {}
    
    # Palavras a ignorar (linhas de totais, cabeçalhos etc.)
    IGNORAR_NOMES = {
        'NOMES', 'NOME', 'TITULAR', 'FUNCIONARIO', 'PIS', 'PROC', 'DATA', 'OBS',
        'TOTAL', 'TOTAIS', 'SUBTOTAL', 'TOTAL GERAL', 'TOTALGERAL', 'GERAL',
        'TOTAL- GERAL', 'TOTAL-GERAL',
    }
    
    for comp, p_data in _state['planilhas'].items():
        df = p_data['df']
        for idx, row in df.iterrows():
            pis = normalizar_pis(row.get('PIS', ''))
            nome_raw = str(row.get('NOMES', ''))
            nome_norm = normalizar(nome_raw)
            
            # Ignora linhas de totais, cabeçalhos e nomes inválidos
            if not nome_norm or len(nome_norm) <= 4:
                continue
            if nome_norm in IGNORAR_NOMES:
                continue
            # Ignora se não tiver espaço (não é nome completo)
            if ' ' not in nome_norm:
                continue
            # Ignora se contiver palavras de total
            if any(p in nome_norm for p in ['TOTAL', 'GERAL', 'SUBTOTAL']):
                continue
            
            match = False
            if termo_pis and pis == termo_pis:
                match = True
            elif not termo_pis and (termo in nome_norm or similaridade(termo, nome_norm) > 0.6):
                match = True
                
            if match:
                chave = f"{pis}::{nome_raw}"
                if chave not in resultados_dict:
                    resultados_dict[chave] = {
                        'pis': pis,
                        'nome_atual': nome_raw,
                        'proc': '',
                        'competencias': []
                    }
                # Processo (PROC.) da planilha — pega o primeiro valor valido
                if not resultados_dict[chave]['proc']:
                    for _k in ('PROC.', 'Processo', 'PROC', 'PROCESSO'):
                        _v = row.get(_k, '')
                        _s = str(_v).strip()
                        if _s and _s.lower() not in ('nan', 'none'):
                            if _s.endswith('.0'):
                                _s = _s[:-2]
                            resultados_dict[chave]['proc'] = _s
                            break
                if comp not in resultados_dict[chave]['competencias']:
                    resultados_dict[chave]['competencias'].append(comp)
                    
    resultados = list(resultados_dict.values())
    # Sort so that items appearing in more spreadsheets come first
    resultados.sort(key=lambda x: -len(x['competencias']))
    
    return jsonify({'ok': True, 'resultados': resultados})

@app.route('/api/corrigir-xls-fisico', methods=['POST'])
def corrigir_xls_fisico():
    data = request.json
    pis_alvo = normalizar_pis(data.get('pis'))
    novo_nome = str(data.get('novo_nome', '')).strip().upper()
    
    if not pis_alvo or not novo_nome:
        return jsonify({'ok': False, 'erro': 'PIS ou Novo Nome faltando.'})
        
    arquivos_alterados = 0
    linhas_alteradas = 0
    avisos = []
    
    import pythoncom
    pythoncom.CoInitialize()

    # Helper functions
    def arquivo_em_uso(caminho):
        """Verifica se o arquivo está aberto/bloqueado (ex: no Excel) antes de tentar editar via COM."""
        try:
            with open(caminho, 'r+b'):
                pass
            return False
        except (PermissionError, OSError):
            return True

    def get_correcoes_xls(caminho):
        indices = []
        try:
            wb = xlrd.open_workbook(caminho)
            try:
                sheet = wb.sheet_by_name('FGTS EM ATRASO - PROCESSOS')
            except:
                sheet = wb.sheet_by_index(0)
            for row in range(sheet.nrows):
                try:
                    val_pis = normalizar_pis(sheet.cell_value(row, 2))
                    if not val_pis:
                        val_pis = normalizar_pis(sheet.cell_value(row, 1))
                except:
                    val_pis = ""
                if val_pis == pis_alvo:
                    try:
                        val_nome = str(sheet.cell_value(row, 3)).strip().upper()
                        if val_nome != novo_nome:
                            indices.append(row)
                    except:
                        pass
        except Exception as e:
            print(f"Erro xlrd: {e}")
        return indices
        
    # Varre TODAS as planilhas carregadas
    for comp, p_data in _state['planilhas'].items():
        caminho_planilha = p_data['caminho']
        if not os.path.exists(caminho_planilha):
            continue
            
        ext = os.path.splitext(caminho_planilha)[1].lower()
        alterados_nesta = 0
        
        if ext == '.xls':
            indices = get_correcoes_xls(caminho_planilha)
            if indices and arquivo_em_uso(caminho_planilha):
                avisos.append(f'"{p_data["nome_original"]}" está aberto no Excel — feche o arquivo e tente novamente.')
            elif indices:
                import win32com.client
                xl = win32com.client.Dispatch('Excel.Application')
                try:
                    xl.Application.DisplayAlerts = False
                    xl.Visible = False
                    abs_path = os.path.abspath(caminho_planilha)
                    wb = xl.Workbooks.Open(abs_path)
                    try:
                        ws = wb.Sheets('FGTS EM ATRASO - PROCESSOS')
                    except:
                        ws = wb.Sheets(1)
                        
                    for r in indices:
                        ws.Cells(r + 1, 4).Value = novo_nome
                        alterados_nesta += 1
                        
                    wb.Close(SaveChanges=True)
                except Exception as e:
                    import traceback
                    print(f"Erro COM win32 ao editar {caminho_planilha}: {e}")
                    traceback.print_exc()
                finally:
                    try:
                        xl.Quit()
                    except Exception:
                        pass
                    
        elif ext == '.ods':
            # Support for .ods using odfpy
            def _ods_set_cell_value(rows, row_idx, col_idx, value):
                if row_idx >= len(rows): return
                cells = rows[row_idx].getElementsByType(TableCell)
                while len(cells) <= col_idx:
                    new_cell = TableCell()
                    rows[row_idx].addElement(new_cell)
                    cells = rows[row_idx].getElementsByType(TableCell)
                cell = cells[col_idx]
                for old_p in cell.getElementsByType(P):
                    cell.removeChild(old_p)
                cell.setAttribute('valuetype', 'string')
                p = P(text=value)
                cell.addElement(p)
                
            try:
                # Use pandas to find indices quickly
                df_ods = pd.read_excel(caminho_planilha, sheet_name=None, header=None, engine='odf')
                aba_nome = 'FGTS EM ATRASO - PROCESSOS' if 'FGTS EM ATRASO - PROCESSOS' in df_ods else list(df_ods.keys())[0]
                aba = df_ods[aba_nome]
                indices_ods = []
                for i in range(len(aba)):
                    try:
                        val_pis = normalizar_pis(aba.iloc[i, 2])
                        if not val_pis:
                            val_pis = normalizar_pis(aba.iloc[i, 1])
                    except:
                        val_pis = ""
                    if val_pis == pis_alvo:
                        try:
                            val_nome = str(aba.iloc[i, 3]).strip().upper()
                            if val_nome != novo_nome:
                                indices_ods.append(i) # 0-indexed pandas matches doc rows generally well enough in this context, but better to just use df
                        except:
                            pass
                            
                if indices_ods:
                    doc = opendocument.load(caminho_planilha)
                    tables = doc.spreadsheet.getElementsByType(Table)
                    table = None
                    for t in tables:
                        if t.getAttribute('name') == 'FGTS EM ATRASO - PROCESSOS':
                            table = t
                            break
                    if table is None and tables:
                        table = tables[0]
                        
                    if table:
                        rows = table.getElementsByType(TableRow)
                        for idx in indices_ods:
                            _ods_set_cell_value(rows, idx, 3, novo_nome)
                            alterados_nesta += 1
                        doc.save(caminho_planilha)
            except Exception as e:
                print(f"Erro ao salvar ods físico: {e}")
                
        else:
            try:
                wb = load_workbook(caminho_planilha)
                try:
                    ws = wb['FGTS EM ATRASO - PROCESSOS']
                except:
                    ws = wb.active
                    
                for row in range(1, ws.max_row + 1):
                    val_pis = normalizar_pis(ws.cell(row=row, column=3).value)
                    if not val_pis:
                        val_pis = normalizar_pis(ws.cell(row=row, column=2).value)
                        
                    if val_pis == pis_alvo:
                        try:
                            val_nome = str(ws.cell(row=row, column=4).value).strip().upper()
                            if val_nome != novo_nome:
                                ws.cell(row=row, column=4).value = novo_nome
                                alterados_nesta += 1
                        except:
                            pass
                            
                if alterados_nesta > 0:
                    wb.save(caminho_planilha)
            except Exception as e:
                print(f"Erro openpyxl: {e}")
                
        if alterados_nesta > 0:
            arquivos_alterados += 1
            linhas_alteradas += alterados_nesta
            novo_df = confinicial.ler_planilha_fgts(caminho_planilha)
            if novo_df is not None:
                _state['planilhas'][comp]['df'] = novo_df

    pythoncom.CoUninitialize()
    
    return jsonify({
        'ok': True,
        'arquivos_alterados': arquivos_alterados,
        'linhas_alteradas': linhas_alteradas,
        'avisos': avisos
    })

@app.route('/api/corrigir-pis-fisico', methods=['POST'])
def corrigir_pis_fisico():
    data = request.json
    nome_alvo = str(data.get('nome', '')).strip().upper()
    pis_antigo = normalizar_pis(data.get('pis_antigo'))
    novo_pis = normalizar_pis(data.get('novo_pis'))
    
    if not nome_alvo or not novo_pis:
        return jsonify({'ok': False, 'erro': 'Nome ou Novo PIS faltando.'})
        
    arquivos_alterados = 0
    linhas_alteradas = 0
    avisos = []
    
    import pythoncom
    pythoncom.CoInitialize()

    # Helper functions
    def arquivo_em_uso(caminho):
        """Verifica se o arquivo está aberto/bloqueado (ex: no Excel) antes de tentar editar via COM."""
        try:
            with open(caminho, 'r+b'):
                pass
            return False
        except (PermissionError, OSError):
            return True

    def get_correcoes_xls(caminho):
        indices_e_colunas = []
        try:
            wb = xlrd.open_workbook(caminho)
            try:
                sheet = wb.sheet_by_name('FGTS EM ATRASO - PROCESSOS')
            except:
                sheet = wb.sheet_by_index(0)
            for row in range(sheet.nrows):
                try:
                    val_nome = str(sheet.cell_value(row, 3))
                except:
                    val_nome = ""
                if normalizar(val_nome) == normalizar(nome_alvo):
                    val_pis_c = normalizar_pis(sheet.cell_value(row, 2))
                    val_pis_b = normalizar_pis(sheet.cell_value(row, 1))
                    if val_pis_c == pis_antigo and val_pis_c != novo_pis:
                        indices_e_colunas.append((row, 3)) # coluna C (1-indexed em win32com é 3)
                    elif val_pis_b == pis_antigo and val_pis_b != novo_pis:
                        indices_e_colunas.append((row, 2)) # coluna B
                    elif not pis_antigo and val_pis_c != novo_pis:
                        indices_e_colunas.append((row, 3))
        except Exception as e:
            print(f"Erro xlrd em PIS: {e}")
        return indices_e_colunas

    # Varre TODAS as planilhas carregadas
    for comp, p_data in _state['planilhas'].items():
        caminho_planilha = p_data['caminho']
        if not os.path.exists(caminho_planilha):
            continue
            
        ext = os.path.splitext(caminho_planilha)[1].lower()
        alterados_nesta = 0
        
        if ext == '.xls':
            itens = get_correcoes_xls(caminho_planilha)
            if itens and arquivo_em_uso(caminho_planilha):
                avisos.append(f'"{p_data["nome_original"]}" está aberto no Excel — feche o arquivo e tente novamente.')
            elif itens:
                import win32com.client
                xl = win32com.client.Dispatch('Excel.Application')
                try:
                    xl.Application.DisplayAlerts = False
                    xl.Visible = False
                    abs_path = os.path.abspath(caminho_planilha)
                    wb = xl.Workbooks.Open(abs_path)
                    try:
                        ws = wb.Sheets('FGTS EM ATRASO - PROCESSOS')
                    except:
                        ws = wb.Sheets(1)
                        
                    for r, col in itens:
                        ws.Cells(r + 1, col).Value = novo_pis
                        alterados_nesta += 1
                        
                    wb.Close(SaveChanges=True)
                except Exception as e:
                    import traceback
                    print(f"Erro COM win32 ao editar PIS {caminho_planilha}: {e}")
                    traceback.print_exc()
                finally:
                    try:
                        xl.Quit()
                    except Exception:
                        pass
                    
        elif ext == '.ods':
            # Support for .ods using odfpy
            def _ods_set_cell_value(rows, row_idx, col_idx, value):
                if row_idx >= len(rows): return
                cells = rows[row_idx].getElementsByType(TableCell)
                while len(cells) <= col_idx:
                    new_cell = TableCell()
                    rows[row_idx].addElement(new_cell)
                    cells = rows[row_idx].getElementsByType(TableCell)
                cell = cells[col_idx]
                for old_p in cell.getElementsByType(P):
                    cell.removeChild(old_p)
                cell.setAttribute('valuetype', 'string')
                p = P(text=value)
                cell.addElement(p)
                
            try:
                df_ods = pd.read_excel(caminho_planilha, sheet_name=None, header=None, engine='odf')
                aba_nome = 'FGTS EM ATRASO - PROCESSOS' if 'FGTS EM ATRASO - PROCESSOS' in df_ods else list(df_ods.keys())[0]
                aba = df_ods[aba_nome]
                indices_ods = []
                for i in range(len(aba)):
                    try:
                        val_nome = str(aba.iloc[i, 3])
                    except:
                        val_nome = ""
                    if normalizar(val_nome) == normalizar(nome_alvo):
                        val_pis_c = normalizar_pis(aba.iloc[i, 2])
                        val_pis_b = normalizar_pis(aba.iloc[i, 1])
                        if val_pis_c == pis_antigo and val_pis_c != novo_pis:
                            indices_ods.append((i, 2))
                        elif val_pis_b == pis_antigo and val_pis_b != novo_pis:
                            indices_ods.append((i, 1))
                        elif not pis_antigo and val_pis_c != novo_pis:
                            indices_ods.append((i, 2))
                            
                if indices_ods:
                    doc = opendocument.load(caminho_planilha)
                    tables = doc.spreadsheet.getElementsByType(Table)
                    table = None
                    for t in tables:
                        if t.getAttribute('name') == 'FGTS EM ATRASO - PROCESSOS':
                            table = t
                            break
                    if table is None and tables:
                        table = tables[0]
                        
                    if table:
                        rows = table.getElementsByType(TableRow)
                        for idx, col_idx in indices_ods:
                            _ods_set_cell_value(rows, idx, col_idx, novo_pis)
                            alterados_nesta += 1
                        doc.save(caminho_planilha)
            except Exception as e:
                print(f"Erro ao salvar ods físico: {e}")
                
        else:
            try:
                wb = load_workbook(caminho_planilha)
                try:
                    ws = wb['FGTS EM ATRASO - PROCESSOS']
                except:
                    ws = wb.active
                    
                for row in range(1, ws.max_row + 1):
                    try:
                        val_nome = str(ws.cell(row=row, column=4).value)
                    except:
                        val_nome = ""
                    if normalizar(val_nome) == normalizar(nome_alvo):
                        val_pis_c = normalizar_pis(ws.cell(row=row, column=3).value)
                        val_pis_b = normalizar_pis(ws.cell(row=row, column=2).value)
                        
                        if val_pis_c == pis_antigo and val_pis_c != novo_pis:
                            ws.cell(row=row, column=3).value = novo_pis
                            alterados_nesta += 1
                        elif val_pis_b == pis_antigo and val_pis_b != novo_pis:
                            ws.cell(row=row, column=2).value = novo_pis
                            alterados_nesta += 1
                        elif not pis_antigo and val_pis_c != novo_pis:
                            ws.cell(row=row, column=3).value = novo_pis
                            alterados_nesta += 1
                            
                if alterados_nesta > 0:
                    wb.save(caminho_planilha)
            except Exception as e:
                print(f"Erro openpyxl: {e}")
                
        if alterados_nesta > 0:
            arquivos_alterados += 1
            linhas_alteradas += alterados_nesta
            novo_df = confinicial.ler_planilha_fgts(caminho_planilha)
            if novo_df is not None:
                _state['planilhas'][comp]['df'] = novo_df

    pythoncom.CoUninitialize()
    
    return jsonify({
        'ok': True,
        'arquivos_alterados': arquivos_alterados,
        'linhas_alteradas': linhas_alteradas,
        'avisos': avisos
    })

@app.route('/api/renomear-pdf-conciliacao', methods=['POST'])
def renomear_pdf_conciliacao():
    data = request.json
    comp = data.get('competencia')
    nome_antigo = data.get('nome_antigo')
    nome_novo = data.get('nome_novo')
    
    if not comp or not nome_antigo or not nome_novo:
        return jsonify({'ok': False, 'erro': 'Dados incompletos para renomear.'})
        
    if not _state['pasta_raiz']:
        return jsonify({'ok': False, 'erro': 'Pasta raiz de PDFs não configurada.'})
        
    # Encontra a subpasta correta
    pasta_pdfs = None
    for d in os.listdir(_state['pasta_raiz']):
        caminho_dir = os.path.join(_state['pasta_raiz'], d)
        if os.path.isdir(caminho_dir):
            m, a = confinicial.extrair_competencia_pasta(d)
            if f"{m}-{a}" == comp:
                pasta_pdfs = caminho_dir
                break
                
    if not pasta_pdfs:
        return jsonify({'ok': False, 'erro': f'Pasta de PDFs para a competência {comp} não encontrada.'})
        
    caminho_antigo = os.path.join(pasta_pdfs, nome_antigo + '.pdf')

    # Se o nome antigo nao existir com .pdf, procura com .tif/.tiff
    ext_real = '.pdf'
    if not os.path.exists(caminho_antigo):
        caminho_antigo = os.path.join(pasta_pdfs, nome_antigo + '.tif')
        ext_real = '.tif'
    if not os.path.exists(caminho_antigo):
        caminho_antigo = os.path.join(pasta_pdfs, nome_antigo + '.tiff')
        ext_real = '.tiff'
    if not os.path.exists(caminho_antigo):
        for f in os.listdir(pasta_pdfs):
            if os.path.splitext(f)[0] == nome_antigo:
                caminho_antigo = os.path.join(pasta_pdfs, f)
                ext_real = os.path.splitext(f)[1]
                break
        else:
            return jsonify({'ok': False, 'erro': 'Arquivo PDF antigo não encontrado na pasta física.'})

    caminho_novo = os.path.join(pasta_pdfs, nome_novo + ext_real)

    if os.path.exists(caminho_novo):
        return jsonify({'ok': False, 'erro': 'Já existe um arquivo com este novo nome na pasta.'})

    try:
        os.rename(caminho_antigo, caminho_novo)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'erro': str(e)})

@app.route('/api/gerar-conciliacao', methods=['POST'])
def gerar_conciliacao():
    if not _state['pasta_raiz']:
        return jsonify({'ok': False, 'erro': 'Selecione a Pasta Raiz de PDFs na aba de Configuração primeiro.'})
    if not _state['planilhas']:
        return jsonify({'ok': False, 'erro': 'Nenhuma planilha carregada.'})
        
    root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True)
    pasta_destino = filedialog.askdirectory(title="Selecione a PASTA DE DESTINO para salvar os relatórios e organizar")
    root.destroy()
    
    if not pasta_destino:
        return jsonify({'ok': False, 'erro': 'Geração cancelada (nenhuma pasta destino selecionada).'})
        
    # Mapear as subpastas da raiz por competencia
    mapa_pastas = {}
    for d in os.listdir(_state['pasta_raiz']):
        caminho_dir = os.path.join(_state['pasta_raiz'], d)
        if os.path.isdir(caminho_dir):
            mes, ano = confinicial.extrair_competencia_pasta(d)
            if mes and ano:
                mapa_pastas[f"{mes}-{ano}"] = caminho_dir

    resultados = []
    _progresso_conc.update(atual=0, total=len(_state['planilhas']), rodando=True, label='')

    for comp, p_data in _state['planilhas'].items():
        _progresso_conc['label'] = comp
        _progresso_conc['atual'] += 1
        df = p_data['df'].copy()
        nome_orig = p_data['nome_original']

        pasta_pdfs = mapa_pastas.get(comp)
        if not pasta_pdfs:
            # Planilha sem pasta correspondente
            df['Status PDF'] = "PASTA NÃO ENCONTRADA"
            df['Nome do Arquivo Encontrado'] = ""
            df_final = df
            total_pdfs = 0
        else:
            total_pdfs = len([f for f in os.listdir(pasta_pdfs) if f.lower().endswith(('.pdf', '.tif', '.tiff'))])
            # Executa a verificação principal
            df_final = confinicial.verificar_pdfs(df, pasta_pdfs)
            # Organiza os arquivos
            confinicial.organizar_pdfs_por_resultado(df_final, pasta_pdfs, pasta_destino)
            
        # Salva o arquivo de relatorio
        nome_saida = f"Conciliacao_{comp}.xlsx"
        caminho_saida = os.path.join(pasta_destino, nome_saida)
        try:
            confinicial.salvar_com_resumo(df_final, caminho_saida, total_pdfs=total_pdfs)
            resultados.append(f"Gerado: {nome_saida}")
        except Exception as e:
            resultados.append(f"Erro ao salvar {nome_saida}: {str(e)}")
            
    # Ao final da geracao, limpa o estado
    _progresso_conc['rodando'] = False
    _state['planilhas'] = {}
    return jsonify({'ok': True, 'resultados': resultados, 'destino': pasta_destino})

if __name__ == '__main__':
    print('=' * 50)
    print('  Gerenciador de PDFs')
    print('  Acesse: http://localhost:5000')
    print('=' * 50)
    try:
        # Servidor de producao leve (Python puro, roda no Windows).
        # Remove o aviso "development server" do Flask e e mais robusto.
        from waitress import serve
        serve(app, host='127.0.0.1', port=5000, threads=8)
    except ImportError:
        # waitress nao instalado -> usa o servidor embutido do Flask.
        # Funciona igual; so mostra o aviso de "development server".
        # Para remover o aviso: pip install waitress
        app.run(debug=False, port=5000, use_reloader=False, threaded=True)
