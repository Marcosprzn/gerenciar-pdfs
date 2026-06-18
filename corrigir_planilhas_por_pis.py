#!/usr/bin/env python3
"""
Corrige nomes em multiplas planilhas FGTS usando uma planilha REFERENCIA ja corrigida.
A correcao e feita cruzando pelo PIS de cada pessoa.
Converte .xls para .xlsx automaticamente (mantendo 100% da formatacao via Excel).
"""
import os, re, datetime, math
import tkinter as tk
from tkinter import filedialog

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# ============================================================
# NORMALIZACAO DE PIS
# ============================================================

def normalizar_pis(valor):
    if valor is None:
        return ''
    if isinstance(valor, float):
        if math.isnan(valor) or valor == 0:
            return ''
        return str(int(valor)).zfill(11)
    texto = str(valor).strip()
    if not texto or texto == '0':
        return ''
    digitos = re.sub(r'\D', '', texto)
    return digitos.zfill(11)


# ============================================================
# CONVERSAO .xls -> .xlsx
# ============================================================

def converter_xls_para_xlsx(caminho_xls):
    """Converte .xls para .xlsx usando o Excel (win32com). 100% identico.
    Retorna o caminho do .xlsx gerado, ou None se falhar."""
    caminho_xlsx = os.path.splitext(caminho_xls)[0] + '.xlsx'
    if os.path.exists(caminho_xlsx):
        os.remove(caminho_xlsx)

    try:
        import win32com.client, pythoncom
        pythoncom.CoInitialize()
        xl = win32com.client.Dispatch('Excel.Application')
        try:
            xl.DisplayAlerts = False
        except:
            pass
        wb = xl.Workbooks.Open(caminho_xls)
        wb.SaveAs(caminho_xlsx, FileFormat=51)  # 51 = xlOpenXMLWorkbook
        wb.Close()
        xl.Quit()
        pythoncom.CoUninitialize()
        return caminho_xlsx
    except Exception as e:
        try:
            pythoncom.CoUninitialize()
        except:
            pass
        # Fallback: conversao via pandas (perde formatacao, 1 aba)
        print(f'     (Excel indisponivel, convertendo com pandas: {e})')
        try:
            import xlrd
            rb = xlrd.open_workbook(caminho_xls)
            nome_aba = 'FGTS EM ATRASO - PROCESSOS'
            if nome_aba not in rb.sheet_names():
                nome_aba = rb.sheet_names()[0]
            df_raw = pd.read_excel(caminho_xls, sheet_name=nome_aba, header=None)
            df_raw.to_excel(caminho_xlsx, sheet_name='FGTS EM ATRASO - PROCESSOS', index=False, header=False)
            return caminho_xlsx
        except:
            return None


# ============================================================
# DIALOGOS
# ============================================================

def dialogo_arquivo(titulo, tipos):
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    root.update()
    arquivo = filedialog.askopenfilename(title=titulo, filetypes=tipos)
    root.destroy()
    return arquivo

def dialogo_arquivos(titulo, tipos):
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    root.update()
    arquivos = filedialog.askopenfilenames(title=titulo, filetypes=tipos)
    root.destroy()
    return arquivos


# ============================================================
# LEITURA DA PLANILHA
# ============================================================

def ler_nomes_por_pis(caminho):
    """Le a planilha e retorna [(indice_linha, pis, nome)] para nomes validos."""
    if not caminho or not os.path.exists(caminho):
        return []
    ext = os.path.splitext(caminho)[1].lower()
    engine_pd = 'odf' if ext == '.ods' else None
    try:
        df_raw = pd.read_excel(
            caminho, sheet_name='FGTS EM ATRASO - PROCESSOS',
            header=None, engine=engine_pd
        )
    except Exception as e:
        print(f'  ERRO ao ler {os.path.basename(caminho)}: {e}')
        return []

    lista = []
    encontrou = False
    gap = 0
    for i in range(len(df_raw)):
        nome = df_raw.iloc[i, 3]
        if not isinstance(nome, str) or not nome.strip():
            if encontrou:
                gap += 1
                if gap > 40:
                    break
            continue
        nome = nome.strip().upper()
        ignorar = ['NOMES', 'NOME', 'TITULAR', 'FUNCIONARIO', 'PIS', 'PROC.', 'DATA', 'OBS', 'TOTAL', 'TOTAIS', 'SUBTOTAL']
        if nome in ignorar:
            continue
        if re.match(r"^[A-ZÁÀÂÃÉÊÍÓÔÕÚÜÇ0-9 \.\-]+$", nome) and ' ' in nome and len(nome) > 5:
            encontrou = True
            gap = 0
            pis = normalizar_pis(df_raw.iloc[i, 2])
            lista.append((i, pis, nome))
    return lista


# ============================================================
# MAIN
# ============================================================

def processar():
    print('=' * 60)
    print('  CORRECAO DE PLANILHAS POR PIS')
    print('=' * 60)
    print()
    print('Selecione a planilha REFERENCIA (ja corrigida manualmente)')
    print()

    ref_path = dialogo_arquivo(
        'Selecione a planilha REFERENCIA (corrigida)',
        [('Planilhas', '*.xls *.xlsx *.ods')]
    )
    if not ref_path:
        print('Nenhuma planilha selecionada.')
        return

    print(f'Referencia: {os.path.basename(ref_path)}')
    print(f'\nLendo referencia...')
    lista_ref = ler_nomes_por_pis(ref_path)
    if not lista_ref:
        print('Nao foi possivel ler a planilha referencia.')
        return

    mapa_pis = {}
    for idx, pis, nome in lista_ref:
        if pis and pis not in mapa_pis:
            mapa_pis[pis] = nome

    print(f'  {len(lista_ref)} nomes lidos, {len(mapa_pis)} PIS unicos')

    # Seleciona alvos
    print(f'\nSelecione as planilhas para CORRIGIR')
    alvos = dialogo_arquivos(
        'Selecione as planilhas para CORRIGIR',
        [('Planilhas', '*.xls *.xlsx *.ods')]
    )
    if not alvos:
        print('Nenhuma planilha selecionada.')
        return
    alvos = [a for a in alvos if os.path.normpath(a) != os.path.normpath(ref_path)]
    if not alvos:
        print('Nenhuma planilha para corrigir.')
        return

    print(f'  {len(alvos)} planilha(s) para corrigir')
    for a in alvos:
        print(f'    - {os.path.basename(a)}')

    print(f'\nATENCAO: Os arquivos serao convertidos para .xlsx e corrigidos!')
    resp = input('Continuar? (S/n): ').strip().lower()
    if resp == 'n':
        print('Operacao cancelada.')
        return

    # Log
    log_lines = [
        f'Correcao de planilhas por PIS - {datetime.datetime.now().strftime("%d/%m/%Y %H:%M")}',
        f'Referencia: {ref_path}',
        '', '=' * 70, ''
    ]
    total_geral_corrigidos = 0
    arquivos_gerados = []

    for alvo in alvos:
        nome_arq = os.path.basename(alvo)
        ext = os.path.splitext(alvo)[1].lower()
        print(f'\n--- Processando: {nome_arq} ---')

        # Converte .xls ou .ods para .xlsx
        caminho_xlsx = None
        if ext == '.xls':
            print(f'  Convertendo para .xlsx...')
            caminho_xlsx = converter_xls_para_xlsx(alvo)
            if not caminho_xlsx:
                print(f'  ERRO ao converter {nome_arq}')
                continue
        elif ext == '.ods':
            nome_base = os.path.splitext(nome_arq)[0]
            caminho_xlsx = os.path.join(os.path.dirname(alvo), f'{nome_base}.xlsx')
            try:
                df_raw = pd.read_excel(alvo, sheet_name='FGTS EM ATRASO - PROCESSOS', header=None, engine='odf')
                df_raw.to_excel(caminho_xlsx, index=False, header=False)
                wb = load_workbook(caminho_xlsx)
                ws = wb.active
                ws.title = 'FGTS EM ATRASO - PROCESSOS'
                wb.save(caminho_xlsx)
            except Exception as e:
                print(f'  ERRO ao converter {nome_arq}: {e}')
                continue
        else:
            caminho_xlsx = alvo  # ja e .xlsx

        # Le o .xlsx para encontrar correcoes
        lista_alvo = ler_nomes_por_pis(caminho_xlsx)
        if not lista_alvo:
            print(f'  Nenhum nome valido encontrado.')
            if caminho_xlsx != alvo and os.path.exists(caminho_xlsx):
                os.remove(caminho_xlsx)
            continue

        print(f'  {len(lista_alvo)} nomes lidos')

        correcoes = []
        for idx, pis, nome_atual in lista_alvo:
            if not pis:
                continue
            nome_ref = mapa_pis.get(pis)
            if nome_ref and nome_ref != nome_atual:
                correcoes.append((idx, pis, nome_atual, nome_ref))

        for idx, pis, antigo, novo in correcoes:
            msg = f'  {antigo:45s} -> {novo}  (PIS: {pis})'
            print(msg)
            log_lines.append(msg)

        if not correcoes:
            print(f'  Nenhuma correcao necessaria.')
            if caminho_xlsx != alvo:
                print(f'  Convertido para: {caminho_xlsx}')
                arquivos_gerados.append(caminho_xlsx)
            continue

        print(f'  {len(correcoes)} nome(s) corrigido(s)')

        # Aplica correcoes no .xlsx com openpyxl (preserva formatacao)
        wb = load_workbook(caminho_xlsx)
        try:
            ws = wb['FGTS EM ATRASO - PROCESSOS']
        except:
            ws = wb.active
        for idx, pis, antigo, novo in correcoes:
            ws.cell(row=idx + 1, column=4).value = novo
        wb.save(caminho_xlsx)

        print(f'  Salvo: {os.path.basename(caminho_xlsx)} ({len(correcoes)} correcao(oes))')
        log_lines.append(f'{nome_arq}: {len(correcoes)} correcao(oes)')
        log_lines.append('')
        arquivos_gerados.append(caminho_xlsx)
        total_geral_corrigidos += len(correcoes)

    # Log final
    sep = '=' * 70
    log_lines.append(sep)
    log_lines.append('RESUMO')
    log_lines.append(f'  Total corrigidos: {total_geral_corrigidos}')
    log_lines.append(f'  Arquivos gerados: {len(arquivos_gerados)}')

    agora = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    log_dir = os.path.dirname(alvos[0]) if alvos else '.'
    log_path = os.path.join(log_dir, f'relatorio_correcoes_{agora}.txt')
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(log_lines))

    print(f'\n{"=" * 60}')
    print(f'  PROCESSO CONCLUIDO!')
    print(f'  Total corrigidos: {total_geral_corrigidos}')
    print(f'  Relatorio: {log_path}')
    if arquivos_gerados:
        print(f'  Arquivos gerados:')
        for a in arquivos_gerados:
            print(f'    - {a}')
    print(f'{"=" * 60}')

    input('\nPressione Enter para sair...')


if __name__ == '__main__':
    processar()
