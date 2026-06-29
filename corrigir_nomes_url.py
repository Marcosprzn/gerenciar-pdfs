#!/usr/bin/env python3
"""
Corrige nomes de pessoas dentro de planilhas .xls/.ods que contem
codificacao URL (%20, %2520, etc).
Ex: 'JOSE%2520PEREIRA' -> 'JOSE PEREIRA'
Seleciona uma pasta, escaneia todas as planilhas dentro e corrige.
"""
import os, sys, re
import tkinter as tk
from tkinter import filedialog
from urllib.parse import unquote
import xlrd
from xlutils.copy import copy as xl_copy

SHEET_NAME = 'FGTS EM ATRASO - PROCESSOS'

def decodificar(texto):
    if not isinstance(texto, str):
        return texto
    if '%' not in texto:
        return texto
    antigo = None
    while antigo != texto:
        antigo = texto
        texto = unquote(texto)
    return texto

def corrigir_xls(caminho):
    try:
        wb_rd = xlrd.open_workbook(caminho, formatting_info=True)
    except Exception:
        wb_rd = xlrd.open_workbook(caminho)

    if SHEET_NAME not in wb_rd.sheet_names():
        return 0

    sheet = wb_rd.sheet_by_name(SHEET_NAME)
    correcoes = []

    for row in range(sheet.nrows):
        val = str(sheet.cell_value(row, 3))
        if '%' in val:
            novo = decodificar(val)
            if novo != val:
                correcoes.append((row, val, novo))

    if not correcoes:
        return 0

    wb_wt = xl_copy(wb_rd)
    ws = wb_wt.get_sheet(wb_rd.sheet_names().index(SHEET_NAME))

    for idx, antigo, novo in correcoes:
        ws.write(idx, 3, novo)
        print(f"    L{idx+1}: {antigo[:60]:60s} -> {novo}")

    wb_wt.save(caminho)
    return len(correcoes)

def corrigir_ods(caminho):
    from odf import opendocument
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = opendocument.load(caminho)
    tables = doc.spreadsheet.getElementsByType(Table)
    table = None
    for t in tables:
        if t.getAttribute('name') == SHEET_NAME:
            table = t
            break
    if table is None:
        return 0

    rows = table.getElementsByType(TableRow)
    correcoes = []

    for idx, row_elem in enumerate(rows):
        cells = row_elem.getElementsByType(TableCell)
        if len(cells) < 4:
            continue
        cell = cells[3]
        for p in cell.getElementsByType(P):
            if p.firstChild is not None and '%' in p.firstChild.data:
                novo = decodificar(p.firstChild.data)
                if novo != p.firstChild.data:
                    correcoes.append((idx, p.firstChild.data, novo, cell))

    for idx, antigo, novo, cell in correcoes:
        for p in cell.getElementsByType(P):
            cell.removeChild(p)
        cell.setAttribute('valuetype', 'string')
        cell.addElement(P(text=novo))
        print(f"    L{idx+1}: {antigo[:60]:60s} -> {novo}")

    if correcoes:
        doc.save(caminho)

    return len(correcoes)

def corrigir_xlsx(caminho):
    from openpyxl import load_workbook

    wb = load_workbook(caminho)
    if SHEET_NAME not in wb.sheetnames:
        return 0

    ws = wb[SHEET_NAME]
    correcoes = []

    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=4)
        if cell.value and isinstance(cell.value, str) and '%' in cell.value:
            novo = decodificar(cell.value)
            if novo != cell.value:
                correcoes.append((row, cell.value, novo, cell))

    for idx, antigo, novo, cell in correcoes:
        cell.value = novo
        print(f"    L{idx}: {antigo[:60]:60s} -> {novo}")

    if correcoes:
        wb.save(caminho)

    return len(correcoes)


root = tk.Tk()
root.withdraw()
root.attributes('-topmost', True)
root.update()

pasta = filedialog.askdirectory(title="Selecione a PASTA com as planilhas")
root.destroy()

if not pasta:
    print("Nenhuma pasta selecionada.")
    sys.exit(1)

total_arquivos = 0
total_correcoes = 0

for f in sorted(os.listdir(pasta)):
    ext = os.path.splitext(f)[1].lower()
    caminho = os.path.join(pasta, f)

    if not os.path.isfile(caminho):
        continue

    print(f"\n--- {f} ---")

    try:
        if ext == '.xls':
            qtd = corrigir_xls(caminho)
        elif ext in ('.xlsx', '.xlsm'):
            qtd = corrigir_xlsx(caminho)
        elif ext == '.ods':
            qtd = corrigir_ods(caminho)
        else:
            continue

        if qtd:
            print(f"  {qtd} nome(s) corrigido(s)")
            total_correcoes += qtd
        else:
            print(f"  Nenhum nome com % encontrado.")
        total_arquivos += 1

    except Exception as e:
        print(f"  ERRO: {e}")

print(f"\nResumo: {total_arquivos} planilha(s) lida(s), {total_correcoes} correcao(oes)")
input("\nPressione Enter para sair...")
