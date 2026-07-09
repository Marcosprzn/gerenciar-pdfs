#!/usr/bin/env python3
"""
INSPETOR DE PLANILHA (para decidir como filtrar as pessoas excluidas)

Mostra, para cada pessoa, os valores das colunas ao redor da J e se a linha
esta oculta. Assim da pra ver o que diferencia as pessoas "excluidas" (vermelho
/ J vazio / ocultas) das validas.

Como usar:
    python inspecionar_planilha.py
    -> selecione a PLANILHA (.xls/.xlsx/.ods)
    -> envie o arquivo 'inspecao_planilha.txt' gerado (na mesma pasta).

So le, nao altera nada. PIS aparece mascarado.
"""
import os
import sys
import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import confinicial

SHEET = "FGTS EM ATRASO - PROCESSOS"
# Colunas (0-based): B=1 proc, C=2 pis, D=3 nome, ... J=9
COL_PROC, COL_PIS, COL_NOME, COL_J = 1, 2, 3, 9


def _sel():
    try:
        import tkinter as tk
        from tkinter import filedialog
        r = tk.Tk(); r.withdraw(); r.attributes('-topmost', True)
        p = filedialog.askopenfilename(
            title="Selecione a PLANILHA para inspecionar",
            filetypes=[("Planilhas", "*.xlsx *.xls *.ods"), ("Todos", "*.*")])
        r.destroy()
        return p
    except Exception:
        return sys.argv[1] if len(sys.argv) > 1 else None


def _mascara_pis(v):
    d = ''.join(c for c in str(v) if c.isdigit())
    return ('***' + d[-4:]) if len(d) >= 4 else '***'


def _val(v):
    """Formata um valor de celula; '' se vazio/nan."""
    if v is None:
        return ''
    s = str(v).strip()
    if s.lower() in ('', 'nan', 'none'):
        return ''
    if s.endswith('.0'):
        s = s[:-2]
    return s


def ler_ocultas(caminho):
    """Tenta descobrir quais linhas (0-based) estao ocultas. Retorna set ou None
    se nao der pra ler nesse formato."""
    ext = os.path.splitext(caminho)[1].lower()
    try:
        if ext == '.xls':
            import xlrd
            wb = xlrd.open_workbook(caminho, formatting_info=True)
            sh = wb.sheet_by_name(SHEET) if SHEET in wb.sheet_names() else wb.sheet_by_index(0)
            ocultas = set()
            for rx, info in getattr(sh, 'rowinfo_map', {}).items():
                if getattr(info, 'hidden', 0):
                    ocultas.add(rx)
            return ocultas
        elif ext in ('.xlsx', '.xlsm'):
            from openpyxl import load_workbook
            wb = load_workbook(caminho)
            ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
            ocultas = set()
            for rx, dim in ws.row_dimensions.items():
                if dim.hidden:
                    ocultas.add(rx - 1)  # openpyxl e 1-based
            return ocultas
    except Exception as e:
        return ('ERRO', str(e))
    return None


def main():
    caminho = _sel()
    if not caminho:
        print("Nenhuma planilha selecionada.")
        return

    import pandas as pd
    ext = os.path.splitext(caminho)[1].lower()
    engine = "odf" if ext == '.ods' else None
    df = pd.read_excel(caminho, sheet_name=SHEET, header=None, engine=engine)

    ocultas = ler_ocultas(caminho)
    tem_ocultas = isinstance(ocultas, set)

    linhas = []
    def w(s=""):
        linhas.append(str(s))

    w("=" * 78)
    w("INSPECAO DE PLANILHA — " + datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    w("Arquivo: " + os.path.basename(caminho))
    if tem_ocultas:
        w(f"Leitura de linhas ocultas: OK ({len(ocultas)} linhas ocultas encontradas)")
    elif isinstance(ocultas, tuple):
        w("Leitura de linhas ocultas: FALHOU -> " + ocultas[1])
    else:
        w("Leitura de linhas ocultas: nao suportada neste formato")
    w("=" * 78)
    w("Colunas mostradas: E,F,G,H,I (antes do J) e J. '.' = vazio.")
    w("")
    w("%-4s %-3s %-30s | %-6s | %-25s | %-6s | %s" % (
        "LIN", "OC", "NOME (col D)", "J(col)", "E F G H I (antes do J)", "PIS", "PROC"))
    w("-" * 100)

    n_total = n_j_vazio = n_j_vazio_resto_vazio = n_j_vazio_oculta = 0
    for i in range(len(df)):
        nome = df.iat[i, COL_NOME] if COL_NOME < df.shape[1] else None
        if not confinicial.e_um_nome_valido(nome):
            continue
        n_total += 1
        j = _val(df.iat[i, COL_J]) if COL_J < df.shape[1] else ''
        antes = [(_val(df.iat[i, c]) if c < df.shape[1] else '') for c in range(4, 9)]  # E..I
        oculta = (tem_ocultas and i in ocultas)
        j_vazio = (j == '')
        resto_vazio = all(v == '' for v in antes)
        if j_vazio:
            n_j_vazio += 1
            if resto_vazio:
                n_j_vazio_resto_vazio += 1
            if oculta:
                n_j_vazio_oculta += 1
        antes_str = ' '.join((v[:4] if v else '.') for v in antes)
        w("%-4d %-3s %-30s | %-6s | %-25s | %-6s | %s" % (
            i,
            "OC" if oculta else "",
            str(nome)[:30],
            (j[:6] if j else '.'),
            antes_str,
            _mascara_pis(df.iat[i, COL_PIS] if COL_PIS < df.shape[1] else ''),
            _val(df.iat[i, COL_PROC] if COL_PROC < df.shape[1] else '')[:12],
        ))

    w("")
    w("=" * 78)
    w("RESUMO:")
    w(f"  Pessoas (linhas com nome valido): {n_total}")
    w(f"  Com J vazio:                      {n_j_vazio}")
    w(f"  Com J vazio E colunas E..I tambem vazias: {n_j_vazio_resto_vazio}")
    if tem_ocultas:
        w(f"  Com J vazio E linha oculta:       {n_j_vazio_oculta}")
    w("=" * 78)

    saida = os.path.join(os.path.dirname(caminho), "inspecao_planilha.txt")
    with open(saida, "w", encoding="utf-8") as f:
        f.write("\n".join(linhas))
    print("\nRelatorio salvo em:\n  " + saida + "\nEnvie esse arquivo.")


if __name__ == "__main__":
    main()
