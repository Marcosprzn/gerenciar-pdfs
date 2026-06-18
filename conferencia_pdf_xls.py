#!/usr/bin/env python3
"""
Conferência PDF vs Planilha FGTS
Compara nomes de arquivos PDF com nomes em planilha (aba "FGTS EM ATRASO - PROCESSOS")
"""
import pandas as pd
import os, re, sys, unicodedata
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from difflib import SequenceMatcher

# ============================================================
# FUNCOES DE EXTRACAO DE DATA
# ============================================================

def extrair_competencia_excel(nome_arquivo):
    """Extrai mes e ano do nome do arquivo Excel. Ex: '01_07 USIVALE' -> ('01', '2008')"""
    # Procura 2 digitos (mes), separador, 2-4 digitos (ano)
    match = re.search(r"(?<!\d)(\d{2})[-_ ](\d{2,4})(?!\d)", nome_arquivo)
    if match:
        mes = match.group(1)
        ano_str = match.group(2)
        ano = "20" + ano_str if len(ano_str) == 2 else ano_str
        return mes, ano
    return None, None


def extrair_competencia_pasta(nome_pasta):
    """Extrai mes e ano do nome da pasta. Ex: '01-SEFIP FGTS - COMPETENCIA 01-2007' -> ('01', '2007')"""
    # Formato MM AAAA ou MM-AAAA
    match = re.search(r"(?<!\d)(\d{2})\s*[-_/\.]\s*(\d{4})(?!\d)", nome_pasta)
    if match:
        return match.group(1), match.group(2)
    # Formato texto MES AAAA
    nome_norm = normalizar_texto(nome_pasta.upper())
    meses = {
        'JANEIRO': '01', 'FEVEREIRO': '02', 'MARCO': '03', 'ABRIL': '04',
        'MAIO': '05', 'JUNHO': '06', 'JULHO': '07', 'AGOSTO': '08',
        'SETEMBRO': '09', 'OUTUBRO': '10', 'NOVEMBRO': '11', 'DEZEMBRO': '12'
    }
    for mes_texto, mes_num in meses.items():
        if mes_texto in nome_norm:
            match_ano = re.search(r"(?<!\d)(\d{4})(?!\d)", nome_norm)
            if match_ano:
                return mes_num, match_ano.group(1)
    return None, None


# ============================================================
# FUNCOES AUXILIARES
# ============================================================

def normalizar_texto(texto):
    """Remove acentos, padroniza maiusculas, remove pontuacao."""
    if not isinstance(texto, str): return ""
    nfkd = unicodedata.normalize('NFKD', texto)
    texto_sem_acento = "".join([c for c in nfkd if not unicodedata.combining(c)])
    texto_upper = texto_sem_acento.upper()
    texto_sem_ponto = texto_upper.replace('.', '')
    texto_sem_hifen = texto_sem_ponto.replace('-', ' ').replace('_', ' ')
    texto_limpo = " ".join(texto_sem_hifen.split())
    return texto_limpo


def calcular_similaridade(a, b):
    return SequenceMatcher(None, a, b).ratio()


def verificar_abreviacao(nome_planilha, nome_pdf):
    """Verifica se um nome pode ser abreviacao do outro."""
    conectores = ['DE', 'DA', 'DO', 'DOS', 'DAS']
    tokens_p = [t for t in nome_planilha.split() if t not in conectores]
    tokens_f = [t for t in nome_pdf.split() if not t.isdigit() and t not in conectores]
    if abs(len(tokens_p) - len(tokens_f)) > 2:
        return False
    idx_p = 0
    idx_f = 0
    match_count = 0
    while idx_p < len(tokens_p) and idx_f < len(tokens_f):
        t_p = tokens_p[idx_p]
        t_f = tokens_f[idx_f]
        if t_p == t_f:
            match_count += 1
            idx_p += 1
            idx_f += 1
        elif (len(t_p) <= 4 and t_f.startswith(t_p)) or (len(t_f) <= 4 and t_p.startswith(t_f)):
            match_count += 1
            idx_p += 1
            idx_f += 1
        else:
            if len(tokens_p) > len(tokens_f):
                idx_p += 1
            elif len(tokens_f) > len(tokens_p):
                idx_f += 1
            else:
                idx_p += 1
                idx_f += 1
    return match_count >= max(len(tokens_p), len(tokens_f)) - 1


def e_um_nome_valido(texto):
    """Verifica se o texto e um nome de pessoa valido (nao e cabecalho/lixo)."""
    if not isinstance(texto, str): return False
    texto = texto.strip().upper()
    ignorar = ["NOMES", "NOME", "TITULAR", "FUNCIONARIO", "PIS", "PROC.", "DATA", "OBS", "TOTAL"]
    if texto in ignorar: return False
    # Nome de pessoa: letras, acentos, espaco, e opcionalmente digitos/numeros
    if re.match(r"^[A-ZÁÀÂÃÉÊÍÓÔÕÚÜÇ0-9 \.\-]+$", texto) and " " in texto and len(texto) > 5:
        return True
    return False


# ============================================================
# LEITURA DA PLANILHA
# ============================================================

def ler_planilha_fgts(caminho_arquivo):
    """
    Le a planilha FGTS (aba 'FGTS EM ATRASO - PROCESSOS').
    Retorna DataFrame com colunas: PROC., PIS, NOMES, TIPO_LISTA
    TIPO_LISTA = 'PADRAO' (1a secao) ou '115' (2a secao)
    """
    if not caminho_arquivo:
        return None
    if not os.path.exists(caminho_arquivo):
        print(f'  ERRO: arquivo nao encontrado: {caminho_arquivo}')
        return None

    extensao = os.path.splitext(caminho_arquivo)[1].lower()
    print(f'  Lendo planilha: {os.path.basename(caminho_arquivo)}  ({extensao})')

    try:
        engine_pd = "odf" if extensao == '.ods' else None
        df_raw = pd.read_excel(
            caminho_arquivo,
            sheet_name="FGTS EM ATRASO - PROCESSOS",
            header=None,
            engine=engine_pd
        )
    except Exception as e:
        print(f'  ERRO ao ler planilha: {e}')
        try:
            xls = pd.ExcelFile(caminho_arquivo)
            print(f'  Abas disponiveis: {xls.sheet_names}')
        except:
            pass
        if extensao == '.ods':
            print('  DICA: pip install odfpy')
        return None

    idx_proc, idx_pis, idx_nomes = 1, 2, 3

    if df_raw.shape[1] < 4:
        print(f'  ERRO: planilha tem apenas {df_raw.shape[1]} coluna(s)')
        return None

    dados = []
    encontrou_primeiro_inicio = False
    gap_linhas_vazias = 0
    bloco_atual = "PADRAO"
    LIMITE_GAP_TOTAL = 40
    LIMITE_MUDANCA_BLOCO = 5

    for i in range(len(df_raw)):
        linha = df_raw.iloc[i]
        val_nome = linha[idx_nomes]
        eh_nome = e_um_nome_valido(val_nome)

        if eh_nome:
            if not encontrou_primeiro_inicio:
                print(f'  Primeiro nome encontrado na linha {i+1}')
            encontrou_primeiro_inicio = True
            gap_linhas_vazias = 0
            dados.append({
                "PROC.": linha[idx_proc],
                "PIS": linha[idx_pis],
                "NOMES": val_nome,
                "TIPO_LISTA": bloco_atual
            })
        else:
            if encontrou_primeiro_inicio:
                gap_linhas_vazias += 1
                if gap_linhas_vazias > LIMITE_MUDANCA_BLOCO and bloco_atual != "115":
                    bloco_atual = "115"
                    print(f'  Mudou para secao 115 (gap de {gap_linhas_vazias} linhas vazias)')
                if gap_linhas_vazias > LIMITE_GAP_TOTAL:
                    break

    df = pd.DataFrame(dados)
    qtde_padrao = len(df[df['TIPO_LISTA'] == 'PADRAO']) if not df.empty else 0
    qtde_115 = len(df[df['TIPO_LISTA'] == '115']) if not df.empty else 0
    print(f'  Total: {len(df)} nomes validos ({qtde_padrao} PADRAO, {qtde_115} como 115)')
    return df


# ============================================================
# BUSCA DE PDFs
# ============================================================

def buscar_arquivos_pasta(pasta, incluir_subpastas=False):
    """Busca arquivos na pasta (excluindo planilhas)."""
    EXT_PLANILHAS = ('.xls', '.xlsx', '.ods')
    arquivos = []
    for raiz, _, arquivos_lista in os.walk(pasta):
        if raiz != pasta and not incluir_subpastas:
            continue
        for f in arquivos_lista:
            if not f.lower().endswith(EXT_PLANILHAS):
                p = os.path.join(raiz, f)
                if p not in arquivos:
                    arquivos.append(p)
    return arquivos


def verificar_pdfs(df, arquivos_pdf):
    """
    Cruza os nomes da planilha com os arquivos encontrados.
    Retorna o DataFrame com colunas Status PDF e Nome do Arquivo Encontrado.
    """
    # Prepara lista de arquivos disponiveis
    arquivos_disponiveis = []
    for caminho in arquivos_pdf:
        nome_arquivo = os.path.basename(caminho)
        arquivos_disponiveis.append({
            "real": nome_arquivo,
            "norm": normalizar_texto(os.path.splitext(nome_arquivo)[0]),
            "caminho": caminho
        })

    status_lista = []
    arquivo_encontrado_lista = []

    for index, row in df.iterrows():
        nome_planilha = row['NOMES']
        tipo_lista = row.get('TIPO_LISTA', 'PADRAO')

        nome_norm = normalizar_texto(nome_planilha)
        match_status = "NAO ENCONTRADO NO PDF"
        match_arquivo_obj = None

        # 1. BUSCA EXATA
        for arq in arquivos_disponiveis:
            if nome_norm in arq["norm"]:
                match_status = "ENCONTRADO"
                match_arquivo_obj = arq
                break

        # 2. BUSCA POR ABREVIACAO
        if match_status == "NAO ENCONTRADO NO PDF":
            for arq in arquivos_disponiveis:
                if verificar_abreviacao(nome_norm, arq["norm"]):
                    match_status = "ENCONTRADO COM ABREVIACAO"
                    match_arquivo_obj = arq
                    break

        # 3. BUSCA APROXIMADA (similaridade)
        if match_status == "NAO ENCONTRADO NO PDF":
            melhor_ratio = 0
            candidato_obj = None
            for arq in arquivos_disponiveis:
                ratio = calcular_similaridade(nome_norm, arq["norm"])
                if ratio > melhor_ratio:
                    melhor_ratio = ratio
                    candidato_obj = arq
            if melhor_ratio >= 0.80:
                match_status = "POSSIVEL ERRO NOMINAL"
                match_arquivo_obj = candidato_obj

        # STATUS FINAL
        if match_arquivo_obj:
            arquivos_disponiveis.remove(match_arquivo_obj)
            if tipo_lista == "115" and "ENCONTRADO" in match_status:
                match_status = "ENCONTRADO COMO 115"
            status_lista.append(match_status)
            arquivo_encontrado_lista.append(os.path.splitext(match_arquivo_obj["real"])[0])
        else:
            status_lista.append("NAO ENCONTRADO NO PDF")
            arquivo_encontrado_lista.append("")

    df['Status PDF'] = status_lista
    df['Nome do Arquivo Encontrado'] = arquivo_encontrado_lista

    if 'TIPO_LISTA' in df.columns:
        df = df.drop(columns=['TIPO_LISTA'])

    # 4. SOBRAS (PDFs que nao estao na planilha)
    sobras = []
    for arq in arquivos_disponiveis:
        sobras.append({
            "PROC.": "",
            "PIS": "",
            "NOMES": "",
            "Status PDF": "PDF NA PASTA, MAS NAO NA PLANILHA",
            "Nome do Arquivo Encontrado": os.path.splitext(arq["real"])[0]
        })
    if sobras:
        df = pd.concat([df, pd.DataFrame(sobras)], ignore_index=True)

    return df


# ============================================================
# FORMATACAO CONDICIONAL DO EXCEL
# ============================================================

def aplicar_formatacao_condicional(caminho_arquivo):
    """Aplica cores e formatacao ao arquivo Excel gerado."""
    try:
        wb = load_workbook(caminho_arquivo)
        ws = wb.active

        borda_fina = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        fill_cinza = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        fill_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fonte_vermelha = Font(color="FF0000", bold=True)
        fonte_verde = Font(color="006400", bold=True)
        fonte_azul = Font(color="0000FF", bold=True)
        fonte_preta = Font(color="000000", bold=True)
        fonte_roxo = Font(color="800080", bold=True)

        col_idx = None
        for cell in ws[1]:
            if cell.value and "Status PDF" in str(cell.value):
                col_idx = cell.column
                break

        for row in ws.iter_rows():
            for cell in row:
                cell.border = borda_fina
                if cell.row == 1:
                    cell.fill = fill_cinza
                    cell.font = Font(bold=True)
                elif col_idx and cell.column == col_idx:
                    val = str(cell.value)
                    if "NAO ENCONTRADO" in val:
                        cell.font = fonte_vermelha
                    elif val == "ENCONTRADO COMO 115":
                        cell.font = fonte_roxo
                    elif "ENCONTRADO" in val:
                        cell.font = fonte_verde
                    elif val == "POSSIVEL ERRO NOMINAL":
                        cell.fill = fill_amarelo
                        cell.font = fonte_preta
                    elif val == "PDF NA PASTA, MAS NAO NA PLANILHA":
                        cell.font = fonte_azul

        for column_cells in ws.columns:
            length = 0
            col_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > length:
                        length = len(str(cell.value))
                except:
                    pass
            if length > 50:
                length = 50
            ws.column_dimensions[col_letter].width = length + 2

        wb.save(caminho_arquivo)
    except Exception as e:
        print(f"  Aviso: erro na formatacao: {e}")


# ============================================================
# ORGANIZAR PDFs EM PASTAS 115 660
# ============================================================

def organizar_pdfs_por_resultado(resultado):
    """
    Cria duas pastas (115 e 660) na mesma pasta de origem e copia os PDFs
    conforme o resultado da conferencia.
    - PASTA 115: PDFs com status 'ENCONTRADO COMO 115'
    - PASTA 660: todos os demais PDFs encontrados
    """
    df = resultado['df']
    pasta_origem = resultado['pasta_origem']
    nome_pasta = os.path.basename(pasta_origem)

    # Cria os nomes das novas pastas
    # Ex: "01-SEFIP FGTS - COMPETENCIA 01-2007"
    #   -> "01-SEFIP FGTS 115 - COMPETENCIA 01-2007"
    #   -> "01-SEFIP FGTS 660 - COMPETENCIA 01-2007"
    padrao_subst = r'( - COMPETENCIA)'
    pasta_115 = re.sub(padrao_subst, r' 115\1', nome_pasta, flags=re.IGNORECASE)
    pasta_660 = re.sub(padrao_subst, r' 660\1', nome_pasta, flags=re.IGNORECASE)

    # Se o padrao nao encontrou, usa fallback
    if pasta_115 == nome_pasta:
        pasta_115 = nome_pasta + ' 115'
        pasta_660 = nome_pasta + ' 660'

    caminho_115 = os.path.join(os.path.dirname(pasta_origem), pasta_115)
    caminho_660 = os.path.join(os.path.dirname(pasta_origem), pasta_660)

    import shutil

    # Contadores
    count_115 = 0
    count_660 = 0

    # Filtra apenas linhas que tem um PDF correspondente
    for _, row in df.iterrows():
        status = str(row.get('Status PDF', ''))
        nome_arquivo = str(row.get('Nome do Arquivo Encontrado', ''))

        if not nome_arquivo or nome_arquivo == '':
            continue
        if status == 'PDF NA PASTA, MAS NAO NA PLANILHA':
            continue

        # Encontra o caminho completo do PDF
        caminho_pdf = None
        for arq in resultado['arquivos_pdf']:
            if os.path.splitext(os.path.basename(arq))[0] == nome_arquivo:
                caminho_pdf = arq
                break

        if not caminho_pdf or not os.path.exists(caminho_pdf):
            continue

        # Decide destino
        if status == 'ENCONTRADO COMO 115':
            os.makedirs(caminho_115, exist_ok=True)
            shutil.copy2(caminho_pdf, os.path.join(caminho_115, os.path.basename(caminho_pdf)))
            count_115 += 1
        else:
            os.makedirs(caminho_660, exist_ok=True)
            shutil.copy2(caminho_pdf, os.path.join(caminho_660, os.path.basename(caminho_pdf)))
            count_660 += 1

    print(f'\n  PDFs organizados:')
    if count_115 > 0:
        print(f'    {count_115} na pasta 115: {caminho_115}')
    if count_660 > 0:
        print(f'    {count_660} na pasta 660: {caminho_660}')

    return count_115 > 0 or count_660 > 0


# ============================================================
# FUNCAO PRINCIPAL - PROCESSAR UMA PASTA
# ============================================================

def processar_pasta(pasta, incluir_subpastas=False, caminho_planilha_forcado=None):
    """
    Processa uma pasta: busca arquivos, le planilha, cruza dados.
    Retorna (df_resultado, caminho_planilha, nome_base) ou None se falhar.
    """
    nome_pasta = os.path.basename(pasta)
    print(f'\n>>> Lendo pasta: {nome_pasta}')

    # Mostra conteudo
    try:
        todos = os.listdir(pasta)
        print(f'  Conteudo ({len(todos)} item(ns)):')
        for f in sorted(todos)[:25]:
            full = os.path.join(pasta, f)
            if os.path.isdir(full):
                print(f'    [PASTA]    {f}')
            else:
                _, e = os.path.splitext(f)
                print(f'    [{e.upper() or "?"}]  {f}')
        if len(todos) > 25:
            print(f'    ... e mais {len(todos) - 25}')
    except Exception as ex:
        print(f'  ERRO ao listar: {ex}')

    # Busca arquivos (PDFs ou qualquer arquivo que nao seja planilha)
    arquivos = buscar_arquivos_pasta(pasta, incluir_subpastas)
    print(f'  Arquivos (exceto planilhas): {len(arquivos)}')
    if not arquivos:
        print('  Nenhum arquivo encontrado.')
        return None

    # Data da pasta
    mes, ano = extrair_competencia_pasta(nome_pasta)

    # Localiza planilha
    if caminho_planilha_forcado:
        caminho_planilha = caminho_planilha_forcado
    elif mes and ano:
        # Procura na pasta
        candidatas = []
        for f in os.listdir(pasta):
            if f.lower().endswith(('.xls', '.xlsx', '.ods')):
                candidatas.append(os.path.join(pasta, f))
        if len(candidatas) == 1:
            caminho_planilha = candidatas[0]
        else:
            # Filtra por data no nome
            chave = f'{mes}_{ano}'
            chave2 = f'{mes}{ano}'
            chave3 = f'{mes}_{ano[-2:]}'
            for c in candidatas:
                nome = os.path.basename(c)
                if chave in nome or chave2 in nome or chave3 in nome:
                    caminho_planilha = c
                    break
            else:
                caminho_planilha = None
                if candidatas:
                    print(f'  Planilhas encontradas mas sem data correspondente:')
                    for c in candidatas:
                        print(f'    - {os.path.basename(c)}')
    else:
        caminho_planilha = None

    if not caminho_planilha:
        print(f'  Planilha nao encontrada nesta pasta.')
        print(f'  (se estiver em outra pasta, selecione manualmente)')
        return None

    print(f'  Planilha: {os.path.basename(caminho_planilha)}')

    # Le planilha
    df = ler_planilha_fgts(caminho_planilha)
    if df is None or df.empty:
        print(f'  Nenhum nome valido encontrado na planilha.')
        return None

    # Cruza com PDFs
    print(f'  Cruzando {len(df)} nomes com {len(arquivos)} arquivo(s)...')
    df = verificar_pdfs(df, arquivos)

    nome_base = os.path.splitext(os.path.basename(caminho_planilha))[0]
    return {
        'df': df,
        'caminho_planilha': caminho_planilha,
        'nome_base': nome_base,
        'arquivos_pdf': arquivos,
        'pasta_origem': pasta,
    }


# ============================================================
# RESUMO NO CONSOLE
# ============================================================

def exibir_resumo(resultado):
    """Exibe resumo do resultado no terminal."""
    df = resultado['df']
    print(f'\n{"=" * 60}')
    print(f'  RESUMO DA CONFERENCIA')
    print(f'{"=" * 60}')

    totais = df['Status PDF'].value_counts()
    for status in ['ENCONTRADO', 'ENCONTRADO COMO 115', 'ENCONTRADO COM ABREVIACAO',
                    'POSSIVEL ERRO NOMINAL', 'NAO ENCONTRADO NO PDF',
                    'PDF NA PASTA, MAS NAO NA PLANILHA']:
        qtde = totais.get(status, 0)
        if qtde > 0:
            print(f'  {qtde:3d}x  {status}')

    print(f'  {"-" * 50}')
    print(f'  Total de nomes na planilha: {len(df[df["Status PDF"] != "PDF NA PASTA, MAS NAO NA PLANILHA"])}')
    print(f'  Total de PDFs: {len(df[df["Status PDF"] == "PDF NA PASTA, MAS NAO NA PLANILHA"]) + len(df[df["Status PDF"].str.contains("ENCONTRADO", na=False)])}')
    print(f'  {"=" * 60}')


# ============================================================
# SALVAR RESULTADO
# ============================================================

def salvar_resultado(resultado, caminho_saida):
    """Salva o DataFrame em Excel com formatacao."""
    df = resultado['df']

    # Ordenacao
    prioridade = {
        "NAO ENCONTRADO NO PDF": 1,
        "POSSIVEL ERRO NOMINAL": 2,
        "PDF NA PASTA, MAS NAO NA PLANILHA": 3,
        "ENCONTRADO COMO 115": 4,
        "ENCONTRADO COM ABREVIACAO": 5,
        "ENCONTRADO": 6,
    }
    df['Prioridade'] = df['Status PDF'].map(prioridade).fillna(8)
    df = df.sort_values(by=['Prioridade', 'NOMES', 'Nome do Arquivo Encontrado'])
    df = df.drop(columns=['Prioridade'])

    df.to_excel(caminho_saida, index=False)
    aplicar_formatacao_condicional(caminho_saida)
    print(f'  Arquivo salvo: {caminho_saida}')


# ============================================================
# DIALOGOS TKINTER
# ============================================================

def dialogo_pasta(titulo):
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    root.update()
    pasta = filedialog.askdirectory(title=titulo)
    root.destroy()
    return pasta


def dialogo_arquivo(titulo, tipos):
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    root.update()
    arquivo = filedialog.askopenfilename(title=titulo, filetypes=tipos)
    root.destroy()
    return arquivo


def dialogo_salvar(titulo, sugestao):
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    root.update()
    caminho = filedialog.asksaveasfilename(
        title=titulo,
        defaultextension='.xlsx',
        filetypes=[('Planilha Excel', '*.xlsx')],
        initialfile=sugestao,
    )
    root.destroy()
    return caminho


# ============================================================
# MAIN
# ============================================================

def main():
    print('=' * 60)
    print('  CONFERENCIA PDF vs PLANILHA FGTS')
    print('=' * 60)

    print('\nModos de operacao:')
    print('  1 — Individual (selecionar UMA pasta)')
    print('  2 — Lote (selecionar pasta RAIZ — processa todas as subpastas)')
    while True:
        modo = input('\nOpcao (1 ou 2): ').strip()
        if modo in ('1', '2'):
            break

    if modo == '1':
        # ===== INDIVIDUAL =====
        pasta = dialogo_pasta('Selecione a pasta com os arquivos')
        if not pasta:
            print('Nenhuma pasta selecionada.')
            return

        resp = input('\nIncluir subpastas? (S/n): ').strip().lower()
        incluir_sub = resp != 'n'

        resultado = processar_pasta(pasta, incluir_sub)
        if not resultado:
            print('\nPlanilha nao encontrada nesta pasta.')
            resp2 = input('Selecionar planilha manualmente em outra pasta? (S/n): ').strip().lower()
            if resp2 != 'n':
                plan = dialogo_arquivo(
                    'Selecione a planilha (.xls / .xlsx / .ods)',
                    [('Planilhas', '*.xls *.xlsx *.ods')]
                )
                if plan:
                    resultado = processar_pasta(pasta, incluir_sub, caminho_planilha_forcado=plan)
            if not resultado:
                print('Nao foi possivel processar.')
                return

        exibir_resumo(resultado)

        timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
        sugestao = f'resultado_{resultado["nome_base"]}_{timestamp}.xlsx'
        caminho = dialogo_salvar('Salvar resultado como...', sugestao)
        if not caminho:
            print('Salvamento cancelado.')
            return

        salvar_resultado(resultado, caminho)

        # Pergunta se quer organizar PDFs em pastas 115/660
        print()
        resp3 = input('Organizar PDFs em pastas 115 e 660? (S/n): ').strip().lower()
        if resp3 != 'n':
            organizar_pdfs_por_resultado(resultado)

    else:
        # ===== LOTE =====
        pasta_raiz = dialogo_pasta('Selecione a pasta RAIZ (com subpastas de competencia)')
        if not pasta_raiz:
            print('Nenhuma pasta selecionada.')
            return

        resp = input('\nIncluir PDFs de sub-subpastas? (S/n): ').strip().lower()
        incluir_sub = resp != 'n'

        # Mapeia subpastas por competencia
        mapa_pastas = {}
        for item in os.listdir(pasta_raiz):
            caminho = os.path.join(pasta_raiz, item)
            if os.path.isdir(caminho):
                mes, ano = extrair_competencia_pasta(item)
                if mes and ano:
                    mapa_pastas[f"{mes}_{ano}"] = caminho

        if not mapa_pastas:
            print('Nenhuma subpasta com data no nome encontrada.')
            return

        print(f'\nSubpastas encontradas: {len(mapa_pastas)}')
        for chave, caminho in sorted(mapa_pastas.items()):
            print(f'  {chave} -> {os.path.basename(caminho)}')

        # Seleciona planilhas
        print('\nAgora selecione a(s) planilha(s) .xls/.xlsx/.ods')
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        root.update()
        planilhas = filedialog.askopenfilenames(
            title='Selecione as planilhas FGTS',
            filetypes=[('Planilhas', '*.xls *.xlsx *.ods')]
        )
        root.destroy()

        if not planilhas:
            print('Nenhuma planilha selecionada.')
            return

        print(f'{len(planilhas)} planilha(s) selecionada(s)')

        # Seleciona pasta de destino
        pasta_destino = dialogo_pasta('Selecione a pasta para salvar os resultados')
        if not pasta_destino:
            print('Nenhuma pasta de destino selecionada.')
            return

        total_processadas = 0
        for arquivo in planilhas:
            nome_base = os.path.basename(arquivo)
            print(f'\n--- Processando: {nome_base} ---')

            mes_ex, ano_ex = extrair_competencia_excel(nome_base)
            df = ler_planilha_fgts(arquivo)

            if df is None or df.empty:
                print(f'  Nenhum dado valido na planilha.')
                continue

            if not mes_ex or not ano_ex:
                print(f'  Mes/ano nao identificado no nome do arquivo.')
                df['Status PDF'] = 'COMPETENCIA DESCONHECIDA'
                df['Nome do Arquivo Encontrado'] = ''
                nome_saida = f'Relatorio_{os.path.splitext(nome_base)[0]}_sem_pasta.xlsx'
            else:
                chave_comp = f'{mes_ex}_{ano_ex}'
                if chave_comp in mapa_pastas:
                    pasta_pdfs = mapa_pastas[chave_comp]
                    arquivos_pdf = buscar_arquivos_pasta(pasta_pdfs, incluir_sub)
                    print(f'  Cruzando com pasta: {os.path.basename(pasta_pdfs)} ({len(arquivos_pdf)} arquivo(s))')
                    df = verificar_pdfs(df, arquivos_pdf)
                    nome_saida = f'{mes_ex}_{ano_ex}.xlsx'
                else:
                    print(f'  Pasta de PDFs para {mes_ex}/{ano_ex} nao encontrada.')
                    df['Status PDF'] = 'PASTA NAO ENCONTRADA'
                    df['Nome do Arquivo Encontrado'] = ''
                    nome_saida = f'{mes_ex}_{ano_ex}_sem_pasta.xlsx'

            # Ordena e salva
            prioridade = {
                "PASTA NAO ENCONTRADA": 1,
                "COMPETENCIA DESCONHECIDA": 1,
                "NAO ENCONTRADO NO PDF": 2,
                "POSSIVEL ERRO NOMINAL": 3,
                "PDF NA PASTA, MAS NAO NA PLANILHA": 4,
                "ENCONTRADO COMO 115": 5,
                "ENCONTRADO COM ABREVIACAO": 6,
                "ENCONTRADO": 7,
            }
            df['Prioridade'] = df['Status PDF'].map(prioridade).fillna(8)
            df = df.sort_values(by=['Prioridade', 'NOMES', 'Nome do Arquivo Encontrado'])
            df = df.drop(columns=['Prioridade'])

            caminho_saida = os.path.join(pasta_destino, nome_saida)
            try:
                df.to_excel(caminho_saida, index=False)
                aplicar_formatacao_condicional(caminho_saida)
                print(f'  Salvo: {nome_saida}')
                total_processadas += 1
            except Exception as e:
                print(f'  Erro ao salvar: {e}')

        print(f'\n{"=" * 60}')
        print(f'  PROCESSO CONCLUIDO! {total_processadas}/{len(planilhas)} processadas')
        print(f'  Pasta: {pasta_destino}')
        print(f'{"=" * 60}')
        try:
            os.startfile(pasta_destino)
        except:
            pass

    input('\nPressione Enter para sair...')


if __name__ == '__main__':
    main()
