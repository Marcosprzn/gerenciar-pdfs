import os
import re
import shutil
import tkinter as tk
from tkinter import filedialog
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from difflib import SequenceMatcher
import pandas as pd
try:
    from tqdm import tqdm
except ImportError:   # tqdm e opcional (so a barra de progresso no terminal)
    def tqdm(iteravel=None, *args, **kwargs):
        return iteravel if iteravel is not None else []
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    import llm_matcher
except ImportError:
    llm_matcher = None

EXT = ('.pdf', '.tif', '.tiff')
USAR_LLM = False

# --- FUNÇÕES DE EXTRAÇÃO DE COMPETÊNCIA ---

def extrair_competencia_excel(nome_arquivo):
    """
    Extrai o mês e ano do nome do arquivo Excel. Ex: '04_08 USIVALE_OK FEC' -> ('04', '2008')
    """
    # Procura um padrão de 2 dígitos (mês), seguido de um separador (_ , - ou espaço) e 2 a 4 dígitos (ano)
    # (?<!\d) e (?!\d) garantem que os números não estejam grudados em outros números
    match = re.search(r"(?<!\d)(\d{2})[-_ ](\d{2,4})(?!\d)", nome_arquivo)
    if match:
        mes = match.group(1)
        ano_str = match.group(2)
        # Se o ano tiver 2 dígitos (ex: '08'), adiciona '20' na frente ('2008')
        ano = "20" + ano_str if len(ano_str) == 2 else ano_str
        return mes, ano
    return None, None

def extrair_competencia_pasta(nome_pasta):
    """
    Extrai o mês e ano do nome da pasta. 
    Lida com '04-2008' ou 'DEZEMBRO - 2008'
    """
    # Tentativa 1: Buscar diretamente no nome original da pasta (preserva hifens e pontuações)
    match_orig = re.search(r"(?<!\d)(\d{2})\s*[-_/\.]\s*(\d{4})(?!\d)", nome_pasta)
    if match_orig:
        return match_orig.group(1), match_orig.group(2)

    # Se não achou na primeira, aplica a limpeza e busca novamente
    nome_norm = normalizar_texto(nome_pasta.upper())
    
    # Tentativa 2: Formato numérico MM YYYY (como o hífen virou espaço na limpeza, buscamos por espaço)
    match_num = re.search(r"(?<!\d)(\d{2})\s+(\d{4})(?!\d)", nome_norm)
    if match_num:
        return match_num.group(1), match_num.group(2)
        
    # Tentativa 3: Formato texto MES YYYY (ex: DEZEMBRO 2008)
    meses = {
        'JANEIRO': '01', 'FEVEREIRO': '02', 'MARCO': '03', 'ABRIL': '04', 
        'MAIO': '05', 'JUNHO': '06', 'JULHO': '07', 'AGOSTO': '08', 
        'SETEMBRO': '09', 'OUTUBRO': '10', 'NOVEMBRO': '11', 'DEZEMBRO': '12'
    }
    
    for mes_texto, mes_num in meses.items():
        if mes_texto in nome_norm:
            match_ano = re.search(r"(?<!\d)(\d{4})(?!\d)", nome_norm)
            if match_ano:
                return mes_num, match_ano.group(1)
                
    return None, None

# --- FUNÇÕES AUXILIARES DE INTERFACE ---

def selecionar_arquivos_excel():
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    caminhos = filedialog.askopenfilenames(
        title="1. Selecione as Planilhas FGTS (Pode selecionar várias)",
        filetypes=[("Planilhas", "*.xlsx *.xls *.ods"), ("Todos", "*.*")]
    )
    root.destroy()
    return caminhos

def selecionar_pasta_raiz():
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    pasta = filedialog.askdirectory(title="2. Selecione a PASTA RAIZ (que contém as subpastas de competência)")
    root.destroy()
    return pasta

def selecionar_pasta_salvamento():
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    pasta = filedialog.askdirectory(
        title="3. Selecione a PASTA DE DESTINO para salvar os relatórios"
    )
    root.destroy()
    return pasta

# normalizar_texto / levenshtein / calcular_similaridade vem de utils.py (compartilhado)
from utils import normalizar_texto, levenshtein, calcular_similaridade
# comparar_nomes: matcher de nomes por tokens (ver match_nomes.py e test_match_nomes.py)
# comparar_tokens + tokens: versao pre-calculada (mais rapida em lacos grandes)
from match_nomes import comparar_nomes, comparar_tokens, tokens as _mn_tokens

RE_VARIANTE = re.compile(r'(REC\s*\.?\s*\d+|\b115\b)', re.IGNORECASE)

def extrair_variante(nome_sem_ext):
    """Detecta sufixo como 'REC. 115', 'REC 115' ou '115' no nome do arquivo.
    Retorna (nome_stem, variante) onde variante e 'REC115' (sem pontuacao) ou None."""
    m = RE_VARIANTE.search(nome_sem_ext)
    if m:
        raw = m.group(1)
        tag = re.sub(r'[\s\.]', '', raw).upper()
        if tag == '115':
            tag = 'REC115'
        stem = nome_sem_ext[:m.start()].strip() + ' ' + nome_sem_ext[m.end():].strip()
        return stem.strip(), tag
    return nome_sem_ext, None

def verificar_abreviacao(nome_planilha, nome_pdf):
    conectores = ['DE', 'DA', 'DO', 'DOS', 'DAS']
    tokens_p = [t for t in nome_planilha.split() if t not in conectores]
    tokens_f = [t for t in nome_pdf.split() if not t.isdigit() and t not in conectores]

    if abs(len(tokens_p) - len(tokens_f)) > 2:
        return False

    idx_p = 0
    idx_f = 0
    match_count = 0

    while idx_p < len(tokens_p) and idx_f < len(tokens_f):
        t_p = tokens_p[idx_p]
        t_f = tokens_f[idx_f]

        if t_p == t_f:
            match_count += 1
            idx_p += 1; idx_f += 1
            continue

        if (len(t_p) <= 4 and t_f.startswith(t_p)) or (len(t_f) <= 4 and t_p.startswith(t_f)):
            match_count += 1
            idx_p += 1; idx_f += 1
            continue

        # Tokens longos/medios diferentes no mesmo indice com listas de mesmo tamanho
        if len(t_p) >= 4 and len(t_f) >= 4 and len(tokens_p) == len(tokens_f):
            if levenshtein(t_p, t_f) <= max(1, min(len(t_p), len(t_f)) // 3):
                match_count += 1; idx_p += 1; idx_f += 1; continue
            return False

        # Tokens longos diferentes: consome ambos (nao e nome omitido, e nome diferente)
        if len(t_p) >= 4 and len(t_f) >= 4:
            idx_p += 1; idx_f += 1
        # Pula token curto extra na lista maior
        elif len(tokens_p) > len(tokens_f):
            idx_p += 1
        elif len(tokens_f) > len(tokens_p):
            idx_f += 1
        else:
            idx_p += 1; idx_f += 1

    return match_count >= max(len(tokens_p), len(tokens_f)) - 1

def e_um_nome_valido(texto):
    if not isinstance(texto, str): return False
    texto = texto.strip().upper()
    
    palavra_limpa = texto.replace(' ', '').replace('.', '').replace('-', '').replace('_', '')
    ignorar = [
        "NOMES", "NOME", "TITULAR", "FUNCIONARIO", "PIS", "PROC", "DATA", "OBS",
        "TOTAL", "TOTAIS", "SUBTOTAL", "TOTALGERAL", "SITUACAO", "COMPETENCIA",
        "VALOR", "DEPOSITO"
    ]
    if palavra_limpa in ignorar:
        return False
        
    if texto.startswith('TOTAL') or texto.startswith('SUBTOTAL'):
        return False

    if re.match(r"^[A-ZÁÀÂÃÉÊÍÓÔÕÚÜÇ0-9 \.\-]+$", texto) and " " in texto and len(texto) > 5:
        return True
    return False

# --- FUNÇÕES PRINCIPAIS ---

def ler_planilha_fgts(caminho_arquivo):
    if not caminho_arquivo: return None
    try:
        # Define a engine correta com base na extensão (necessário para ler .ods corretamente)
        extensao = os.path.splitext(caminho_arquivo)[1].lower()
        engine_pd = "odf" if extensao == '.ods' else None
        
        df_raw = pd.read_excel(caminho_arquivo, sheet_name="FGTS EM ATRASO - PROCESSOS", header=None, engine=engine_pd)
    except Exception as e:
        print(f"\n[!] Erro ao ler a planilha {os.path.basename(caminho_arquivo)}: {e}")
        if extensao == '.ods':
            print(" -> DICA: Para ler arquivos .ods, certifique-se de que a biblioteca 'odfpy' está instalada.")
            print(" -> Rode no seu terminal/prompt: pip install odfpy")
        return None

    idx_proc, idx_pis, idx_nomes = 1, 2, 3
    IDX_COL_J = 9                     # coluna J (valor)
    IDX_VALORES_ANTES_J = range(4, 9) # colunas E..I (valores antes do J)
    dados = []
    excluidos_sem_valor = 0

    def _celula_vazia(v):
        # Vazio de verdade (branco/NaN). '0' e '-' contam como valor (nao exclui).
        return pd.isna(v) or str(v).strip().lower() in ('', 'nan', 'none')
    
    encontrou_primeiro_inicio = False
    gap_linhas_vazias = 0
    bloco_atual = "PADRAO" 
    
    LIMITE_GAP_TOTAL = 40
    LIMITE_MUDANCA_BLOCO = 5

    for i in range(len(df_raw)):
        linha = df_raw.iloc[i]
        val_nome = linha[idx_nomes]
        
        # Ignora linhas de total (ex: "T O T A I S", "TOTAL GERAL")
        val_col0 = str(linha[0]).strip().upper() if pd.notna(linha[0]) else ''
        if re.search(r'T\s*O\s*T\s*(A\s*I\s*S|A\s*L)', val_col0):
            continue

        eh_nome = e_um_nome_valido(val_nome)

        if eh_nome:
            encontrou_primeiro_inicio = True
            gap_linhas_vazias = 0

            # Filtra pessoas EXCLUIDAS: sem valor na coluna J E sem valor nas
            # colunas antes dela (E..I). Assim quem so esqueceu o J, mas tem os
            # outros valores, NAO e excluido por engano.
            ncols = len(linha)
            j_vazio = IDX_COL_J >= ncols or _celula_vazia(linha[IDX_COL_J])
            antes_vazio = all(c >= ncols or _celula_vazia(linha[c]) for c in IDX_VALORES_ANTES_J)
            if j_vazio and antes_vazio:
                excluidos_sem_valor += 1
                continue

            val_pis = linha[idx_pis]
            # Se a coluna do PIS estiver vazia, procura nas colunas anteriores (deslocamento de célula)
            if pd.isna(val_pis) or str(val_pis).strip() in ('', '-', '0'):
                for col_idx in range(idx_nomes):
                    c_val = linha[col_idx]
                    if pd.notna(c_val):
                        val_str = str(c_val).strip()
                        if val_str.endswith('.0'): val_str = val_str[:-2]
                        # Verifica se parece um PIS (11 dígitos)
                        digitos = re.sub(r'\D', '', val_str)
                        if len(digitos) == 11:
                            val_pis = val_str
                            break

            dados.append({
                "PROC.": linha[idx_proc],
                "PIS": val_pis,
                "NOMES": val_nome,
                "TIPO_LISTA": bloco_atual 
            })
        else:
            if encontrou_primeiro_inicio:
                gap_linhas_vazias += 1
                if gap_linhas_vazias > LIMITE_MUDANCA_BLOCO:
                    bloco_atual = "115"
                if gap_linhas_vazias > LIMITE_GAP_TOTAL:
                    break

    if excluidos_sem_valor:
        print(f"  {excluidos_sem_valor} pessoa(s) ignorada(s) (sem valor na coluna J e nas colunas anteriores).")
    df_final = pd.DataFrame(dados)
    df_final.attrs['excluidos_sem_valor'] = excluidos_sem_valor
    return df_final

def verificar_pdfs(df, pasta_pdfs):
    if not pasta_pdfs or not os.path.exists(pasta_pdfs): 
        return df

    # Filtra arquivos que NAO sao nomes de pessoa (ex: documentos do sistema)
    EXCLUIR_PADROES = [
        'ARQUIVO SEFIP', 'SEFIP', 'GRRF', 'DEPOSITADO',
        'COMPROVANTE', 'RECIBO', 'EXTRATO', 'FOLHA',
        'RELATORIO', 'GUIA', 'GPS', 'GFIP',
        'PROTOCOLO', 'COMPENSACAO', 'DECLARACAO',
        'ARQUIVO FGTS', 'ARQUIVO FGST', 'IMG_', 'Thumbs',
    ]

    arquivos_excluidos = []
    arquivos_brutos = []
    for f in os.listdir(pasta_pdfs):
        if not f.lower().endswith(EXT):
            continue
        nome_sem_ext = os.path.splitext(f)[0].upper()
        if any(p in nome_sem_ext for p in EXCLUIR_PADROES):
            arquivos_excluidos.append(f)
            continue
        arquivos_brutos.append(f)

    df.attrs['arquivos_excluidos'] = sorted(arquivos_excluidos)

    # Categoriza PDFs por variante (REC 115 ou normal)
    arquivos_disponiveis = []
    for f in arquivos_brutos:
        stem, var = extrair_variante(os.path.splitext(f)[0])
        _norm = normalizar_texto(stem)
        arquivos_disponiveis.append({
            "real": f,
            "norm": _norm,
            "tokens": _mn_tokens(_norm),   # pre-calculado (otimizacao)
            "variante": var,
        })

    def _filtrar_por_variante(lista, variante):
        return [a for a in lista if a["variante"] == variante]

    def _filtrar_sem_variante(lista):
        return [a for a in lista if not a["variante"]]

    def _buscar_em_grupo(nome_norm, grupo, usar_llm=False, niveis=("IGUAL", "ABREVIACAO", "DUVIDA")):
        """Busca nome_norm no grupo de PDFs usando o matcher por tokens
        (match_nomes.comparar_nomes). Retorna (status, arquivo_obj) ou (None, None).

        `niveis` define QUAIS veredictos sao aceitos nesta chamada, para permitir
        casamento em passadas por confianca (ver o loop principal):
          IGUAL      -> ENCONTRADO
          ABREVIACAO -> ENCONTRADO COM ABREVIAÇÃO
          DUVIDA     -> POSSÍVEL ERRO NOMINAL (ou LLM decide, se ligado)
          DIFERENTE  -> ignorado
        Desempate entre candidatos do mesmo veredicto: maior score, depois
        tamanho de nome mais proximo.
        """
        if not grupo:
            return None, None

        nome_tokens = _mn_tokens(nome_norm)   # pre-calcula uma vez por chamada
        melhor_igual = melhor_igual_chave = None
        melhor_abrev = melhor_abrev_chave = None
        duvidas = []

        for arq in grupo:
            veredicto, score = comparar_tokens(nome_tokens, arq["tokens"])
            chave = (score, -abs(len(arq["norm"]) - len(nome_norm)))
            if veredicto == "IGUAL":
                if melhor_igual is None or chave > melhor_igual_chave:
                    melhor_igual, melhor_igual_chave = arq, chave
            elif veredicto == "ABREVIACAO":
                if melhor_abrev is None or chave > melhor_abrev_chave:
                    melhor_abrev, melhor_abrev_chave = arq, chave
            elif veredicto == "DUVIDA":
                duvidas.append((score, arq))

        if "IGUAL" in niveis and melhor_igual is not None:
            return "ENCONTRADO", melhor_igual
        if "ABREVIACAO" in niveis and melhor_abrev is not None:
            return "ENCONTRADO COM ABREVIAÇÃO", melhor_abrev

        if "DUVIDA" in niveis and duvidas:
            duvidas.sort(key=lambda x: x[0], reverse=True)
            if usar_llm and llm_matcher:
                # O LLM arbitra os casos duvidosos, do mais provavel ao menos.
                indeciso = None
                for score, arq in duvidas:
                    resp = llm_matcher.verificar_com_llm(nome_norm, arq["norm"])
                    if resp is True:
                        print(f'    [LLM] {nome_norm} x {arq["norm"]} -> SIM (duvida)')
                        return "ENCONTRADO VIA LLM", arq
                    elif resp is False:
                        print(f'    [LLM] {nome_norm} x {arq["norm"]} -> NAO')
                        continue
                    elif indeciso is None:
                        indeciso = arq  # LLM em duvida -> guarda para revisao humana
                if indeciso is not None:
                    return "POSSÍVEL ERRO NOMINAL", indeciso
                return None, None
            # Sem LLM: a duvida mais provavel vai para revisao humana.
            return "POSSÍVEL ERRO NOMINAL", duvidas[0][1]

        return None, None

    status_lista = [""] * len(df)
    arquivo_encontrado_lista = [""] * len(df)
    total_pessoas = len(df)

    # Indice por TUPLA DE TOKENS (ignora numeros/conectores) -> match exato O(1).
    # Chavear pela tupla (nao pela string com o numero) trata "ARNALDO ... SOUZA 1"
    # e "...2" como o mesmo nome "ARNALDO ... SOUZA" (o numero e so desambiguador).
    from collections import defaultdict
    tok_map = defaultdict(list)
    for a in arquivos_disponiveis:
        tok_map[tuple(a["tokens"])].append(a)
    disp_ids = {id(a) for a in arquivos_disponiveis}   # disponibilidade O(1)

    nomes_lista = list(df['NOMES'])
    tipos_lista = list(df['TIPO_LISTA']) if 'TIPO_LISTA' in df.columns else ['PADRAO'] * total_pessoas

    # Match em PASSADAS POR CONFIANCA: exatos (IGUAL) de todas as linhas primeiro,
    # depois abreviacoes, e so por ultimo as duvidas — sempre com o que sobrou.
    for niveis in [("IGUAL",), ("ABREVIACAO",), ("DUVIDA",)]:
        for ordem in ["115", "PADRAO"]:   # 115 tem prioridade nos arquivos
            indices_ordem = [idx for idx in range(total_pessoas)
                             if not status_lista[idx] and tipos_lista[idx] == ordem]
            if not indices_ordem:
                continue

            for idx in tqdm(indices_ordem, desc=f"{niveis[0]} {ordem}", unit="pess", leave=False):
                tipo_lista = tipos_lista[idx]
                nome_norm = normalizar_texto(nomes_lista[idx])
                match_obj = None
                status = None
                var_diff = False

                if niveis == ("IGUAL",):
                    # Match exato O(1) pela tupla de tokens (trata numero no fim).
                    cands = [a for a in tok_map.get(tuple(_mn_tokens(nome_norm)), [])
                             if id(a) in disp_ids]
                    if cands:
                        quer_115 = (tipo_lista == "115")
                        # prefere PDF do grupo certo; senao aceita outro (VAR DIFF)
                        mesmo_grupo = [a for a in cands if (a["variante"] == "REC115") == quer_115]
                        match_obj = (mesmo_grupo or cands)[0]
                        var_diff = (match_obj["variante"] == "REC115") != quer_115
                        status = "ENCONTRADO"
                else:
                    # ABREVIACAO / DUVIDA via _buscar_em_grupo (grupo por variante).
                    if tipo_lista == "115":
                        grupo_primario = _filtrar_por_variante(arquivos_disponiveis, "REC115")
                        grupo_secundario = _filtrar_sem_variante(arquivos_disponiveis)
                    else:
                        grupo_primario = _filtrar_sem_variante(arquivos_disponiveis)
                        grupo_secundario = _filtrar_por_variante(arquivos_disponiveis, "REC115")

                    status, match_obj = _buscar_em_grupo(nome_norm, grupo_primario, USAR_LLM, niveis)
                    if not match_obj:
                        status, match_obj = _buscar_em_grupo(nome_norm, grupo_secundario, USAR_LLM, niveis)
                        if match_obj:
                            var_diff = True

                if match_obj:
                    disp_ids.discard(id(match_obj))
                    arquivos_disponiveis.remove(match_obj)
                    if tipo_lista == "115" and "ENCONTRADO" in status:
                        status = "ENCONTRADO COMO 115"
                    if var_diff and "ENCONTRADO" in status:
                        status = status + " (VAR DIFF)"
                    status_lista[idx] = status
                    arquivo_encontrado_lista[idx] = os.path.splitext(match_obj["real"])[0]
    # end for niveis

    # Linhas sem match em nenhuma passada -> nao encontrado
    for idx in range(len(status_lista)):
        if not status_lista[idx]:
            status_lista[idx] = "NÃO ENCONTRADO NO .PDF"
            arquivo_encontrado_lista[idx] = ""

    # Pos-processamento: linha duplicada na planilha (mesmo PIS de alguem que
    # JA recebeu PDF) nao e "faltando PDF" -> marca como DUPLICADO. Ex.: a mesma
    # pessoa foi digitada 2x na planilha, mas existe 1 unico PDF dela.
    def _pis_digitos(v):
        return ''.join(c for c in str(v) if c.isdigit()).lstrip('0')
    pis_lista = list(df['PIS']) if 'PIS' in df.columns else [''] * len(df)
    pis_com_pdf = set()
    for i in range(len(status_lista)):
        if arquivo_encontrado_lista[i]:
            p = _pis_digitos(pis_lista[i]) if i < len(pis_lista) else ''
            if p:
                pis_com_pdf.add(p)
    for i in range(len(status_lista)):
        if status_lista[i] == "NÃO ENCONTRADO NO .PDF":
            p = _pis_digitos(pis_lista[i]) if i < len(pis_lista) else ''
            if p and p in pis_com_pdf:
                status_lista[i] = "DUPLICADO NA PLANILHA"

    df['Status PDF'] = status_lista
    df['Nome do Arquivo Encontrado'] = arquivo_encontrado_lista

    if 'TIPO_LISTA' in df.columns:
        df = df.drop(columns=['TIPO_LISTA'])

    # 4. SOBRAS (PDFs que sobraram)
    sobras = []
    for arq in arquivos_disponiveis:
        var_tag = f" [{arq['variante']}]" if arq['variante'] else ""
        sobras.append({
            "PROC.": "", "PIS": "", "NOMES": "",
            "Status PDF": f"PDF NA PASTA, MAS NÃO NA PLANILHA{var_tag}",
            "Nome do Arquivo Encontrado": os.path.splitext(arq["real"])[0]
        })

    if sobras:
        df = pd.concat([df, pd.DataFrame(sobras)], ignore_index=True)

    return df

def organizar_pdfs_por_resultado(df, pasta_pdfs, pasta_destino):
    """
    Cria duas pastas (115 e 660) DENTRO da pasta_destino e copia os PDFs
    conforme o resultado da conferencia.
    - PASTA 115: PDFs com status 'ENCONTRADO COMO 115'
    - PASTA 660: todos os demais PDFs encontrados
    """
    if not pasta_pdfs or not os.path.exists(pasta_pdfs):
        print(" -> Pasta de PDFs nao encontrada.")
        return
    if not pasta_destino or not os.path.exists(pasta_destino):
        print(" -> Pasta de destino nao encontrada.")
        return

    nome_pasta = os.path.basename(pasta_pdfs)
    
    # Cria os nomes das novas pastas
    # Ex: "01-SEFIP FGTS - COMPETENCIA 01-2007"
    #   -> "01-SEFIP FGTS 115 - COMPETENCIA 01-2007"
    #   -> "01-SEFIP FGTS 660 - COMPETENCIA 01-2007"
    padrao_subst = r'( - COMPETENCIA)'
    pasta_115 = re.sub(padrao_subst, r' 115\1', nome_pasta, flags=re.IGNORECASE)
    pasta_660 = re.sub(padrao_subst, r' 660\1', nome_pasta, flags=re.IGNORECASE)

    if pasta_115 == nome_pasta:
        pasta_115 = nome_pasta + ' 115'
        pasta_660 = nome_pasta + ' 660'

    # Cria dentro da pasta de destino (onde o .xlsx foi salvo)
    caminho_115 = os.path.join(pasta_destino, pasta_115)
    caminho_660 = os.path.join(pasta_destino, pasta_660)

    count_115 = 0
    count_660 = 0

    for _, row in df.iterrows():
        status = str(row.get('Status PDF', ''))
        nome_arquivo = str(row.get('Nome do Arquivo Encontrado', ''))

        if not nome_arquivo or nome_arquivo == '':
            continue
        if status == 'PDF NA PASTA, MAS NÃO NA PLANILHA':
            continue

        # Reconstrói o caminho completo do PDF na pasta ORIGINAL
        caminho_pdf = os.path.join(pasta_pdfs, nome_arquivo + '.pdf')
        if not os.path.exists(caminho_pdf):
            caminho_pdf = os.path.join(pasta_pdfs, nome_arquivo + '.tif')
        if not os.path.exists(caminho_pdf):
            caminho_pdf = os.path.join(pasta_pdfs, nome_arquivo + '.tiff')
        if not os.path.exists(caminho_pdf):
            # Tenta com a extensao real da pasta
            for f in os.listdir(pasta_pdfs):
                nome_base = os.path.splitext(f)[0]
                if nome_base == nome_arquivo:
                    caminho_pdf = os.path.join(pasta_pdfs, f)
                    break
            else:
                continue

        if 'ENCONTRADO COMO 115' in status:
            os.makedirs(caminho_115, exist_ok=True)
            shutil.copy2(caminho_pdf, os.path.join(caminho_115, os.path.basename(caminho_pdf)))
            count_115 += 1
        else:
            os.makedirs(caminho_660, exist_ok=True)
            shutil.copy2(caminho_pdf, os.path.join(caminho_660, os.path.basename(caminho_pdf)))
            count_660 += 1

    print(f" -> PDFs organizados:")
    if count_115 > 0:
        print(f"    {count_115} na pasta 115: {caminho_115}")
    if count_660 > 0:
        print(f"    {count_660} na pasta 660: {caminho_660}")

def aplicar_formatacao_condicional(caminho_arquivo):
    try:
        wb = load_workbook(caminho_arquivo)
        ws = wb.active
        
        borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fill_cinza = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        fill_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fonte_vermelha = Font(color="FF0000", bold=True)
        fonte_verde = Font(color="006400", bold=True)
        fonte_azul = Font(color="0000FF", bold=True)
        fonte_preta = Font(color="000000", bold=True)
        fonte_roxo = Font(color="800080", bold=True) 

        col_idx = None
        for cell in ws[1]:
            if cell.value == "Status PDF":
                col_idx = cell.column
                break
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = borda_fina
                if cell.row == 1:
                    cell.fill = fill_cinza
                    cell.font = Font(bold=True)
                elif col_idx and cell.column == col_idx:
                    val = str(cell.value)
                    if "NÃO ENCONTRADO" in val:
                        cell.font = fonte_vermelha
                    elif val.startswith("ENCONTRADO COMO 115"):
                        cell.font = fonte_roxo 
                    elif "ENCONTRADO" in val: 
                        cell.font = fonte_verde
                    elif val == "POSSÍVEL ERRO NOMINAL":
                        cell.fill = fill_amarelo
                        cell.font = fonte_preta
                    elif val.startswith("PDF NA PASTA, MAS NÃO NA PLANILHA"):
                        cell.font = fonte_azul
        
        for column_cells in ws.columns:
            length = 0
            col_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > length: length = len(str(cell.value))
                except: pass
            if length > 50: length = 50
            ws.column_dimensions[col_letter].width = length + 2

        wb.save(caminho_arquivo)
    except Exception as e:
        print(f"Erro na formatação: {e}")

def salvar_com_resumo(df, caminho_saida, total_pdfs=None):
    """Salva o DataFrame em Excel com uma aba Resumo antes dos dados."""
    def _status_base(s):
        for prefixo in [
            "NÃO ENCONTRADO NO .PDF",
            "POSSÍVEL ERRO NOMINAL",
            "ENCONTRADO VIA LLM",
            "ENCONTRADO COMO 115",
            "ENCONTRADO COM ABREVIAÇÃO",
            "ENCONTRADO",
        ]:
            if str(s).startswith(prefixo):
                return prefixo
        if str(s).startswith("PDF NA PASTA"):
            return "PDF NA PASTA, MAS NÃO NA PLANILHA"
        if str(s).startswith("PASTA NÃO ENCONTRADA"):
            return "PASTA NÃO ENCONTRADA"
        if str(s).startswith("COMPETÊNCIA DESCONHECIDA"):
            return "COMPETÊNCIA DESCONHECIDA"
        return str(s)

    prioridade = {
        "PASTA NÃO ENCONTRADA": 1,
        "COMPETÊNCIA DESCONHECIDA": 1,
        "NÃO ENCONTRADO NO .PDF": 2,
        "POSSÍVEL ERRO NOMINAL": 3,
        "PDF NA PASTA, MAS NÃO NA PLANILHA": 4,
        "ENCONTRADO COMO 115": 5,
        "ENCONTRADO COM ABREVIAÇÃO": 6,
        "ENCONTRADO": 7,
    }
    df = df.copy()
    df['Prioridade'] = df['Status PDF'].apply(lambda s: prioridade.get(_status_base(s), 8))
    df = df.sort_values(by=['Prioridade', 'NOMES', 'Nome do Arquivo Encontrado'])
    df = df.drop(columns=['Prioridade'])

    # Salva dados principais primeiro (usa openpyxl como engine padrao)
    with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Dados', index=False)

    # Abre para adicionar aba Resumo e formatar
    wb = load_workbook(caminho_saida)
    print(f"     Aba 'Dados' criada com {len(df)} registros")

    # Calcula estatisticas (agrupa por status base para capturar variantes)
    df['Status Base'] = df['Status PDF'].apply(_status_base)
    totais = df['Status Base'].value_counts()
    
    qtd_pdf_na_planilha = totais.get('PDF NA PASTA, MAS NÃO NA PLANILHA', 0)
    qtd_encontrado = totais.get('ENCONTRADO', 0)
    qtd_encontrado_115 = totais.get('ENCONTRADO COMO 115', 0)
    qtd_abreviacao = totais.get('ENCONTRADO COM ABREVIAÇÃO', 0)
    qtd_erro_nominal = totais.get('POSSÍVEL ERRO NOMINAL', 0)
    qtd_nao_encontrado = totais.get('NÃO ENCONTRADO NO .PDF', 0)
    qtd_pasta_nao_encontrada = totais.get('PASTA NÃO ENCONTRADA', 0)
    qtd_comp_desconhecida = totais.get('COMPETÊNCIA DESCONHECIDA', 0)
    df = df.drop(columns=['Status Base'])

    total_planilha = qtd_encontrado + qtd_encontrado_115 + qtd_abreviacao + qtd_erro_nominal + qtd_nao_encontrado + qtd_pasta_nao_encontrada + qtd_comp_desconhecida
    total_encontrados = qtd_encontrado + qtd_encontrado_115 + qtd_abreviacao

    # Cria aba Resumo
    ws = wb.create_sheet('Resumo', 0)  # Insere como primeira aba

    # Estilos
    titulo_font = Font(bold=True, size=14)
    cab_font = Font(bold=True, size=11, color='FFFFFF')
    cab_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    bdr = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    fill_ok = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_err = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    fill_115 = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    font_ok = Font(color='006100', bold=True)
    font_err = Font(color='9C0006', bold=True)
    font_115 = Font(color='9C5700', bold=True)

    ws.merge_cells('A1:B1')
    ws['A1'] = 'RESUMO DA CONFERÊNCIA'
    ws['A1'].font = titulo_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Linhas do resumo
    resumo_linhas = [
        ('', ''),
        ('Planilha', f'{os.path.basename(caminho_saida)}'),
        ('', ''),
        ('PESSOAS NA PLANILHA', total_planilha),
        ('ARQUIVOS PDF NA PASTA', total_pdfs if total_pdfs else ''),
        ('', ''),
    ]
    if qtd_encontrado > 0:
        resumo_linhas.append(('[OK] Encontrados', qtd_encontrado))
    if qtd_encontrado_115 > 0:
        resumo_linhas.append(('[OK] Encontrados como 115', qtd_encontrado_115))
    if qtd_abreviacao > 0:
        resumo_linhas.append(('[~] Encontrados com abreviação', qtd_abreviacao))
    if qtd_erro_nominal > 0:
        resumo_linhas.append(('[?] Possível erro nominal', qtd_erro_nominal))
    if qtd_nao_encontrado > 0:
        resumo_linhas.append(('[X] Não encontrados no PDF', qtd_nao_encontrado))
    if qtd_pasta_nao_encontrada > 0:
        resumo_linhas.append(('[!] Pasta não encontrada', qtd_pasta_nao_encontrada))
    if qtd_comp_desconhecida > 0:
        resumo_linhas.append(('[!] Competência desconhecida', qtd_comp_desconhecida))
    if qtd_pdf_na_planilha > 0:
        resumo_linhas.append(('[PDF] PDF na pasta mas não na planilha', qtd_pdf_na_planilha))

    resumo_linhas.extend([
        ('', ''),
        ('TOTAL ENCONTRADOS (OK)', total_encontrados),
        ('TOTAL NÃO ENCONTRADOS', qtd_nao_encontrado + qtd_pasta_nao_encontrada + qtd_comp_desconhecida),
        ('', ''),
    ])

    for i, (desc, val) in enumerate(resumo_linhas, 3):
        ws.cell(row=i, column=1, value=desc).border = bdr
        c = ws.cell(row=i, column=2, value=val)
        c.border = bdr
        c.alignment = Alignment(horizontal='center')

        # Colore linhas importantes
        if 'Encontrados' in desc and '115' not in desc and 'abrevia' not in desc:
            c.fill = fill_ok
            c.font = font_ok
        elif '115' in desc:
            c.fill = fill_115
            c.font = font_115
        elif 'Não encontrados' in desc or 'não encontrados' in desc:
            c.fill = fill_err
            c.font = font_err

    # --- Lista de nao encontrados e PDFs extras ---
    linha_atual = 3 + len(resumo_linhas) + 1

    nao_encontrados = df[df['Status PDF'].apply(lambda s: str(s).startswith('NÃO ENCONTRADO NO .PDF'))]
    pdfs_extras = df[df['Status PDF'].apply(lambda s: str(s).startswith('PDF NA PASTA, MAS NÃO NA PLANILHA'))]

    if not nao_encontrados.empty:
        linha_atual += 1
        ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=2)
        ws.cell(row=linha_atual, column=1, value='NÃO ENCONTRADOS NO PDF:').font = Font(bold=True, color='FF0000')
        linha_atual += 1
        for _, r in nao_encontrados.iterrows():
            ws.cell(row=linha_atual, column=1, value=r['NOMES'])
            ws.cell(row=linha_atual, column=2, value=r.get('PROC.', ''))
            linha_atual += 1

    if not pdfs_extras.empty:
        linha_atual += 1
        ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=2)
        ws.cell(row=linha_atual, column=1, value='PDFS EXTRAS (na pasta mas nao na planilha):').font = Font(bold=True, color='0000FF')
        linha_atual += 1
        for _, r in pdfs_extras.iterrows():
            ws.cell(row=linha_atual, column=1, value=r['Nome do Arquivo Encontrado'])
            linha_atual += 1

    # --- Lista de arquivos ignorados (documentos do sistema, SEFIP, etc.) ---
    arquivos_ignorados = df.attrs.get('arquivos_excluidos', [])
    if arquivos_ignorados:
        linha_atual += 1
        ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=2)
        ws.cell(row=linha_atual, column=1, value='ARQUIVOS IGNORADOS (documentos do sistema):').font = Font(bold=True, color='666666')
        linha_atual += 1
        for nome in arquivos_ignorados:
            ws.cell(row=linha_atual, column=1, value=nome).font = Font(italic=True, color='888888')
            linha_atual += 1

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 25

    # Formata a aba Dados
    ws_dados = wb['Dados']
    bdr2 = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    fill_cinza = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    fonte_vermelha = Font(color='FF0000', bold=True)
    fonte_verde = Font(color='006400', bold=True)
    fonte_azul = Font(color='0000FF', bold=True)
    fonte_preta = Font(color='000000', bold=True)
    fonte_roxo = Font(color='800080', bold=True)
    fill_amarelo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    col_idx = None
    for cell in ws_dados[1]:
        if cell.value and 'Status PDF' in str(cell.value):
            col_idx = cell.column
            break

    for row in ws_dados.iter_rows():
        for cell in row:
            cell.border = bdr2
            if cell.row == 1:
                cell.fill = fill_cinza
                cell.font = Font(bold=True)
            elif col_idx and cell.column == col_idx:
                val = str(cell.value)
                if 'NÃO ENCONTRADO' in val:
                    cell.font = fonte_vermelha
                elif val.startswith('ENCONTRADO COMO 115'):
                    cell.font = fonte_roxo
                elif 'ENCONTRADO' in val:
                    cell.font = fonte_verde
                elif val == 'POSSÍVEL ERRO NOMINAL':
                    cell.fill = fill_amarelo
                    cell.font = fonte_preta
                elif val.startswith('PDF NA PASTA, MAS NÃO NA PLANILHA'):
                    cell.font = fonte_azul

    # Ajusta largura das colunas
    for col_cells in ws_dados.columns:
        length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                if len(str(cell.value)) > length:
                    length = len(str(cell.value))
            except:
                pass
        if length > 50:
            length = 50
        ws_dados.column_dimensions[col_letter].width = length + 2

    wb.save(caminho_saida)


# --- EXECUÇÃO EM MASSA ---
if __name__ == "__main__":
    print('=' * 60)
    print('  CONFERENCIA FGTS x PDFs')
    print('=' * 60)
    print()
    print('Modos:')
    print('  1 — Simular (sem alterar nada)')
    print('  2 — Executar')
    print('  3 — Executar com LLM (Gemini)')
    while True:
        modo = input('\nOpcao (1, 2 ou 3): ').strip()
        if modo in ('1', '2', '3'):
            break

    if modo == '3':
        USAR_LLM = True
        print('\nModo LLM ativado.')
        if llm_matcher:
            resp = llm_matcher.verificar_com_llm('TESTE', 'TESTE')
            if resp is None:
                print('  ATENCAO: LLM nao disponivel. Configure a chave em config_llm.py')
                input('\nPressione Enter para continuar mesmo assim...')
            else:
                print('  LLM OK')
        else:
            print('  ATENCAO: llm_matcher.py nao encontrado na pasta.')
            input('\nPressione Enter para continuar mesmo assim...')
        resp = llm_matcher.verificar_com_llm('TESTE', 'TESTE')
        if resp is None:
            print('  ATENCAO: LLM nao disponivel. Configure a chave em config_llm.py')
            input('\nPressione Enter para continuar mesmo assim...')
        else:
            print('  LLM OK')

    arquivos_excel = selecionar_arquivos_excel()
    
    if arquivos_excel:
        print(f"{len(arquivos_excel)} planilhas selecionadas.")
        pasta_raiz = selecionar_pasta_raiz()
        
        if pasta_raiz:
            pasta_salvamento = selecionar_pasta_salvamento()
            
            if pasta_salvamento:
                print(f"\nMapeando pastas dentro de: {pasta_raiz} ...")
                
                # Mapeia todas as subpastas pela competência (MM_YYYY)
                mapa_pastas = {}
                for item in os.listdir(pasta_raiz):
                    caminho_completo = os.path.join(pasta_raiz, item)
                    if os.path.isdir(caminho_completo):
                        mes, ano = extrair_competencia_pasta(item)
                        if mes and ano:
                            mapa_pastas[f"{mes}_{ano}"] = caminho_completo

                # Itera sobre as planilhas do Excel para gerar arquivos individuais
                for arquivo in arquivos_excel:
                    nome_base_excel = os.path.basename(arquivo)
                    print(f"\n--- Processando: {nome_base_excel} ---")
                    
                    mes_ex, ano_ex = extrair_competencia_excel(nome_base_excel)
                    df_dados = ler_planilha_fgts(arquivo)
                    
                    if df_dados is None or df_dados.empty:
                        print(f" -> Nenhuma informação válida encontrada na planilha.")
                        continue

                    if not mes_ex or not ano_ex:
                        print(f" -> AVISO: Mês/ano não identificado. Ignorando PDFs.")
                        df_dados["Status PDF"] = "COMPETÊNCIA DESCONHECIDA"
                        df_dados["Nome do Arquivo Encontrado"] = ""
                        nome_saida = f"Relatorio_{os.path.splitext(nome_base_excel)[0]} - sem pasta.xlsx"
                    else:
                        chave_comp = f"{mes_ex}_{ano_ex}"
                        
                        if chave_comp in mapa_pastas:
                            nome_saida = f"{mes_ex}_{ano_ex}.xlsx"
                            pasta_alvo = mapa_pastas[chave_comp]
                            print(f" -> Cruzando com a pasta: {os.path.basename(pasta_alvo)}")
                            df_dados = verificar_pdfs(df_dados, pasta_alvo)
                        else:
                            nome_saida = f"{mes_ex}_{ano_ex} - sem pasta.xlsx"
                            print(f" -> AVISO: Pasta de PDFs de {mes_ex}/{ano_ex} não encontrada.")
                            df_dados["Status PDF"] = "PASTA NÃO ENCONTRADA"
                            df_dados["Nome do Arquivo Encontrado"] = ""

                    # Salvamento com resumo + formatacao
                    caminho_saida = os.path.join(pasta_salvamento, nome_saida)
                    try:
                        # Conta quantos PDFs existem na pasta (para o resumo)
                        total_pdfs_pasta = 0
                        if 'chave_comp' in locals() and chave_comp in mapa_pastas:
                            pasta_alvo = mapa_pastas[chave_comp]
                            EXCLUIR_PADROES = [
                            'ARQUIVO SEFIP', 'SEFIP', 'GRRF', 'DEPOSITADO',
                            'COMPROVANTE', 'RECIBO', 'EXTRATO', 'FOLHA',
                            'RELATORIO', 'GUIA', 'GPS', 'GFIP', 'SEFIP',
                            'PROTOCOLO', 'COMPENSACAO', 'DECLARACAO',
                        ]
                        for f in os.listdir(pasta_alvo):
                            if not f.lower().endswith(('.xls', '.xlsx', '.ods')):
                                nome_sem_ext = os.path.splitext(f)[0].upper()
                                if any(p in nome_sem_ext for p in EXCLUIR_PADROES):
                                    continue
                                total_pdfs_pasta += 1

                        salvar_com_resumo(df_dados, caminho_saida, total_pdfs_pasta)
                        print(f" -> Salvo com sucesso: {nome_saida}")

                        # Mostra detalhes no console
                        nao_encontrados = df_dados[df_dados['Status PDF'] == 'NÃO ENCONTRADO NO .PDF']
                        if not nao_encontrados.empty:
                            print(f" -> PESSOAS NÃO ENCONTRADAS NO PDF ({len(nao_encontrados)}):")
                            for _, r in nao_encontrados.iterrows():
                                print(f"    {r['NOMES']}")

                        erros_nominais = df_dados[df_dados['Status PDF'] == 'POSSÍVEL ERRO NOMINAL']
                        if not erros_nominais.empty:
                            print(f" -> POSSÍVEIS ERROS NOMINAIS ({len(erros_nominais)}):")
                            for _, r in erros_nominais.iterrows():
                                print(f"    Planilha: {r['NOMES']}  ->  PDF: {r['Nome do Arquivo Encontrado']}")

                        pdfs_extras = df_dados[df_dados['Status PDF'] == 'PDF NA PASTA, MAS NÃO NA PLANILHA']
                        if not pdfs_extras.empty:
                            print(f" -> PDFS EXTRAS (sem nome na planilha) ({len(pdfs_extras)}):")
                            for _, r in pdfs_extras.iterrows():
                                print(f"    {r['Nome do Arquivo Encontrado']}")

                        # Mostra arquivos ignorados (sistema)
                        arquivos_ign = df_dados.attrs.get('arquivos_excluidos', [])
                        if arquivos_ign:
                            print(f" -> ARQUIVOS IGNORADOS (documentos do sistema):")
                            for nome in arquivos_ign:
                                print(f"    {nome}")

                    except Exception as e:
                        print(f" -> Erro ao salvar {nome_saida}: {e}")

                    # Organizar PDFs em pastas 115/660 (dentro da pasta de destino)
                    if 'chave_comp' in locals() and chave_comp in mapa_pastas:
                        pasta_alvo = mapa_pastas[chave_comp]
                        print(f" -> Organizando PDFs em pastas 115/660...")
                        organizar_pdfs_por_resultado(df_dados, pasta_alvo, pasta_salvamento)
                
                print(f"\nPROCESSO CONCLUÍDO! Verifique a pasta: {pasta_salvamento}")
                os.startfile(pasta_salvamento) # Abre a pasta onde salvou tudo no final
                
            else:
                print("Pasta de destino não selecionada. Processo cancelado.")
        else:
            print("Pasta Raiz não selecionada.")
    else:
        print("Nenhuma planilha selecionada.")